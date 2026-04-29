import io
import os
import sys
from datetime import datetime
import streamlit as st
import pandas as pd
from PIL import Image
import psycopg2
import numpy as np
import plotly.graph_objects as go
import plotly.express as px

def _fmt_num(v) -> str:
    """Formatea un número con máximo 2 decimales y sin ceros finales.

    Ejemplos: 5.0 → '5', 5.50 → '5.5', 5.123 → '5.12', -3.0 → '-3'
    """
    try:
        f = float(v)
    except (TypeError, ValueError):
        return str(v)
    if f == int(f):
        return str(int(f))
    return f"{f:.2f}".rstrip("0")


# =========================================================
# INTERNACIONALIZACIÓN — solo textos fijos de interfaz
# =========================================================
_TRANSLATIONS: dict[str, dict[str, str]] = {
    "es": {
        "app_title":          "Motor Estratégico de Capacidad Productiva",
        "app_subtitle":       "Planificación por líneas y simulación de mix",
        "plant_select":       "Seleccionar planta",
        "plant_new":          "Nueva planta",
        "plant_add":          "Añadir planta",
        "nav_label":          "Pantalla:",
        "nav_global":         "🌐 Global",
        "nav_planning":       "📊 Planificación",
        "nav_config":         "⚙️ Configuración (Power User)",
        "nav_results":        "📈 Resultados",
        "nav_mix":            "🧭 Capacidad según mix",
        "dark_mode_label":    "🌙  Modo oscuro",
        "lang_label":         "🌍 Idioma",
        "params_header":      "Parámetros de planificación",
        "param_hours_week":   "Horas por semana",
        "param_shifts":       "Turnos",
        "param_availability": "Disponibilidad",
        "param_efficiency":   "Eficiencia",
        "param_days_year":    "Días abiertos al año",
        "param_days_week":    "Días abiertos por semana",
        "caption_hours_eff":  "Horas efectivas planta:",
        "caption_weeks_equiv":"Semanas equivalentes:",
        "unit_week":          "h/semana",
        "unit_year":          "sem/año",
        "btn_save_params":    "Guardar parámetros de esta planta",
        "msg_params_saved":   "Parámetros guardados para esta planta",
        "tab_global_header":  "🌐 Visión Global Multiplanta",
        "tab_plan_header":    "Selección de modelo por línea",
        "plan_col_model":     "Modelo / Variante de prueba",
        "plan_col_demand":    "Demanda (UDS/SEM)",
        "tab_cfg_header":     "Configuración (power user)",
        "tab_cfg_caption":    "Aquí se mantienen modelos, tiempos, estaciones y compatibilidades. Usuario normal NO debería tocar esto.",
        "btn_save_models":    "💾 Guardar modelos",
        "msg_models_saved":   "Modelos guardados",
        "btn_save_times":     "💾 Guardar tiempos",
        "msg_times_saved":    "Tiempos guardados",
        "btn_save_stations":  "💾 Guardar estaciones / operarios",
        "msg_stations_saved": "Guardado",
        "btn_save_compat":    "💾 Guardar compatibilidades",
        "msg_compat_saved":   "Compatibilidades guardadas",
        "btn_save_benches":   "💾 Guardar bancos",
        "msg_benches_saved":  "Configuración de bancos guardada",
        "btn_save_da":        "💾 Guardar asignación D&A",
        "msg_da_saved":       "Asignación D&A guardada",
        "tab_res_header":     "Resultados de capacidad",
        "res_hours_eff":      "Horas efectivas planta:",
        "res_no_lines":       "Sin líneas planificadas. Selecciona modelos y demanda en Planificación.",
        "res_detail_header":  "## 🔎 Detalle fino por línea y subproceso",
        "res_detail_caption": "Desglose real por subproceso. El cuello de botella es el subproceso con menor capacidad.",
        "tab_mix_header":     "Capacidad según mix",
        # Sidebar mensajes
        "plant_added":            "Planta añadida",
        "plant_name_empty":       "Escribe un nombre de planta",
        # Global
        "global_info":            "Esta vista muestra información agregada de **TODAS las plantas** simultáneamente, independiente del selector de planta del sidebar.",
        "global_scenario_header": "### 📊 Selector de Escenario",
        "global_capacity_header": "### 📈 Resumen Global de Capacidad",
        "global_avail_header":    "### ⚡ Capacidad vs Disponibilidad",
        "global_avail_caption":   "Introduce la disponibilidad anual (horas) para cada planta.",
        "global_model_header":    "### 🔧 Capacidad por Modelo",
        "global_model_select":    "Seleccionar modelo:",
        "global_model_info":      "Selecciona un modelo específico para ver su capacidad en todas las plantas.",
        "global_no_models":       "No hay modelos activos en ninguna planta.",
        "global_mods_header":     "### 🔧 Modificaciones Necesarias",
        "global_mods_caption":    "Registro de mejoras planificadas",
        "global_mod_expander":    "➕ Añadir modificación",
        "global_mod_name":        "Nombre de la modificación",
        "global_mod_plant":       "Planta",
        "global_mod_hours":       "Horas estimadas",
        "global_mod_btn":         "Añadir",
        "global_milestones_header": "### 📊 Resumen de Hitos",
        "global_metric_count":    "Nº Hitos",
        "global_metric_hours":    "Horas Totales",
        "global_distrib_header":  "### 📊 Distribución de Capacidad por Planta",
        "global_usage_header":    "### 📋 Uso de Líneas por Planta",
        # Configuración
        "cfg_filter_model":       "Filtrar por modelo",
        "cfg_filter_process":     "Filtrar por proceso",
        "cfg_showing_rows":       "Mostrando {shown} de {total} filas. Las filas ocultas se conservan al guardar.",
        "cfg_benches_by_type":    "### Bancos disponibles por tipo",
        # Resultados
        "res_no_results_yet":     "No hay resultados aún. Selecciona modelos/demanda en Planificación.",
        "res_export_btn":         "⬇️ Exportar a Excel",
        "res_compare_header":     "### 🔀 Comparativa entre escenarios",
        "res_compare_select":     "Escenario de comparación",
        "res_compare_none":       "(Sin comparación)",
        "res_compare_export_btn": "⬇️ Exportar comparativa a Excel",
        "res_compare_no_data":    "No se pudo cargar el escenario de comparación.",
        "res_chart_header":       "## 📊 Representación gráfica de Demanda vs Capacidad",
        "res_no_data":            "No hay datos suficientes (revisa estaciones o tiempos).",
        "res_panel_n_deficit":    "Líneas con déficit",
        "res_panel_max_sat":      "Saturación máxima",
        "res_panel_n_critical":   "Líneas ≥ 90 % saturación",
        "res_panel_bottleneck":   "Cuello principal",
        "res_sorted_note":        "Tabla ordenada por criticidad: déficit primero, luego alta saturación.",
        "res_params_expander":    "⚙ Parámetros por línea",
        "res_fte_info":            "📋 El plan actual exige **{total_fte} personas equivalentes** bajo las condiciones actuales de planta.\nLa mayor carga se concentra en **{top_line}** con **{top_fte} personas eq.** ({top_pct} % del plan).\n\n*Personas eq. = HH proceso/sem ÷ (horas/sem × disponibilidad × eficiencia) de cada línea. Los turnos no entran: una persona no trabaja dos turnos a la vez. No representa plantilla asignada ni FTE confirmados.*",
        "res_save_line_btn":       "💾 Guardar parámetros por línea",
        "res_save_line_ok":        "Parámetros por línea guardados.",
        "res_save_line_err":       "Error al guardar. Comprueba la conexión.",
        "res_proc_no_procs":       "Sin procesos para esta línea.",
        "res_proc_global_shifts":  "Turnos globales de línea:",
        "res_proc_save_ok":        "Guardado.",
        "res_proc_save_err":       "Error al guardar.",
        "res_metric_cap_total":    "Cap. total (uds/sem)",
        "res_metric_dem_total":    "Dem. total (uds/sem)",
        "res_metric_deficit":      "Déficit (uds/sem)",
        "res_metric_sat_max":      "Sat. máxima (%)",
        "res_metric_crit_lines":   "Líneas críticas",
        # Simulación anual
        "nav_sim_annual":          "📅 Simulación anual",
        "sim_tab_header":          "Simulación anual — Capacidad vs Demanda",
        "sim_no_scenario":         "Sin escenario activo. Ve a Planificación, activa un escenario y asigna modelos antes de usar esta simulación.",
        "sim_no_lines":            "El escenario activo no tiene líneas planificadas. Ve a Planificación y asigna modelos a las líneas.",
        "sim_lines_planned":       "líneas planificadas",
        "sim_cap_base_label":      "Capacidad base provisional (hipótesis inicial a validar)",
        "sim_cap_per_line":        "h/sem por línea (global)",
        "sim_n_lines":             "Líneas planificadas",
        "sim_cap_total_base":      "Cap. base planta (h/sem)",
        "sim_cap_base_note":       "⚠ Hipótesis inicial: hours_eff global × nº líneas planificadas. No aplica overrides individuales por línea. Validar con negocio antes de continuar al cálculo semanal.",
        # Mix
        "mix_info":               "La planta produce **horas configurables**.\nLa capacidad no es un valor fijo, sino un **rango estructural** determinado por el mix posible de modelos en cada línea.\nAquí se muestran los valores **Máximo / Promedio / Mínimo** por planta y por línea, en unidades y en horas (semana y año).",
        "mix_level1":             "### Nivel 1 — Global planta (rango estructural)",
        "mix_level2":             "### Nivel 2 — Por línea (rango estructural)",
        "mix_level3":             "## Nivel 3 — Simulador de ocupación estructural por modelo",
        "mix_sim_caption":        "Este simulador **NO** cambia la planificación real. Solo sirve para explorar, en términos de **horas estructurales (h/sem)**, cuánto 'peso' podría llegar a ocupar cada modelo dentro del **techo estructural** de la planta.",
        "mix_no_combos":          "No hay combinaciones válidas para calcular el rango. Revisa compatibilidades, estaciones y/o tiempos.",
        "mix_no_valid_models":    "No hay modelos válidos (capacidad estructural > 0) para construir el simulador.",
        "mix_plant_agg":          "### Agregado planta",
        "mix_per_model":          "### Por modelo",
        "mix_slider_label":       "Ocupación simulada (% de planta)",
        "mix_slider_help":        "Este % es una ocupación estructural teórica (no planificación real).",
        # Configuración — títulos limpios (sin sufijos de fichero)
        "cfg_models_header":      "## Gestión de modelos",
        "cfg_times_header":       "## Tiempos por modelo y proceso",
        "cfg_times_info":         "**Machine time** = tiempo automático fijo no reducible (test automático, horno, robot, ciclo máquina). No depende del nº de operarios.\n\n**Labor time** = horas-hombre secuenciales necesarias por unidad (preparación, conexión, montaje manual, supervisión, retirada).\n\nLa capacidad se calcula mediante:\n\n`cycle_time_real = max(machine_time, labor_time / operarios)`\n\n`capacity = (horas_efectivas × estaciones) / cycle_time_real`\n\nEn procesos manuales puros, machine_time puede ser 0.",
        "cfg_stations_header":    "## Configuración de estaciones y operarios",
        "cfg_compat_header":      "## Compatibilidad modelo ↔ línea",
        "cfg_compat_expand_all":  "Desplegar todas",
        "cfg_compat_collapse_all":"Plegar todas",
        "cfg_line_label":         "Línea",
        "cfg_benches_section":    "## Bancos de prueba",
        "cfg_benches_caption":    "Configuración de bancos disponibles por planta y asignación de tipo de prueba a cada valor D&A. **Esta asignación es una simplificación operativa de esta fase**: dentro de una misma familia puede haber equipos con prueba LV y equipos con prueba MV. Solo aplica a los valores D&A: SL, SD, LL, LD, XD, XL.",
        "cfg_da_header":          "### Asignación valor D&A → tipo de banco",
        "cfg_da_caption":         "Usa los valores exactos de D&A tal como existen hoy en el motor (SL, SD, LL, LD, XD, XL). En esta fase no se baja al nivel de modelo real. La asignación es una simplificación operativa: una misma familia puede tener equipos LV y equipos MV que esta fase no distingue todavía.",
        "cfg_col_bench_type":     "Tipo de banco",
        "cfg_col_bench_qty":      "Cantidad de bancos",
        "cfg_col_da_value":       "Valor D&A en el motor",
        "cfg_col_da_variant":     "Variante de prueba",
        "cfg_col_bench_apply":    "Banco aplicable",
        "cfg_filter_da":          "Filtrar por valor D&A",
        "cfg_filter_da_all":      "(Todos)",
        "cfg_filter_status":      "Estado",
        "cfg_status_all":         "Todos",
        "cfg_status_active":      "Activos",
        "cfg_status_inactive":    "Inactivos",
        "cfg_filter_nave":        "Nave",
        "cfg_filter_proc_label":  "Proceso",
        "cfg_compat_export":      "⬇ Exportar compatibilidades",
        # Planificación
        "plan_no_models":         "sin modelos compatibles activos (revisa compatibilidades/modelos).",
        # Resultados — bancos
        "res_bench_header":       "## 🏭 Análisis de bancos de prueba",
        "res_bench_info":         "**La información de bancos de prueba es una referencia adicional para detectar posibles limitaciones. Todavía no sustituye el cálculo oficial de capacidad de la línea.**\n\nPara cada línea D&A puedes especificar la **Variante de prueba** (LV o MV) en la pestaña Planificación. Si no se especifica, se aplica la regla general configurada para esa familia. La asignación por familia sigue siendo una simplificación operativa cuando la variante no se informa.",
        "res_bench_warning":      "Configuración de bancos no disponible para esta planta. Rellena las tablas en ⚙️ Configuración → Bancos de prueba.",
        "res_no_da_lines":        "No hay líneas D&A activas con bancos configurados para esta planta.",
        "res_bench_agg_header":   "### Resumen agregado por tipo de banco",
        "res_bench_agg_text":     "La tabla anterior analiza cada línea D&A **por separado**, mostrando si los bancos disponibles podrían limitarla de forma individual. Sin embargo, en la práctica varias líneas pueden compartir el mismo conjunto de bancos, y lo que determina si hay un problema real es el consumo **conjunto**.\n\nEste resumen agrega todas las líneas que usan el mismo tipo de banco y compara las horas totales disponibles con las horas totales que la demanda actual requeriría:\n\n- **Horas disponibles/sem**: horas efectivas de la planta × número de bancos de ese tipo.\n- **Horas demandadas/sem**: suma de (demanda de cada línea × tiempo de prueba PARA de esa línea). Indica cuántas horas de banco se necesitarían para cubrir la demanda planificada.\n- **Demanda total (UDS/SEM)**: suma de unidades semanales de todas las líneas que usan ese banco.\n- **Capacidad máxima (UDS/SEM)**: capacidad equivalente del conjunto de bancos con la mezcla actual de familias.\n- **Saturación (%)**: nivel de ocupación del conjunto de bancos.\n\nEste bloque es informativo y **no modifica la lógica oficial del motor** (capacidad, saturación, déficit ni cuello de botella).",
        "res_bench_sim_header":   "### 🔭 Simulación estratégica de escenarios de banco",
        "res_bench_sim_warning":  "**Este bloque muestra proyecciones hipotéticas a partir de la configuración actual de bancos.** Los valores de +1 y +2 bancos son estimaciones lineales orientativas. No modifican la capacidad oficial del motor ni ningún otro resultado de la herramienta.",
        "res_bench_q1":           "#### ¿Qué pasaría si añadiera más bancos?",
        "res_bench_q2":           "#### ¿Cuántos bancos hacen falta para cubrir la demanda actual?",
        "res_no_da_data":         "No hay líneas D&A con datos suficientes para calcular el resumen agregado. Comprueba que el proceso PARA existe en tiempos y estaciones.",
        # Global — adicionales
        "global_sim_turns":       "⚠️ Simulación: todas las plantas con **{n} turno(s)**",
        "global_mod_detail_plant": "**Planta:**",
        "global_mod_detail_hours": "**Horas estimadas:**",
        "global_mod_delete":      "🗑️ Eliminar",
        "global_model_selected":  "**Modelo seleccionado:**",
        "global_escenario_label": "Escenario de capacidad:",
        "global_turnos_label":    "Turnos (simulación global):",
        "global_caption_esc":     "Escenario:",
        "global_caption_turns":   "Turnos:",
        "global_disp_prefix":     "Disp.",
        "global_chart_capvsdisp": "Capacidad vs Disponibilidad por Planta (h/año)",
        "global_chart_legend_cap":  "Capacidad",
        "global_chart_legend_disp": "Disponibilidad",
        "global_chart_distrib":   "Distribución de Capacidad ({esc}) - h/año",
        # Mix — panel derecho
        "mix_ceiling_label":      "**Techo estructural planta:**",
        "mix_max_hours_year":     "**Horas máximas/año (estructura):**",
        "mix_excess_label":       "Exceso estructural:",
        "mix_occupancy_label":    "**Ocupación agregada:**",
        "mix_hours_week_label":   "**Horas/sem (equivalentes):**",
        "mix_hours_year_label":   "**Horas/año (equivalentes):**",
        "mix_pot_caption":        "🎚️ Uso del potenciómetro:",
        "mix_equiv_label":        "Equivalente aproximado",
        # Scenario persistence
        "plan_save_btn":          "Guardar escenario actual",
        "plan_save_name_label":   "Nombre del escenario",
        "plan_save_name_default": "Escenario guardado",
        "plan_save_ok":           "Escenario guardado correctamente.",
        "plan_save_no_db":        "Sin conexión a base de datos — escenario no persistido.",
        "plan_scenarios_header":  "Escenarios guardados",
        "plan_load_btn":          "Cargar",
        "plan_activate_btn":      "Usar por defecto",
        "plan_load_ok":           "Escenario cargado.",
        "plan_activate_ok":       "Establecido como escenario por defecto.",
        "plan_no_scenarios":      "No hay escenarios guardados para esta planta.",
        "plan_active_marker":     "✓ activo",
        "plan_delete_btn":        "Borrar",
        "plan_delete_ok":         "Escenario eliminado.",
        "plan_delete_blocked":    "No se puede borrar el escenario activo. Activa otro primero.",
        "plan_duplicate_btn":     "Duplicar",
        "plan_save_changes_btn":  "Guardar cambios",
        "plan_save_new_btn":      "Guardar como nuevo",
        "plan_save_changes_no_sel": "Selecciona un escenario antes de guardar cambios.",
        "plan_save_changes_ok":   "Cambios guardados en el escenario.",
    },
    "en": {
        "app_title":          "Strategic Production Capacity Engine",
        "app_subtitle":       "Line planning and mix simulation",
        "plant_select":       "Select plant",
        "plant_new":          "New plant",
        "plant_add":          "Add plant",
        "nav_label":          "Screen:",
        "nav_global":         "🌐 Global",
        "nav_planning":       "📊 Planning",
        "nav_config":         "⚙️ Configuration (Power User)",
        "nav_results":        "📈 Results",
        "nav_mix":            "🧭 Capacity by mix",
        "dark_mode_label":    "🌙  Dark mode",
        "lang_label":         "🌍 Language",
        "params_header":      "Planning parameters",
        "param_hours_week":   "Hours per week",
        "param_shifts":       "Shifts",
        "param_availability": "Availability",
        "param_efficiency":   "Efficiency",
        "param_days_year":    "Open days per year",
        "param_days_week":    "Open days per week",
        "caption_hours_eff":  "Effective plant hours:",
        "caption_weeks_equiv":"Equivalent weeks:",
        "unit_week":          "h/week",
        "unit_year":          "wk/year",
        "btn_save_params":    "Save plant parameters",
        "msg_params_saved":   "Parameters saved for this plant",
        "tab_global_header":  "🌐 Multi-Plant Global View",
        "tab_plan_header":    "Model selection by line",
        "plan_col_model":     "Model / Test variant",
        "plan_col_demand":    "Demand (UNITS/WEEK)",
        "tab_cfg_header":     "Configuration (power user)",
        "tab_cfg_caption":    "Models, times, stations and compatibilities are managed here. Regular users should NOT modify this.",
        "btn_save_models":    "💾 Save models",
        "msg_models_saved":   "Models saved",
        "btn_save_times":     "💾 Save times",
        "msg_times_saved":    "Times saved",
        "btn_save_stations":  "💾 Save stations / operators",
        "msg_stations_saved": "Saved",
        "btn_save_compat":    "💾 Save compatibilities",
        "msg_compat_saved":   "Compatibilities saved",
        "btn_save_benches":   "💾 Save banks",
        "msg_benches_saved":  "Bank configuration saved",
        "btn_save_da":        "💾 Save D&A assignment",
        "msg_da_saved":       "D&A assignment saved",
        "tab_res_header":     "Capacity results",
        "res_hours_eff":      "Effective plant hours:",
        "res_no_lines":       "No lines planned. Select models and demand in Planning.",
        "res_detail_header":  "## 🔎 Fine detail by line and subprocess",
        "res_detail_caption": "Actual breakdown by subprocess. The bottleneck is the subprocess with the lowest capacity.",
        "tab_mix_header":     "Capacity by mix",
        # Sidebar messages
        "plant_added":            "Plant added",
        "plant_name_empty":       "Enter a plant name",
        # Global
        "global_info":            "This view shows aggregated information from **ALL plants** simultaneously, independent of the sidebar plant selector.",
        "global_scenario_header": "### 📊 Scenario Selector",
        "global_capacity_header": "### 📈 Global Capacity Summary",
        "global_avail_header":    "### ⚡ Capacity vs Availability",
        "global_avail_caption":   "Enter the annual availability (hours) for each plant.",
        "global_model_header":    "### 🔧 Capacity by Model",
        "global_model_select":    "Select model:",
        "global_model_info":      "Select a specific model to see its capacity across all plants.",
        "global_no_models":       "No active models in any plant.",
        "global_mods_header":     "### 🔧 Pending Modifications",
        "global_mods_caption":    "Record of planned improvements",
        "global_mod_expander":    "➕ Add modification",
        "global_mod_name":        "Modification name",
        "global_mod_plant":       "Plant",
        "global_mod_hours":       "Estimated hours",
        "global_mod_btn":         "Add",
        "global_milestones_header": "### 📊 Milestones Summary",
        "global_metric_count":    "No. Milestones",
        "global_metric_hours":    "Total Hours",
        "global_distrib_header":  "### 📊 Capacity Distribution by Plant",
        "global_usage_header":    "### 📋 Line Usage by Plant",
        # Configuration
        "cfg_filter_model":       "Filter by model",
        "cfg_filter_process":     "Filter by process",
        "cfg_showing_rows":       "Showing {shown} of {total} rows. Hidden rows are preserved when saving.",
        "cfg_benches_by_type":    "### Available benches by type",
        # Results
        "res_no_results_yet":     "No results yet. Select models/demand in Planning.",
        "res_export_btn":         "⬇️ Export to Excel",
        "res_compare_header":     "### 🔀 Scenario comparison",
        "res_compare_select":     "Comparison scenario",
        "res_compare_none":       "(No comparison)",
        "res_compare_export_btn": "⬇️ Export comparison to Excel",
        "res_compare_no_data":    "Could not load comparison scenario.",
        "res_chart_header":       "## 📊 Demand vs Capacity chart",
        "res_no_data":            "Insufficient data (check stations or times).",
        "res_panel_n_deficit":    "Lines with deficit",
        "res_panel_max_sat":      "Maximum saturation",
        "res_panel_n_critical":   "Lines ≥ 90 % saturation",
        "res_panel_bottleneck":   "Main bottleneck",
        "res_sorted_note":        "Table sorted by criticality: deficit first, then high saturation.",
        "res_params_expander":    "⚙ Parameters per line",
        "res_fte_info":            "📋 The current plan requires **{total_fte} equivalent people** under current plant conditions.\nThe highest load is concentrated in **{top_line}** with **{top_fte} FTE equiv.** ({top_pct} % of plan).\n\n*FTE equiv. = process HH/week ÷ (hours/week × availability × efficiency) per line. Shifts are excluded: one person cannot work two shifts simultaneously. Does not represent assigned headcount or confirmed FTE.*",
        "res_save_line_btn":       "💾 Save line parameters",
        "res_save_line_ok":        "Line parameters saved.",
        "res_save_line_err":       "Error saving. Check connection.",
        "res_proc_no_procs":       "No processes for this line.",
        "res_proc_global_shifts":  "Line global shifts:",
        "res_proc_save_ok":        "Saved.",
        "res_proc_save_err":       "Error saving.",
        "res_metric_cap_total":    "Total cap. (uds/week)",
        "res_metric_dem_total":    "Total dem. (uds/week)",
        "res_metric_deficit":      "Deficit (uds/week)",
        "res_metric_sat_max":      "Max. saturation (%)",
        "res_metric_crit_lines":   "Critical lines",
        # Annual simulation
        "nav_sim_annual":          "📅 Annual simulation",
        "sim_tab_header":          "Annual simulation — Capacity vs Demand",
        "sim_no_scenario":         "No active scenario. Go to Planning, activate a scenario and assign models before using this simulation.",
        "sim_no_lines":            "The active scenario has no planned lines. Go to Planning and assign models to lines.",
        "sim_lines_planned":       "planned lines",
        "sim_cap_base_label":      "Provisional base capacity (initial hypothesis to validate)",
        "sim_cap_per_line":        "h/week per line (global)",
        "sim_n_lines":             "Planned lines",
        "sim_cap_total_base":      "Plant base capacity (h/week)",
        "sim_cap_base_note":       "⚠ Initial hypothesis: global hours_eff × number of planned lines. Does not apply individual line overrides. Validate with business before proceeding to weekly calculation.",
        # Mix
        "mix_info":               "The plant produces **configurable hours**.\nCapacity is not a fixed value, but a **structural range** determined by the possible model mix on each line.\nThis shows **Maximum / Average / Minimum** values per plant and line, in units and hours (week and year).",
        "mix_level1":             "### Level 1 — Plant global (structural range)",
        "mix_level2":             "### Level 2 — By line (structural range)",
        "mix_level3":             "## Level 3 — Structural occupancy simulator by model",
        "mix_sim_caption":        "This simulator does **NOT** change the actual planning. It only helps explore, in terms of **structural hours (h/week)**, how much 'weight' each model could occupy within the plant's **structural ceiling**.",
        "mix_no_combos":          "No valid combinations to calculate the range. Check compatibilities, stations and/or times.",
        "mix_no_valid_models":    "No valid models (structural capacity > 0) to build the simulator.",
        "mix_plant_agg":          "### Plant aggregate",
        "mix_per_model":          "### By model",
        "mix_slider_label":       "Simulated occupancy (% of plant)",
        "mix_slider_help":        "This % is a theoretical structural occupancy (not real planning).",
        # Configuration — clean titles (no file name suffixes)
        "cfg_models_header":      "## Model management",
        "cfg_times_header":       "## Times by model and process",
        "cfg_times_info":         "**Machine time** = fixed automatic non-reducible time (automatic test, oven, robot, machine cycle). Does not depend on the number of operators.\n\n**Labor time** = sequential man-hours required per unit (preparation, wiring, manual assembly, supervision, retrieval).\n\nCapacity is calculated as:\n\n`cycle_time_real = max(machine_time, labor_time / operators)`\n\n`capacity = (effective_hours × stations) / cycle_time_real`\n\nFor purely manual processes, machine_time may be 0.",
        "cfg_stations_header":    "## Station and operator configuration",
        "cfg_compat_header":      "## Model ↔ line compatibility",
        "cfg_compat_expand_all":  "Expand all",
        "cfg_compat_collapse_all":"Collapse all",
        "cfg_line_label":         "Line",
        "cfg_benches_section":    "## Test benches",
        "cfg_benches_caption":    "Configuration of available benches per plant and assignment of test type to each D&A value. **This assignment is an operational simplification for this phase**: within the same family there may be equipment with LV test and equipment with MV test. Applies only to D&A values: SL, SD, LL, LD, XD, XL.",
        "cfg_da_header":          "### D&A value → bench type assignment",
        "cfg_da_caption":         "Use the exact D&A values as they exist today in the engine (SL, SD, LL, LD, XD, XL). At this phase, we do not go down to the real model level. The assignment is an operational simplification: the same family may have LV and MV equipment that this phase does not yet distinguish.",
        "cfg_col_bench_type":     "Bench type",
        "cfg_col_bench_qty":      "Number of benches",
        "cfg_col_da_value":       "D&A value in engine",
        "cfg_col_da_variant":     "Test variant",
        "cfg_col_bench_apply":    "Applicable bench",
        "cfg_filter_da":          "Filter by D&A value",
        "cfg_filter_da_all":      "(All)",
        "cfg_filter_status":      "Status",
        "cfg_status_all":         "All",
        "cfg_status_active":      "Active",
        "cfg_status_inactive":    "Inactive",
        "cfg_filter_nave":        "Bay",
        "cfg_filter_proc_label":  "Process",
        "cfg_compat_export":      "⬇ Export compatibilities",
        # Planning
        "plan_no_models":         "no active compatible models (check compatibilities/models).",
        # Results — benches
        "res_bench_header":       "## 🏭 Test bench analysis",
        "res_bench_info":         "**Test bench information is an additional reference to detect potential limitations. It does not yet replace the official line capacity calculation.**\n\nFor each D&A line you can specify the **Test variant** (LV or MV) in the Planning tab. If not specified, the general rule configured for that family applies. Family-level assignment remains an operational simplification when the variant is not informed.",
        "res_bench_warning":      "Bench configuration not available for this plant. Fill in the tables in ⚙️ Configuration → Test benches.",
        "res_no_da_lines":        "No active D&A lines with benches configured for this plant.",
        "res_bench_agg_header":   "### Aggregated summary by bench type",
        "res_bench_agg_text":     "The table above analyses each D&A line **separately**, showing whether the available benches could individually limit it. However, in practice several lines may share the same set of benches, and what determines whether there is a real problem is the **combined** consumption.\n\nThis summary aggregates all lines using the same bench type and compares total available hours with total hours that current demand would require:\n\n- **Available hours/week**: effective plant hours × number of benches of that type.\n- **Demanded hours/week**: sum of (line demand × PARA test time for that line). Indicates how many bench hours would be needed to meet planned demand.\n- **Total demand (UDS/WEEK)**: total weekly units of all lines using that bench.\n- **Max capacity (UDS/WEEK)**: equivalent capacity of the bench set with the current family mix.\n- **Saturation (%)**: occupancy level of the bench set.\n\nThis block is informational and **does not modify the official engine logic** (capacity, saturation, deficit or bottleneck).",
        "res_bench_sim_header":   "### 🔭 Strategic bench scenario simulation",
        "res_bench_sim_warning":  "**This block shows hypothetical projections based on the current bench configuration.** Values for +1 and +2 benches are linear indicative estimates. They do not modify the official engine capacity or any other result of the tool.",
        "res_bench_q1":           "#### What would happen if more benches were added?",
        "res_bench_q2":           "#### How many benches are needed to cover current demand?",
        "res_no_da_data":         "No D&A lines with sufficient data to calculate the aggregated summary. Check that the PARA process exists in times and stations.",
        # Global — additional
        "global_sim_turns":       "⚠️ Simulation: all plants with **{n} shift(s)**",
        "global_mod_detail_plant": "**Plant:**",
        "global_mod_detail_hours": "**Estimated hours:**",
        "global_mod_delete":      "🗑️ Delete",
        "global_model_selected":  "**Selected model:**",
        "global_escenario_label": "Capacity scenario:",
        "global_turnos_label":    "Shifts (global simulation):",
        "global_caption_esc":     "Scenario:",
        "global_caption_turns":   "Shifts:",
        "global_disp_prefix":     "Avail.",
        "global_chart_capvsdisp": "Capacity vs Availability by Plant (h/year)",
        "global_chart_legend_cap":  "Capacity",
        "global_chart_legend_disp": "Availability",
        "global_chart_distrib":   "Capacity Distribution ({esc}) - h/year",
        # Mix — right panel
        "mix_ceiling_label":      "**Plant structural ceiling:**",
        "mix_max_hours_year":     "**Max hours/year (structure):**",
        "mix_excess_label":       "Structural excess:",
        "mix_occupancy_label":    "**Aggregate occupancy:**",
        "mix_hours_week_label":   "**Hours/week (equivalent):**",
        "mix_hours_year_label":   "**Hours/year (equivalent):**",
        "mix_pot_caption":        "🎚️ Dial usage:",
        "mix_equiv_label":        "Approximate equivalent",
        # Scenario persistence
        "plan_save_btn":          "Save current scenario",
        "plan_save_name_label":   "Scenario name",
        "plan_save_name_default": "Saved scenario",
        "plan_save_ok":           "Scenario saved successfully.",
        "plan_save_no_db":        "No database connection — scenario not persisted.",
        "plan_scenarios_header":  "Saved scenarios",
        "plan_load_btn":          "Load",
        "plan_activate_btn":      "Set as default",
        "plan_load_ok":           "Scenario loaded.",
        "plan_activate_ok":       "Set as default scenario.",
        "plan_no_scenarios":      "No saved scenarios for this plant.",
        "plan_active_marker":     "✓ active",
        "plan_delete_btn":        "Delete",
        "plan_delete_ok":         "Scenario deleted.",
        "plan_delete_blocked":    "Cannot delete the active scenario. Set another as default first.",
        "plan_duplicate_btn":     "Duplicate",
        "plan_save_changes_btn":  "Save changes",
        "plan_save_new_btn":      "Save as new",
        "plan_save_changes_no_sel": "Select a scenario before saving changes.",
        "plan_save_changes_ok":   "Changes saved to scenario.",
    },
    "eu": {
        "app_title":          "Ekoizpen-ahalmenaren Motor Estrategikoa",
        "app_subtitle":       "Lerro-planifikazioa eta mix simulazioa",
        "plant_select":       "Planta hautatu",
        "plant_new":          "Planta berria",
        "plant_add":          "Planta gehitu",
        "nav_label":          "Pantaila:",
        "nav_global":         "🌐 Global",
        "nav_planning":       "📊 Planifikazioa",
        "nav_config":         "⚙️ Konfigurazioa (Power User)",
        "nav_results":        "📈 Emaitzak",
        "nav_mix":            "🧭 Mix-aren araberako ahalmena",
        "dark_mode_label":    "🌙  Modu iluna",
        "lang_label":         "🌍 Hizkuntza",
        "params_header":      "Planifikazio-parametroak",
        "param_hours_week":   "Asteko orduak",
        "param_shifts":       "Txandak",
        "param_availability": "Erabilgarritasuna",
        "param_efficiency":   "Efizientzia",
        "param_days_year":    "Urteko egun irekiak",
        "param_days_week":    "Asteko egun irekiak",
        "caption_hours_eff":  "Plantaren ordu eraginkorrak:",
        "caption_weeks_equiv":"Aste baliokideak:",
        "unit_week":          "o/aste",
        "unit_year":          "aste/urte",
        "btn_save_params":    "Planta honen parametroak gorde",
        "msg_params_saved":   "Parametroak gordeta planta honentzat",
        "tab_global_header":  "🌐 Planta Anitzeko Ikuspegi Globala",
        "tab_plan_header":    "Eredu-hautaketa lerroaren arabera",
        "plan_col_model":     "Eredua / Proba-aldaera",
        "plan_col_demand":    "Eskaria (UNITATE/ASTE)",
        "tab_cfg_header":     "Konfigurazioa (power user)",
        "tab_cfg_caption":    "Ereduak, denborak, estazioak eta bateragarritasunak kudeatzen dira hemen. Erabiltzaile arruntak EZ luke ukitu behar.",
        "btn_save_models":    "💾 Ereduak gorde",
        "msg_models_saved":   "Ereduak gordeta",
        "btn_save_times":     "💾 Denborak gorde",
        "msg_times_saved":    "Denborak gordeta",
        "btn_save_stations":  "💾 Estazioak / operarioak gorde",
        "msg_stations_saved": "Gordeta",
        "btn_save_compat":    "💾 Bateragarritasunak gorde",
        "msg_compat_saved":   "Bateragarritasunak gordeta",
        "btn_save_benches":   "💾 Bankuak gorde",
        "msg_benches_saved":  "Bankuen konfigurazioa gordeta",
        "btn_save_da":        "💾 D&A esleipena gorde",
        "msg_da_saved":       "D&A esleipena gordeta",
        "tab_res_header":     "Ahalmen-emaitzak",
        "res_hours_eff":      "Plantaren ordu eraginkorrak:",
        "res_no_lines":       "Lerrorik ez planifikatua. Hautatu ereduak eta eskaria Planifikazioan.",
        "res_detail_header":  "## 🔎 Xehetasun zehatza lerro eta azpiprozesuka",
        "res_detail_caption": "Benetako banaketa azpiprozesuka. Botila-lepoa ahalmen txikiena duen azpiprozesua da.",
        "tab_mix_header":     "Mix-aren araberako ahalmena",
        # Sidebar mezuak
        "plant_added":            "Planta gehitu da",
        "plant_name_empty":       "Idatzi planta izen bat",
        # Global
        "global_info":            "Ikuspegi honek **PLANTA GUZTIEN** informazio agregatua erakusten du aldi berean, sidebar-eko planta-hautatzailearekin independentean.",
        "global_scenario_header": "### 📊 Eszenatoki-hautatzailea",
        "global_capacity_header": "### 📈 Ahalmen Global Laburpena",
        "global_avail_header":    "### ⚡ Ahalmena vs Erabilgarritasuna",
        "global_avail_caption":   "Sartu urteko erabilgarritasuna (orduak) planta bakoitzarentzat.",
        "global_model_header":    "### 🔧 Ahalmena ereduaren arabera",
        "global_model_select":    "Eredua hautatu:",
        "global_model_info":      "Hautatu eredu zehatz bat planta guztietan duen ahalmena ikusteko.",
        "global_no_models":       "Ez dago eredu aktiborik inongo plantan.",
        "global_mods_header":     "### 🔧 Beharrezko Aldaketak",
        "global_mods_caption":    "Planeatutako hobetze-erregistroa",
        "global_mod_expander":    "➕ Aldaketa gehitu",
        "global_mod_name":        "Aldaketaren izena",
        "global_mod_plant":       "Planta",
        "global_mod_hours":       "Ordu estimatuak",
        "global_mod_btn":         "Gehitu",
        "global_milestones_header": "### 📊 Hito-laburpena",
        "global_metric_count":    "Hito kop.",
        "global_metric_hours":    "Ordu Guztiak",
        "global_distrib_header":  "### 📊 Ahalmen-banaketa plantaren arabera",
        "global_usage_header":    "### 📋 Lerro-erabilera plantaren arabera",
        # Konfigurazioa
        "cfg_filter_model":       "Ereduaren arabera iragazi",
        "cfg_filter_process":     "Prozesuaren arabera iragazi",
        "cfg_showing_rows":       "{shown} errenkada erakusten, {total}etik. Ezkutuko errenkadak gordeta mantentzen dira.",
        "cfg_benches_by_type":    "### Eskuragarri dauden bankuak motaren arabera",
        # Emaitzak
        "res_no_results_yet":     "Oraindik emaitzarik ez. Hautatu ereduak/eskaria Planifikazioan.",
        "res_export_btn":         "⬇️ Excel-era esportatu",
        "res_compare_header":     "### 🔀 Eszenatokien konparaketa",
        "res_compare_select":     "Konparaketa eszenatokia",
        "res_compare_none":       "(Konparaketarik gabe)",
        "res_compare_export_btn": "⬇️ Konparaketa Excel-era esportatu",
        "res_compare_no_data":    "Ezin izan da konparaketa eszenatokia kargatu.",
        "res_chart_header":       "## 📊 Eskaera vs Ahalmena grafikoa",
        "res_no_data":            "Datu nahikorik ez (egiaztatu estazioak edo denborak).",
        "res_panel_n_deficit":    "Lerroak defizitarekin",
        "res_panel_max_sat":      "Saturazio maximoa",
        "res_panel_n_critical":   "Lerroak ≥ 90 % saturazio",
        "res_panel_bottleneck":   "Eztarri nagusia",
        "res_sorted_note":        "Taula kritikotasunaren arabera ordenatua: defizita lehenik, gero saturazio altua.",
        "res_params_expander":    "⚙ Parametroak lerro bakoitzeko",
        "res_fte_info":            "📋 Uneko planak **{total_fte} baliokide pertsona** eskatzen ditu egungo planta-baldintzekin.\nKarga handiena **{top_line}** lerroan kontzentratzen da: **{top_fte} pertsona bald.** (planaren {top_pct} %).\n\n*Pertsona bald. = prozesuko HH/aste ÷ (ordu/aste × erabilgarritasuna × efizientzia) lerro bakoitzeko. Txandak ez dira sartzen: pertsona batek ezin ditu bi txanda aldi berean egin. Ez da esleitutako langile-taldea ezta baieztatutako FTErik.*",
        "res_save_line_btn":       "💾 Lerroko parametroak gorde",
        "res_save_line_ok":        "Lerroko parametroak gordeta.",
        "res_save_line_err":       "Errorea gordetzean. Egiaztatu konexioa.",
        "res_proc_no_procs":       "Lerro honentzat prozesurik ez.",
        "res_proc_global_shifts":  "Lerroaren turno globalak:",
        "res_proc_save_ok":        "Gordeta.",
        "res_proc_save_err":       "Errorea gordetzean.",
        "res_metric_cap_total":    "Kap. guztira (uds/aste)",
        "res_metric_dem_total":    "Esk. guztira (uds/aste)",
        "res_metric_deficit":      "Defizita (uds/aste)",
        "res_metric_sat_max":      "Sat. maximoa (%)",
        "res_metric_crit_lines":   "Lerro kritikoak",
        # Urteko simulazioa
        "nav_sim_annual":          "📅 Urteko simulazioa",
        "sim_tab_header":          "Urteko simulazioa — Ahalmena vs Eskaria",
        "sim_no_scenario":         "Ez dago eszenatoki aktiborik. Joan Planifikazioera, aktibatu eszenatoki bat eta esleitu ereduak simulazio hau erabili aurretik.",
        "sim_no_lines":            "Eszenatoki aktiboak ez du planifikatutako lerrorik. Joan Planifikazioera eta esleitu ereduak lerroei.",
        "sim_lines_planned":       "planifikatutako lerro",
        "sim_cap_base_label":      "Oinarrizko ahalmen probisionala (hasierako hipotesia, balioztatze beharrekoa)",
        "sim_cap_per_line":        "o/aste lerro bakoitzeko (globala)",
        "sim_n_lines":             "Planifikatutako lerroak",
        "sim_cap_total_base":      "Plantaren oinarrizko ahalmena (o/aste)",
        "sim_cap_base_note":       "⚠ Hasierako hipotesia: ordu_efektibo globala × planifikatutako lerro kopurua. Ez du lerroko override indibidualik aplikatzen. Baliozta ezazu negozioari asteko kalkulura jo aurretik.",
        # Mix
        "mix_info":               "Plantak **konfiguragarriak diren orduak** ekoizten ditu.\nAhalmena ez da balio finko bat, baizik eta lerro bakoitzean posible diren ereduen mixak zehaztutako **tarte estrukturala**.\nHemen **Maximoa / Batez bestekoa / Minimoa** balioak erakusten dira planta eta lerroaren arabera, unitateetan eta orduetan (aste eta urte).",
        "mix_level1":             "### 1. maila — Planta globala (tarte estrukturala)",
        "mix_level2":             "### 2. maila — Lerroz lerro (tarte estrukturala)",
        "mix_level3":             "## 3. maila — Ereduaren araberako okupazio estrukturalaren simulagailua",
        "mix_sim_caption":        "Simulagailu honek ez du benetako planifikazioa **ALDATZEN**. **Ordu estrukturalei (o/aste)** dagokienez, eredu bakoitzak plantaren **teto estrukturalaren** barruan zenbat 'pisu' har lezakeen esploratzeko balio du.",
        "mix_no_combos":          "Ez dago tarte estrukturala kalkulatzeko konbinazio baliodunterik. Egiaztatu bateragarritasunak, estazioak eta/edo denborak.",
        "mix_no_valid_models":    "Ez dago eredu baliodunik (ahalmen estrukturala > 0) simulagailua eraikitzeko.",
        "mix_plant_agg":          "### Planta agregatua",
        "mix_per_model":          "### Ereduaren arabera",
        "mix_slider_label":       "Simulatutako okupazioa (plantaren %)",
        "mix_slider_help":        "% hau okupazio estrukturala teorikoa da (ez benetako planifikazioa).",
        # Konfigurazioa — izenburuak garbiak (fitxategi-atzizkiak gabe)
        "cfg_models_header":      "## Ereduen kudeaketa",
        "cfg_times_header":       "## Denborak ereduaren eta prozesuaren arabera",
        "cfg_times_info":         "**Machine time** = denbora automatiko finko ez-murrizgarria (test automatikoa, labea, robota, makina-zikloa). Ez dago operario kopuruaren mende.\n\n**Labor time** = unitateko beharrezko gizakideko ordu sekuentzialak (prestaketa, konexioa, eskuzko muntaketa, gainbegiraketa, erretiratzea).\n\nAhalmena honela kalkulatzen da:\n\n`cycle_time_real = max(machine_time, labor_time / operarioak)`\n\n`capacity = (ordu_eraginkorrak × estazioak) / cycle_time_real`\n\nEskuzko prozesu hutsetan, machine_time 0 izan daiteke.",
        "cfg_stations_header":    "## Estazioen eta operarioen konfigurazioa",
        "cfg_compat_header":      "## Eredua ↔ lerro bateragarritasuna",
        "cfg_compat_expand_all":  "Guztiak zabaldu",
        "cfg_compat_collapse_all":"Guztiak itxi",
        "cfg_line_label":         "Lerroa",
        "cfg_benches_section":    "## Proba-bankuak",
        "cfg_benches_caption":    "Plantako eskuragarri dauden bankuen konfigurazioa eta proba-motaren esleipena D&A balio bakoitzarentzat. **Esleipen hau fase honen sinplifikazio operatiboa da**: familia beraren barruan egon daitezke LV probarekin eta MV probarekin ekipoak. D&A balioentzat soilik aplikatzen da: SL, SD, LL, LD, XD, XL.",
        "cfg_da_header":          "### D&A balioa → banku-mota esleipena",
        "cfg_da_caption":         "Erabili D&A balio zehatzak gaur egun motor-ean dauden moduan (SL, SD, LL, LD, XD, XL). Fase honetan ez da eredu-mailaraino jaisten. Esleipena sinplifikazio operatiboa da: familia berdinak LV eta MV ekipoak izan ditzake fase honek oraindik bereizten ez dituena.",
        "cfg_col_bench_type":     "Banku-mota",
        "cfg_col_bench_qty":      "Banku kopurua",
        "cfg_col_da_value":       "D&A balioa motor-ean",
        "cfg_col_da_variant":     "Proba-aldaera",
        "cfg_col_bench_apply":    "Banku aplikagarria",
        "cfg_filter_da":          "D&A balioaren arabera iragazi",
        "cfg_filter_da_all":      "(Guztiak)",
        "cfg_filter_status":      "Egoera",
        "cfg_status_all":         "Guztiak",
        "cfg_status_active":      "Aktiboak",
        "cfg_status_inactive":    "Ez-aktiboak",
        "cfg_filter_nave":        "Nabe",
        "cfg_filter_proc_label":  "Prozesua",
        "cfg_compat_export":      "⬇ Esportatu bateragarritasunak",
        # Planifikazioa
        "plan_no_models":         "eredu bateragarri aktiborik gabe (egiaztatu bateragarritasunak/ereduak).",
        # Emaitzak — bankuak
        "res_bench_header":       "## 🏭 Proba-bankuen analisia",
        "res_bench_info":         "**Proba-bankuen informazioa erreferentzia gehigarri bat da muga posibleak hautemateko. Oraindik ez du lerroaren ahalmen-kalkulu ofiziala ordezkatzen.**\n\nD&A lerro bakoitzeko **Proba-aldaera** (LV edo MV) zehazteko aukera daukazu Planifikazio fitxan. Zehaztuta ez badago, familia horretarako konfiguratutako arau orokorra aplikatzen da. Familia-mailako esleipena sinplifikazio operatiboa izaten jarraitzen du aldaera informatzen ez denean.",
        "res_bench_warning":      "Bankuen konfigurazioa ez dago eskuragarri planta honentzat. Bete taulak ⚙️ Konfigurazioa → Proba-bankuak atalean.",
        "res_no_da_lines":        "Ez dago D&A lerro aktiborik planta honetarako konfiguratutako bankuekin.",
        "res_bench_agg_header":   "### Banku-motaren araberako laburpen agregatua",
        "res_bench_agg_text":     "Aurreko taulak D&A lerro bakoitza **bereizita** aztertzen du, eskuragarri dauden bankuek indibidualkiro mugatu dezaketen erakutsiz. Hala ere, praktikan hainbat lerrork banku multzo bera parteka dezakete, eta arazo erreala dagoen zehazteko kontsumo **bateratua** da garrantzitsua.\n\nLaburpen honek banku-mota bera erabiltzen duten lerro guztiak agregatzen ditu eta eskuragarri dauden ordu guztiak egungo eskaerak beharko lukeen ordu guztiekin alderatzen ditu:\n\n- **Eskuragarri dauden orduak/aste**: plantaren ordu eraginkorrak × banku-mota horretako banku kopurua.\n- **Eskatutako orduak/aste**: lerro bakoitzaren eskaera × PARA proba-denbora batura.\n- **Eskaera totala (UDS/ASTE)**: banku hori erabiltzen duten lerro guztien asteko unitate batura.\n- **Ahalmen maximoa (UDS/ASTE)**: egungo familia-mixarekin banku multzoak duen baliokide ahalmena.\n- **Saturazioa (%)**: banku multzoko okupazio-maila.\n\nBloke hau informatibo da eta **ez du motor-aren logika ofiziala aldatzen** (ahalmena, saturazioa, defizita ez eta botila-lepoa).",
        "res_bench_sim_header":   "### 🔭 Bankuen eszenatokien simulazio estrategikoa",
        "res_bench_sim_warning":  "**Bloke honek bankuen egungo konfiguraziotik abiatutako proiekzio hipotetikoak erakusten ditu.** +1 eta +2 bankuen balioak estimazio lineal orientagarriak dira. Ez dute motor-aren ahalmen ofiziala aldatzen.",
        "res_bench_q1":           "#### Zer gertatuko litzateke banku gehiago gehituz gero?",
        "res_bench_q2":           "#### Zenbat banku behar dira egungo eskaria estaltzeko?",
        "res_no_da_data":         "Ez dago laburpen agregatua kalkulatzeko datu nahiko dituzten D&A lerrorik. Egiaztatu PARA prozesua denboretan eta estazioetan existitzen dela.",
        # Global — gehigarriak
        "global_sim_turns":       "⚠️ Simulazioa: planta guztiak **{n} txanda(k)** dituztelarik",
        "global_mod_detail_plant": "**Planta:**",
        "global_mod_detail_hours": "**Ordu estimatuak:**",
        "global_mod_delete":      "🗑️ Ezabatu",
        "global_model_selected":  "**Hautatutako eredua:**",
        "global_escenario_label": "Ahalmen-eszenarioa:",
        "global_turnos_label":    "Txandak (simulazio globala):",
        "global_caption_esc":     "Eszenarioa:",
        "global_caption_turns":   "Txandak:",
        "global_disp_prefix":     "Erab.",
        "global_chart_capvsdisp": "Ahalmena vs Erabilgarritasuna Plantaka (h/urte)",
        "global_chart_legend_cap":  "Ahalmena",
        "global_chart_legend_disp": "Erabilgarritasuna",
        "global_chart_distrib":   "Ahalmen-banaketa ({esc}) - h/urte",
        # Mix — eskuineko panela
        "mix_ceiling_label":      "**Plantaren teto estrukturala:**",
        "mix_max_hours_year":     "**Ordu maximoak/urte (egitura):**",
        "mix_excess_label":       "Gainezkoa estrukturala:",
        "mix_occupancy_label":    "**Okupazio agregatua:**",
        "mix_hours_week_label":   "**Ordu/aste (baliokideak):**",
        "mix_hours_year_label":   "**Ordu/urte (baliokideak):**",
        "mix_pot_caption":        "🎚️ Potentziornetroa erabilita:",
        "mix_equiv_label":        "Baliokide hurbila",
        # Scenario persistence
        "plan_save_btn":          "Uneko eszenatokia gorde",
        "plan_save_name_label":   "Eszenatokiaren izena",
        "plan_save_name_default": "Gordetako eszenatokia",
        "plan_save_ok":           "Eszenatokia ondo gorde da.",
        "plan_save_no_db":        "Ez dago datu-base konexiorik — eszenatokia ez da gorde.",
        "plan_scenarios_header":  "Gordetako eszenatokiak",
        "plan_load_btn":          "Kargatu",
        "plan_activate_btn":      "Lehenetsita utzi",
        "plan_load_ok":           "Eszenatokia kargatu da.",
        "plan_activate_ok":       "Eszenatoki lehenetsi gisa ezarri da.",
        "plan_no_scenarios":      "Ez dago gordetako eszenatokiik planta honentzat.",
        "plan_active_marker":     "✓ aktibo",
        "plan_delete_btn":        "Ezabatu",
        "plan_delete_ok":         "Eszenatokia ezabatu da.",
        "plan_delete_blocked":    "Ezin da aktiboen eszenatokia ezabatu. Beste bat lehenetsi lehenik.",
        "plan_duplicate_btn":     "Bikoiztu",
        "plan_save_changes_btn":  "Aldaketak gorde",
        "plan_save_new_btn":      "Berri gisa gorde",
        "plan_save_changes_no_sel": "Hautatu eszenatoki bat aldaketak gorde aurretik.",
        "plan_save_changes_ok":   "Aldaketak eszenatokian gorde dira.",
    },
}

_LANG_OPTIONS = {"es": "🇪🇸 Español", "en": "🇬🇧 English", "eu": "🟥🟩⬜ Euskara"}


def t(key: str) -> str:
    """Resuelve un texto de interfaz en el idioma activo de la sesión.

    Fallback en cascada: idioma activo → español → la propia clave.
    Nunca lanza excepción: si la clave no existe devuelve la clave tal cual.
    """
    lang = st.session_state.get("lang", "es")
    return (
        _TRANSLATIONS.get(lang, _TRANSLATIONS["es"])
        .get(key, _TRANSLATIONS["es"].get(key, key))
    )


def _build_css(dark: bool) -> str:
    """Genera el bloque <style> completo según modo claro/oscuro."""
    if dark:
        bg         = "#1A1D27"
        bg_sidebar = "#141720"
        bg_card    = "#22263A"
        bdr        = "#2E3447"
        accent     = "#D93050"   # rojo Ingeteam más brillante sobre fondo oscuro
        accent_h4  = "#8BA8CC"
        text_main  = "#CDD4E0"
        text_sub   = "#8090A8"
    else:
        bg         = "#F5F6F8"
        bg_sidebar = "#FFFFFF"
        bg_card    = "#FFFFFF"
        bdr        = "#DDE1E7"
        accent     = "#A6192E"
        accent_h4  = "#2C3E50"
        text_main  = "#1A1A2E"
        text_sub   = "#5A6473"

    dark_only = (
        f"""
[data-testid="stAlert"] {{
    position: relative;
    isolation: isolate;
}}
[data-testid="stAlert"]::before {{
    content: "";
    position: absolute;
    inset: 0;
    background: rgba(10,12,22,0.55);
    z-index: 0;
    pointer-events: none;
    border-radius: 4px;
}}
[data-testid="stAlert"] > div {{
    position: relative;
    z-index: 1;
}}
[data-testid="stAlert"] p,
[data-testid="stAlert"] li {{
    color: #E0E8F8 !important;
}}
[data-baseweb="base-input"],
[data-baseweb="input"] {{
    background-color: {bg_card} !important;
    color: {text_main} !important;
}}
input[type="number"] {{
    color: {text_main} !important;
}}
"""
        if dark else ""
    )

    return f"""<style>
/* ── Fondo principal ──────────────────────────────────────── */
[data-testid="stApp"], .main {{
    background-color: {bg} !important;
}}
.main .block-container {{
    background-color: {bg} !important;
    padding-top: 1.5rem;
}}

/* ── Sidebar ──────────────────────────────────────────────── */
[data-testid="stSidebar"],
[data-testid="stSidebar"] > div:first-child {{
    background-color: {bg_sidebar} !important;
    border-right: 1px solid {bdr};
}}

/* ── Titulares h1 / h2 / h3 ──────────────────────────────── */
h1, h2, h3 {{
    color: {accent} !important;
    font-family: "Trade Gothic", "Helvetica Neue", Arial, sans-serif;
    font-weight: 700;
}}
h2 {{
    padding-bottom: 0.25em;
    border-bottom: 2px solid {accent};
    margin-bottom: 0.6em;
}}

/* ── Subtítulos h4 / h5 / h6 ─────────────────────────────── */
h4, h5, h6 {{
    color: {accent_h4} !important;
    font-family: "Trade Gothic", "Helvetica Neue", Arial, sans-serif;
    font-weight: 600;
}}

/* ── Texto cuerpo ─────────────────────────────────────────── */
.stMarkdown p, .stMarkdown li {{
    color: {text_main} !important;
    font-family: "Trade Gothic", "Helvetica Neue", Arial, sans-serif;
}}
.stCaption p {{
    color: {text_sub} !important;
    font-size: 0.83em;
}}
[data-testid="stWidgetLabel"] p, label {{
    color: {text_sub} !important;
}}

/* ── Sidebar — texto y encabezados ───────────────────────── */
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] [data-testid="stWidgetLabel"] p {{
    color: {text_main} !important;
}}
[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3 {{
    color: {accent} !important;
    border-bottom: none !important;
    padding-bottom: 0 !important;
}}

/* ── Tabs ─────────────────────────────────────────────────── */
button[data-baseweb="tab"] {{
    font-weight: 600 !important;
    color: {text_sub} !important;
    background: transparent !important;
}}
button[data-baseweb="tab"][aria-selected="true"] {{
    color: {accent} !important;
    border-bottom: 3px solid {accent} !important;
    background: transparent !important;
}}
[data-baseweb="tab-list"] {{
    background-color: {bg} !important;
    border-bottom: 1px solid {bdr} !important;
}}

/* ── Cabeceras de tabla HTML ──────────────────────────────── */
thead tr th {{
    background-color: {bg_card} !important;
    color: {accent_h4} !important;
    font-weight: 700 !important;
    border-bottom: 2px solid {bdr} !important;
}}

/* ── Expanders ────────────────────────────────────────────── */
[data-testid="stExpander"] {{
    border: 1px solid {bdr} !important;
    border-radius: 6px !important;
    overflow: hidden;
}}

/* ── Divisores ────────────────────────────────────────────── */
hr {{
    border-color: {bdr} !important;
    margin: 1.2em 0;
}}

/* ── Clases personalizadas existentes en la app ──────────── */
.red-text {{
    color: {accent} !important;
    font-weight: bold;
}}
.small-text {{
    color: {text_sub} !important;
}}
{dark_only}
</style>"""


def resource_path(relative_path: str) -> str:
    """Devuelve la ruta absoluta a un recurso.

    - En desarrollo: usa la carpeta donde está este app.py.
    - En ejecutable PyInstaller: usa la carpeta temporal sys._MEIPASS.
    """
    if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
        base_path = sys._MEIPASS  # PyInstaller onefile/onedir
    else:
        base_path = os.path.abspath(os.path.dirname(__file__))
    return os.path.join(base_path, relative_path)

st.set_page_config(
    page_title="Planificador de Capacidad",
    layout="wide"
)

# --- Conexión a base de datos Neon ---
def get_connection():
    return psycopg2.connect(os.environ["DATABASE_URL"], connect_timeout=10)


ASSETS_DIR = resource_path("assets")
# --- Cargar logo ---
@st.cache_data
def _load_logo():
    return Image.open(os.path.join(ASSETS_DIR, "ingeteam_logo.jpg"))

# Inicializar preferencias de interfaz (una vez por sesión)
if "dark_mode" not in st.session_state:
    st.session_state.dark_mode = False
if "lang" not in st.session_state:
    st.session_state.lang = "es"

# Inyectar CSS según modo actual
st.markdown(_build_css(st.session_state.dark_mode), unsafe_allow_html=True)

# ── Franja de controles compacta — encima del logo ────────────────
_ui_c1, _ui_c2 = st.sidebar.columns(2)
_ui_c1.checkbox(t("dark_mode_label"), key="dark_mode")
_ui_c2.selectbox(
    t("lang_label"),
    options=list(_LANG_OPTIONS.keys()),
    format_func=lambda x: _LANG_OPTIONS[x],
    key="lang",
)

logo = _load_logo()
st.sidebar.image(logo, use_container_width=True)

DATA_DIR = resource_path("data")

# =========================================================
# CSV IO (robusto para Windows/acentos)
# =========================================================
@st.cache_data
def load_csv(name: str) -> pd.DataFrame:
    path = os.path.join(DATA_DIR, name)
    # Intento UTF-8 (incl. BOM), y si falla, latin1 (muy común en Excel/Windows)
    for enc in ("utf-8-sig", "utf-8", "latin1"):
        try:
            df = pd.read_csv(path, encoding=enc)
            break
        except UnicodeDecodeError:
            df = None
    if df is None:
        # último recurso: que reviente con mensaje claro
        df = pd.read_csv(path)

    # Limpieza suave
    for col in df.columns:
        if df[col].dtype == "object":
            df[col] = df[col].astype(str).str.strip()
    return df


def save_csv(df: pd.DataFrame, name: str) -> None:
    path = os.path.join(DATA_DIR, name)
    # Guardamos SIEMPRE como utf-8-sig para compatibilidad con Excel
    df.to_csv(path, index=False, encoding="utf-8-sig")

# =========================================================
# DB IO (Neon / Postgres)
# =========================================================
def _has_db() -> bool:
    return bool(os.getenv("DATABASE_URL"))


@st.cache_data(ttl=60)
def load_table(table: str) -> pd.DataFrame:
    """
    Carga una tabla completa desde Postgres (Neon) y devuelve DataFrame.
    Si no hay DATABASE_URL, cae a CSV local (mismo nombre + .csv) para no romper.
    """
    if not _has_db():
        return load_csv(f"{table}.csv")

    c = get_connection()
    try:
        with c.cursor() as cur:
            cur.execute(f'SELECT * FROM "{table}"')
            cols = [d[0] for d in cur.description]
            rows = cur.fetchall()
        df = pd.DataFrame(rows, columns=cols)

        # Limpieza suave (mismo criterio que CSV)
        for col in df.columns:
            if df[col].dtype == "object":
                df[col] = df[col].astype(str).str.strip()
        return df
    finally:
        c.close()

def save_table(df: pd.DataFrame, table: str) -> None:
    """
    Sobrescribe los datos de la planta activa en Postgres.
    
    Elimina primero los registros de esa planta
    y luego inserta los nuevos.

    Evita borrar datos de otras plantas.
    """
    if not _has_db():
        save_csv(df, f"{table}.csv")
        return

    c = get_connection()
    try:
        work_df = df.copy()
        if "id" in work_df.columns:
            work_df = work_df.drop(columns=["id"])

        cols = list(work_df.columns)
        if not cols:
            return

        placeholders = ",".join(["%s"] * len(cols))
        col_list = ",".join([f'"{cname}"' for cname in cols])

        # Convertimos NaN -> None para psycopg2
        values = [tuple(None if (pd.isna(v)) else v for v in row) for row in work_df[cols].itertuples(index=False, name=None)]

        with c.cursor() as cur:
            if "plant_id" in cols:
                plant_value = int(work_df["plant_id"].iloc[0]) if not work_df.empty else int(st.session_state["plant_id"])
                # Seguridad: asegurar que solo guardamos datos de la planta correcta
                work_df = work_df[work_df["plant_id"] == plant_value].copy()
                values = [tuple(None if (pd.isna(v)) else v for v in row) for row in work_df[cols].itertuples(index=False, name=None)]
                cur.execute(f'DELETE FROM "{table}" WHERE plant_id = %s', (plant_value,))
            else:
                cur.execute(f'TRUNCATE TABLE "{table}"')

            cur.executemany(
                f'INSERT INTO "{table}" ({col_list}) VALUES ({placeholders})',
                values
            )
        c.commit()
    finally:
        c.close()

    # invalidar cache de lecturas
    try:
        load_table.clear()
        load_plant_data.clear()
        load_all_plants_data.clear()
    except Exception:
        pass


# =========================================================
# Scenario persistence (Bloque 2)
# =========================================================
def load_active_scenario(plant_id: int) -> dict | None:
    """Returns dict with keys line_model, line_demand, line_bench_variant for
    the active scenario of plant_id, or None if no active scenario exists."""
    if not _has_db():
        return None
    c = get_connection()
    try:
        with c.cursor() as cur:
            cur.execute(
                'SELECT id, name FROM "scenarios" WHERE plant_id = %s AND is_active = TRUE LIMIT 1',
                (plant_id,)
            )
            row = cur.fetchone()
            if not row:
                return None
            scenario_id, scenario_name = row[0], row[1]
            cur.execute(
                'SELECT line_id, model, demand, bench_variant FROM "scenario_lines" WHERE scenario_id = %s',
                (scenario_id,)
            )
            lines = cur.fetchall()
        result = {"scenario_id": scenario_id, "scenario_name": scenario_name, "line_model": {}, "line_demand": {}, "line_bench_variant": {}}
        for line_id, model, demand, bench_variant in lines:
            result["line_model"][line_id] = model or ""
            result["line_demand"][line_id] = float(demand) if demand is not None else 0.0
            result["line_bench_variant"][line_id] = bench_variant or ""
        return result
    except Exception:
        return None
    finally:
        c.close()


def save_scenario(plant_id: int, name: str, line_model: dict, line_demand: dict, line_bench_variant: dict) -> None:
    """Upsert active scenario for plant_id: deactivate others, insert/replace this one."""
    if not _has_db():
        return
    c = get_connection()
    try:
        with c.cursor() as cur:
            cur.execute('UPDATE "scenarios" SET is_active = FALSE WHERE plant_id = %s', (plant_id,))
            cur.execute(
                'INSERT INTO "scenarios" (plant_id, name, is_active) VALUES (%s, %s, TRUE) RETURNING id',
                (plant_id, name)
            )
            scenario_id = cur.fetchone()[0]
            all_line_ids = set(line_model.keys()) | set(line_demand.keys()) | set(line_bench_variant.keys())
            for lid in all_line_ids:
                cur.execute(
                    'INSERT INTO "scenario_lines" (scenario_id, line_id, model, demand, bench_variant) VALUES (%s, %s, %s, %s, %s)',
                    (
                        scenario_id,
                        lid,
                        line_model.get(lid, ""),
                        float(line_demand.get(lid, 0.0)),
                        line_bench_variant.get(lid, ""),
                    )
                )
        c.commit()
    finally:
        c.close()


def list_scenarios(plant_id: int) -> list[dict]:
    """Returns all scenarios for plant_id ordered by created_at DESC."""
    if not _has_db():
        return []
    c = get_connection()
    try:
        with c.cursor() as cur:
            cur.execute(
                'SELECT id, name, is_active FROM "scenarios" WHERE plant_id = %s ORDER BY created_at DESC',
                (plant_id,)
            )
            rows = cur.fetchall()
        return [{"id": r[0], "name": r[1], "is_active": bool(r[2])} for r in rows]
    except Exception:
        return []
    finally:
        c.close()


def load_line_overrides(plant_id: int) -> dict:
    """Fallback de planta: devuelve overrides por línea cuando no hay escenario activo
    o cuando el escenario activo no tiene overrides propios en scenario_line_overrides."""
    if not _has_db():
        return {}
    c = get_connection()
    try:
        with c.cursor() as cur:
            cur.execute(
                'SELECT line_id, enabled, shifts, availability, efficiency '
                'FROM "line_overrides" WHERE plant_id = %s',
                (plant_id,)
            )
            rows = cur.fetchall()
        return {
            r[0]: {
                "enabled":      bool(r[1]),
                "shifts":       int(r[2]),
                "availability": float(r[3]),
                "efficiency":   float(r[4]),
            }
            for r in rows
        }
    except Exception:
        return {}
    finally:
        c.close()


def save_line_overrides(plant_id: int, overrides: dict) -> bool:
    """Upsert line overrides for plant_id. overrides = {line_id: {enabled, shifts, availability, efficiency}}."""
    if not _has_db():
        return False
    c = get_connection()
    try:
        with c.cursor() as cur:
            for line_id, ov in overrides.items():
                cur.execute(
                    '''
                    INSERT INTO "line_overrides" (plant_id, line_id, enabled, shifts, availability, efficiency)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    ON CONFLICT (plant_id, line_id) DO UPDATE SET
                        enabled      = EXCLUDED.enabled,
                        shifts       = EXCLUDED.shifts,
                        availability = EXCLUDED.availability,
                        efficiency   = EXCLUDED.efficiency
                    ''',
                    (
                        plant_id,
                        line_id,
                        bool(ov.get("enabled", False)),
                        int(ov.get("shifts", 1)),
                        float(ov.get("availability", 1.0)),
                        float(ov.get("efficiency", 1.0)),
                    )
                )
        c.commit()
        return True
    except Exception:
        return False
    finally:
        c.close()


def load_scenario_line_overrides(scenario_id: int) -> dict:
    """Persistencia principal de overrides por línea. Ligada al escenario activo.
    Si devuelve vacío, la carga en Resultados cae a load_line_overrides (fallback por planta)."""
    if not _has_db():
        return {}
    c = get_connection()
    try:
        with c.cursor() as cur:
            cur.execute(
                'SELECT line_id, enabled, shifts, availability, efficiency '
                'FROM "scenario_line_overrides" WHERE scenario_id = %s',
                (scenario_id,)
            )
            rows = cur.fetchall()
        return {
            r[0]: {
                "enabled":      bool(r[1]),
                "shifts":       int(r[2]),
                "availability": float(r[3]),
                "efficiency":   float(r[4]),
            }
            for r in rows
        }
    except Exception:
        return {}
    finally:
        c.close()


def save_scenario_line_overrides(scenario_id: int, overrides: dict) -> bool:
    """Upsert line overrides for a scenario. overrides = {line_id: {enabled, shifts, availability, efficiency}}."""
    if not _has_db():
        return False
    c = get_connection()
    try:
        with c.cursor() as cur:
            for line_id, ov in overrides.items():
                cur.execute(
                    '''
                    INSERT INTO "scenario_line_overrides"
                        (scenario_id, line_id, enabled, shifts, availability, efficiency)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    ON CONFLICT (scenario_id, line_id) DO UPDATE SET
                        enabled      = EXCLUDED.enabled,
                        shifts       = EXCLUDED.shifts,
                        availability = EXCLUDED.availability,
                        efficiency   = EXCLUDED.efficiency
                    ''',
                    (
                        scenario_id,
                        line_id,
                        bool(ov.get("enabled", False)),
                        int(ov.get("shifts", 1)),
                        float(ov.get("availability", 1.0)),
                        float(ov.get("efficiency", 1.0)),
                    )
                )
        c.commit()
        return True
    except Exception:
        return False
    finally:
        c.close()


def load_scenario_process_shifts(scenario_id: int) -> dict:
    """Returns {line_id: {process: shifts}} for a scenario. Only rows with Personalizar=ON."""
    if not _has_db():
        return {}
    c = get_connection()
    try:
        with c.cursor() as cur:
            cur.execute(
                'SELECT line_id, process, shifts '
                'FROM "scenario_process_shift_overrides" WHERE scenario_id = %s',
                (scenario_id,)
            )
            rows = cur.fetchall()
        result: dict = {}
        for line_id, process, sh in rows:
            result.setdefault(line_id, {})[process] = int(sh)
        return result
    except Exception:
        return {}
    finally:
        c.close()


def save_scenario_process_shifts(scenario_id: int, line_id: str, data: dict) -> bool:
    """Replaces process shift overrides for (scenario_id, line_id).
    data = {process: shifts} — only personalized processes. Absence = hereda."""
    if not _has_db():
        return False
    c = get_connection()
    try:
        with c.cursor() as cur:
            cur.execute(
                'DELETE FROM "scenario_process_shift_overrides" '
                'WHERE scenario_id = %s AND line_id = %s',
                (scenario_id, line_id)
            )
            for process, sh in data.items():
                cur.execute(
                    'INSERT INTO "scenario_process_shift_overrides" '
                    '(scenario_id, line_id, process, shifts) VALUES (%s, %s, %s, %s)',
                    (scenario_id, line_id, process, int(sh))
                )
        c.commit()
        return True
    except Exception:
        return False
    finally:
        c.close()


def load_scenario_by_id(scenario_id: int) -> dict | None:
    """Loads line data for a specific scenario id."""
    if not _has_db():
        return None
    c = get_connection()
    try:
        with c.cursor() as cur:
            cur.execute('SELECT name FROM "scenarios" WHERE id = %s', (scenario_id,))
            name_row = cur.fetchone()
            scenario_name = name_row[0] if name_row else ""
            cur.execute(
                'SELECT line_id, model, demand, bench_variant FROM "scenario_lines" WHERE scenario_id = %s',
                (scenario_id,)
            )
            lines = cur.fetchall()
        result = {"scenario_id": scenario_id, "scenario_name": scenario_name, "line_model": {}, "line_demand": {}, "line_bench_variant": {}}
        for line_id, model, demand, bench_variant in lines:
            result["line_model"][line_id] = model or ""
            result["line_demand"][line_id] = float(demand) if demand is not None else 0.0
            result["line_bench_variant"][line_id] = bench_variant or ""
        return result
    except Exception:
        return None
    finally:
        c.close()


def activate_scenario(plant_id: int, scenario_id: int) -> None:
    """Sets the given scenario as active, deactivating all others for plant_id."""
    if not _has_db():
        return
    c = get_connection()
    try:
        with c.cursor() as cur:
            cur.execute('UPDATE "scenarios" SET is_active = FALSE WHERE plant_id = %s', (plant_id,))
            cur.execute('UPDATE "scenarios" SET is_active = TRUE WHERE id = %s', (scenario_id,))
        c.commit()
    finally:
        c.close()


def delete_scenario(scenario_id: int) -> None:
    """Deletes a non-active scenario and its lines. DB guard: only deletes if is_active=FALSE."""
    if not _has_db():
        return
    c = get_connection()
    try:
        with c.cursor() as cur:
            cur.execute('DELETE FROM "scenarios" WHERE id = %s AND is_active = FALSE', (scenario_id,))
        c.commit()
    finally:
        c.close()


def duplicate_scenario(scenario_id: int, plant_id: int) -> tuple[int, str] | tuple[None, None]:
    """Creates a non-active copy of a scenario. Returns (new_id, copy_name) or (None, None) on failure."""
    if not _has_db():
        return None, None
    c = get_connection()
    try:
        with c.cursor() as cur:
            cur.execute('SELECT name FROM "scenarios" WHERE id = %s', (scenario_id,))
            row = cur.fetchone()
            if not row:
                return None, None
            copy_name = f"{row[0]} (copia)"
            cur.execute(
                'INSERT INTO "scenarios" (plant_id, name, is_active) VALUES (%s, %s, FALSE) RETURNING id',
                (plant_id, copy_name)
            )
            new_id = cur.fetchone()[0]
            cur.execute(
                'INSERT INTO "scenario_lines" (scenario_id, line_id, model, demand, bench_variant) '
                'SELECT %s, line_id, model, demand, bench_variant FROM "scenario_lines" WHERE scenario_id = %s',
                (new_id, scenario_id)
            )
            cur.execute(
                'INSERT INTO "scenario_line_overrides" '
                '(scenario_id, line_id, enabled, shifts, availability, efficiency) '
                'SELECT %s, line_id, enabled, shifts, availability, efficiency '
                'FROM "scenario_line_overrides" WHERE scenario_id = %s',
                (new_id, scenario_id)
            )
            cur.execute(
                'INSERT INTO "scenario_process_shift_overrides" '
                '(scenario_id, line_id, process, shifts) '
                'SELECT %s, line_id, process, shifts '
                'FROM "scenario_process_shift_overrides" WHERE scenario_id = %s',
                (new_id, scenario_id)
            )
        c.commit()
        return new_id, copy_name
    finally:
        c.close()


def update_scenario_lines(scenario_id: int, line_model: dict, line_demand: dict, line_bench_variant: dict) -> None:
    """Replaces lines of an existing scenario without changing name or is_active."""
    if not _has_db():
        return
    c = get_connection()
    try:
        with c.cursor() as cur:
            cur.execute('DELETE FROM "scenario_lines" WHERE scenario_id = %s', (scenario_id,))
            all_lids = set(line_model.keys()) | set(line_demand.keys()) | set(line_bench_variant.keys())
            for lid in all_lids:
                cur.execute(
                    'INSERT INTO "scenario_lines" (scenario_id, line_id, model, demand, bench_variant) '
                    'VALUES (%s, %s, %s, %s, %s)',
                    (scenario_id, lid, line_model.get(lid, ""), line_demand.get(lid, 0), line_bench_variant.get(lid, ""))
                )
        c.commit()
    finally:
        c.close()


def rename_scenario(scenario_id: int, name: str) -> None:
    """Updates only the name of an existing scenario."""
    if not _has_db():
        return
    c = get_connection()
    try:
        with c.cursor() as cur:
            cur.execute('UPDATE "scenarios" SET name = %s WHERE id = %s', (name, scenario_id))
        c.commit()
    finally:
        c.close()


def create_scenario_inactive(plant_id: int, name: str, line_model: dict, line_demand: dict, line_bench_variant: dict) -> int | None:
    """Creates a new scenario with is_active=FALSE (does not touch existing active scenario). Returns the new id."""
    if not _has_db():
        return None
    c = get_connection()
    try:
        with c.cursor() as cur:
            cur.execute(
                'INSERT INTO "scenarios" (plant_id, name, is_active) VALUES (%s, %s, FALSE) RETURNING id',
                (plant_id, name)
            )
            new_id = cur.fetchone()[0]
            all_lids = set(line_model.keys()) | set(line_demand.keys()) | set(line_bench_variant.keys())
            for lid in all_lids:
                cur.execute(
                    'INSERT INTO "scenario_lines" (scenario_id, line_id, model, demand, bench_variant) '
                    'VALUES (%s, %s, %s, %s, %s)',
                    (new_id, lid, line_model.get(lid, ""), line_demand.get(lid, 0), line_bench_variant.get(lid, ""))
                )
        c.commit()
        return new_id
    finally:
        c.close()


def _on_scenario_select_change(pid) -> None:
    """on_change callback for the scenario selectbox. Syncs only the name field."""
    _sel = st.session_state.get(f"scenario_select_{pid}")
    _nmap = st.session_state.get(f"_sc_name_map_{pid}", {})
    if _sel in _nmap:
        st.session_state[f"scenario_name_input_{pid}"] = _nmap[_sel]


def ensure_int(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    out = df.copy()
    for c in cols:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0).astype(int)
    return out


@st.cache_data(ttl=300)
def load_plant_data(plant_id: int):
    if _has_db():
        # UNA sola conexión, 4 queries con WHERE plant_id = %s
        # Antes: 4 conexiones independientes × ~500 ms = ~2 s en miss
        # Ahora: 1 conexión + 4 queries = ~500 ms en miss
        _TABLES = (
            "models",
            "models_process_times",
            "lines_process_stations",
            "compatibility",
        )
        c = get_connection()
        _raw = {}
        try:
            with c.cursor() as cur:
                for tname in _TABLES:
                    cur.execute(
                        f'SELECT * FROM "{tname}" WHERE plant_id = %s',
                        (plant_id,),
                    )
                    cols = [d[0] for d in cur.description]
                    df = pd.DataFrame(cur.fetchall(), columns=cols)
                    for col in df.columns:
                        if df[col].dtype == "object":
                            df[col] = df[col].astype(str).str.strip()
                    _raw[tname] = df
        finally:
            c.close()
        models_df   = _raw["models"].copy()
        times_df    = _raw["models_process_times"].copy()
        stations_df = _raw["lines_process_stations"].copy()
        compat_df   = _raw["compatibility"].copy()
    else:
        # Fallback CSV: comportamiento original intacto
        models_df = load_table("models")
        models_df = models_df[models_df["plant_id"] == plant_id].copy()
        times_df = load_table("models_process_times")
        times_df = times_df[times_df["plant_id"] == plant_id].copy()
        stations_df = load_table("lines_process_stations")
        stations_df = stations_df[stations_df["plant_id"] == plant_id].copy()
        compat_df = load_table("compatibility")
        compat_df = compat_df[compat_df["plant_id"] == plant_id].copy()

    # Normalización — igual que antes
    models_df["model"] = models_df["model"].astype(str).str.strip()
    models_df = ensure_int(models_df, ["active"])

    times_df["model"] = times_df["model"].astype(str).str.strip()

    stations_df["line"] = stations_df["line"].astype(str).str.strip()
    stations_df["nave"] = stations_df["nave"].astype(str).str.strip()
    stations_df["line_id"] = stations_df["nave"] + "-" + stations_df["line"]

    compat_df["line"] = compat_df["line"].astype(str).str.strip()
    compat_df["model"] = compat_df["model"].astype(str).str.strip()
    compat_df["nave"] = compat_df["nave"].astype(str).str.strip()
    compat_df["line_id"] = compat_df["nave"] + "-" + compat_df["line"]
    compat_df = ensure_int(compat_df, ["compatible"])

    active_models = models_df.loc[models_df["active"] == 1, "model"].tolist()

    line_ids_nave = sorted(
        stations_df["line_id"].astype(str).str.strip().unique().tolist()
    )

    return {
        "models_df": models_df,
        "times_df": times_df,
        "stations_df": stations_df,
        "compat_df": compat_df,
        "active_models": active_models,
        "line_ids_nave": line_ids_nave,
    }


@st.cache_data
def compute_plant_structural_capacity(
    p_id: int,
    p_name: str,
    all_settings: pd.DataFrame,
    all_models: pd.DataFrame,
    all_times: pd.DataFrame,
    all_stations: pd.DataFrame,
    all_compat: pd.DataFrame,
    shifts_override: int = None,
    hours_week_override: float = None,
    availability_override: float = None,
    efficiency_override: float = None,
) -> dict:
    """Calcula capacidad estructural (max/prom/min) para una planta.
    Los parámetros _override sustituyen al valor de settings cuando no son None."""

    # Obtener settings de la planta
    p_settings = all_settings[all_settings["plant_id"] == p_id]
    if p_settings.empty:
        p_hours_week = 43.0
        p_shifts = 1
        p_availability = 1.0
        p_efficiency = 1.0
        p_days_year = 250
        p_days_week = 5
    else:
        row = p_settings.iloc[0]
        p_hours_week = float(row.get("hours_week", 43.0))
        p_shifts = int(row.get("shifts", 1))
        p_availability = float(row.get("availability", 1.0))
        p_efficiency = float(row.get("efficiency", 1.0))
        p_days_year = int(row.get("days_open_year", 250))
        p_days_week = int(row.get("days_open_week", 5))

    effective_hours_week   = hours_week_override   if hours_week_override   is not None else p_hours_week
    effective_shifts       = shifts_override        if shifts_override        is not None else p_shifts
    effective_availability = availability_override  if availability_override  is not None else p_availability
    effective_efficiency   = efficiency_override    if efficiency_override    is not None else p_efficiency
    p_hours_eff = effective_hours_week * effective_shifts * effective_availability * effective_efficiency
    p_weeks_equiv = p_days_year / max(p_days_week, 1)

    # Filtrar datos por planta
    p_models = all_models[all_models["plant_id"] == p_id].copy()
    p_times = all_times[all_times["plant_id"] == p_id].copy()
    p_stations = all_stations[all_stations["plant_id"] == p_id].copy()
    p_compat = all_compat[all_compat["plant_id"] == p_id].copy()

    if p_models.empty or p_stations.empty:
        return {
            "plant_id": p_id, "plant_name": p_name,
            "hours_eff": p_hours_eff, "weeks_equiv": p_weeks_equiv,
            "max_u_sem": 0.0, "prom_u_sem": 0.0, "min_u_sem": 0.0,
            "max_u_year": 0.0, "prom_u_year": 0.0, "min_u_year": 0.0,
            "max_h_sem": 0.0, "prom_h_sem": 0.0, "min_h_sem": 0.0,
            "max_h_year": 0.0, "prom_h_year": 0.0, "min_h_year": 0.0,
            "lines_count": 0, "line_stats": [],
        }

    # Normalización
    p_models["model"] = p_models["model"].astype(str).str.strip()
    p_times["model"] = p_times["model"].astype(str).str.strip()
    p_stations["line"] = p_stations["line"].astype(str).str.strip()
    p_compat["line"] = p_compat["line"].astype(str).str.strip()
    p_compat["model"] = p_compat["model"].astype(str).str.strip()
    p_compat["nave"] = p_compat["nave"].astype(str).str.strip()
    p_compat["line_id"] = p_compat["nave"] + "-" + p_compat["line"]

    p_models = ensure_int(p_models, ["active"])
    p_compat = ensure_int(p_compat, ["compatible"])

    active_models_p = p_models.loc[p_models["active"] == 1, "model"].tolist()

    if not active_models_p:
        return {
            "plant_id": p_id, "plant_name": p_name,
            "hours_eff": p_hours_eff, "weeks_equiv": p_weeks_equiv,
            "max_u_sem": 0.0, "prom_u_sem": 0.0, "min_u_sem": 0.0,
            "max_u_year": 0.0, "prom_u_year": 0.0, "min_u_year": 0.0,
            "max_h_sem": 0.0, "prom_h_sem": 0.0, "min_h_sem": 0.0,
            "max_h_year": 0.0, "prom_h_year": 0.0, "min_h_year": 0.0,
            "lines_count": 0, "line_stats": [],
        }

    # Compatibilidades activas
    compat_active_p = p_compat[(p_compat["compatible"] == 1) & (p_compat["model"].isin(active_models_p))].copy()
    allowed_by_line_p = compat_active_p.groupby("line_id")["model"].apply(list).to_dict()

    # Cycle times por modelo (para horas visibles)
    _t = p_times.copy()
    _t["cycle_time"] = pd.to_numeric(_t["cycle_time"], errors="coerce").fillna(0.0)
    cycle_by_model_p = _t.groupby("model")["cycle_time"].sum().to_dict()

    # Pre-convertir tipos numéricos una sola vez
    p_times["cycle_time"] = pd.to_numeric(p_times["cycle_time"], errors="coerce").fillna(0.0)

    if "machine_time" in p_times.columns:
        p_times["machine_time"] = pd.to_numeric(p_times["machine_time"], errors="coerce").fillna(0.0)
    else:
        p_times["machine_time"] = 0.0

    if "labor_time" in p_times.columns:
        p_times["labor_time"] = pd.to_numeric(p_times["labor_time"], errors="coerce").fillna(0.0)
    else:
        p_times["labor_time"] = 0.0

    p_stations["stations"] = pd.to_numeric(p_stations["stations"], errors="coerce").fillna(0)
    p_stations["operators_per_station"] = pd.to_numeric(p_stations["operators_per_station"], errors="coerce").fillna(0)

    # Pre-indexar tiempos por modelo
    _time_cols = ["process", "cycle_time", "machine_time", "labor_time"]
    times_by_model = {m: grp[_time_cols] for m, grp in p_times.groupby("model")}

    # Pre-indexar estaciones por line_id
    stations_by_line = {lid: grp[["process", "stations", "operators_per_station"]]
                        for lid, grp in p_stations.groupby("line_id")}

    # Line IDs
    line_ids_p = sorted(p_stations["line_id"].astype(str).str.strip().unique().tolist())

    line_stats_rows = []

    for line_id in line_ids_p:
        parts = line_id.split("-", 1)
        if len(parts) == 2:
            nave, base_line = parts
        else:
            nave = "N1"
            base_line = parts[0]

        models_allowed = allowed_by_line_p.get(line_id, [])
        if not models_allowed:
            continue

        capU_vals = []
        capH_vals = []

        s = stations_by_line.get(line_id)
        if s is None or s.empty:
            continue

        for m in models_allowed:
            t = times_by_model.get(m)
            if t is None or t.empty:
                continue

            merged = pd.merge(s, t, on="process", how="inner")

            if merged.empty:
                continue

            productive = merged[(merged["stations"] > 0)].copy()
            if productive.empty:
                continue

            ops = productive["operators_per_station"].clip(lower=1.0)
            labor_per_op = productive["labor_time"] / ops
            ctr = pd.concat([productive["machine_time"], labor_per_op], axis=1).max(axis=1)

            valid = ctr > 0
            if not valid.any():
                continue

            cap_values = (p_hours_eff * productive.loc[valid, "stations"]) / ctr[valid]

            cap_min = float(cap_values.min())
            if cap_min <= 0:
                continue

            w_m = float(cycle_by_model_p.get(m, 0.0))

            capU_vals.append(cap_min)
            capH_vals.append(cap_min * w_m)

        if not capU_vals:
            continue

        line_stats_rows.append({
            "line_id": line_id, "nave": nave, "line": base_line,
            "max_u": float(np.max(capU_vals)),
            "prom_u": float(np.mean(capU_vals)),
            "min_u": float(np.min(capU_vals)),
            "max_h": float(np.max(capH_vals)),
            "prom_h": float(np.mean(capH_vals)),
            "min_h": float(np.min(capH_vals)),
        })

    if not line_stats_rows:
        return {
            "plant_id": p_id, "plant_name": p_name,
            "hours_eff": p_hours_eff, "weeks_equiv": p_weeks_equiv,
            "max_u_sem": 0.0, "prom_u_sem": 0.0, "min_u_sem": 0.0,
            "max_u_year": 0.0, "prom_u_year": 0.0, "min_u_year": 0.0,
            "max_h_sem": 0.0, "prom_h_sem": 0.0, "min_h_sem": 0.0,
            "max_h_year": 0.0, "prom_h_year": 0.0, "min_h_year": 0.0,
            "lines_count": 0, "line_stats": [],
        }

    # Agregar por planta
    max_u_sem  = sum(r["max_u"]  for r in line_stats_rows)
    prom_u_sem = sum(r["prom_u"] for r in line_stats_rows)
    min_u_sem  = sum(r["min_u"]  for r in line_stats_rows)
    max_h_sem  = sum(r["max_h"]  for r in line_stats_rows)
    prom_h_sem = sum(r["prom_h"] for r in line_stats_rows)
    min_h_sem  = sum(r["min_h"]  for r in line_stats_rows)

    return {
        "plant_id": p_id, "plant_name": p_name,
        "hours_eff": p_hours_eff, "weeks_equiv": p_weeks_equiv,
        "max_u_sem": max_u_sem, "prom_u_sem": prom_u_sem, "min_u_sem": min_u_sem,
        "max_u_year": max_u_sem * p_weeks_equiv,
        "prom_u_year": prom_u_sem * p_weeks_equiv,
        "min_u_year": min_u_sem * p_weeks_equiv,
        "max_h_sem": max_h_sem, "prom_h_sem": prom_h_sem, "min_h_sem": min_h_sem,
        "max_h_year": max_h_sem * p_weeks_equiv,
        "prom_h_year": prom_h_sem * p_weeks_equiv,
        "min_h_year": min_h_sem * p_weeks_equiv,
        "lines_count": len(line_stats_rows),
        "line_stats": line_stats_rows,
    }


@st.cache_data(show_spinner=False)
def compute_all_plants_structural_capacity(
    all_data: dict,
    shifts_override: int = None,
) -> list:
    """Calcula capacidad estructural para TODAS las plantas en una sola llamada cacheada.
    all_data se hashea UNA sola vez en lugar de 5 DataFrames × N plantas."""
    results = []
    for _, plant_row in all_data["plants"].iterrows():
        p_id = int(plant_row["id"])
        p_name = str(plant_row["name"])
        result = compute_plant_structural_capacity(
            p_id, p_name,
            all_data["settings"], all_data["models"], all_data["times"],
            all_data["stations"], all_data["compat"],
            shifts_override=shifts_override,
        )
        results.append(result)
    return results


# =========================================================
# SELECCIÓN DE PLANTA.
# =========================================================
plants_df = load_table("plants")

if plants_df.empty:
    st.sidebar.error("No hay plantas definidas en la tabla plants")
    st.stop()

plant_names = plants_df["name"].astype(str).tolist()

selected_plant_name = st.sidebar.selectbox(
    t("plant_select"),
    plant_names
)

plant_id = int(
    plants_df.loc[plants_df["name"] == selected_plant_name, "id"].iloc[0]
)

st.session_state["plant_id"] = plant_id

# =========================================================
# AÑADIR NUEVA PLANTA
# =========================================================
new_plant_name = st.sidebar.text_input(t("plant_new"))

if st.sidebar.button(t("plant_add")):
    if new_plant_name.strip():
        if _has_db():
            c = get_connection()
            try:
                with c.cursor() as cur:
                    cur.execute(
                        'INSERT INTO "plants" ("name") VALUES (%s)',
                        (new_plant_name.strip(),)
                    )
                c.commit()
            finally:
                c.close()
            try:
                load_table.clear()
                load_plant_data.clear()
                load_all_plants_data.clear()
            except Exception:
                pass
        else:
            next_id = int(plants_df["id"].max()) + 1 if not plants_df.empty else 1
            new_row = pd.DataFrame([{"id": next_id, "name": new_plant_name.strip()}])
            plants_df = pd.concat([plants_df, new_row], ignore_index=True)
            save_csv(plants_df, "plants.csv")

        st.sidebar.success(t("plant_added"))
        st.rerun()
    else:
        st.sidebar.warning(t("plant_name_empty"))
# =========================================================
# APP CONFIG
# =========================================================

st.title(t("app_title"))
st.caption(t("app_subtitle"))

# =========================================================
# SIDEBAR – NAVEGACIÓN
# =========================================================
_PAGES = [
    "🌐 Global",
    "📊 Planificación",
    "⚙️ Configuración (Power User)",
    "📈 Resultados",
    "🧭 Capacidad según mix",
    "📅 Simulación anual",
]
_PAGES_I18N = {p: k for p, k in zip(_PAGES, [
    "nav_global", "nav_planning", "nav_config", "nav_results", "nav_mix", "nav_sim_annual"
])}
st.sidebar.radio(
    t("nav_label"),
    _PAGES,
    key="active_tab",
    format_func=lambda x: t(_PAGES_I18N[x]),
)

st.sidebar.divider()

# =========================================================
# SIDEBAR – PARÁMETROS (SIEMPRE VISIBLES)
# =========================================================
st.sidebar.header(t("params_header"))

settings_df = load_table("settings")
settings_df = settings_df[settings_df["plant_id"] == plant_id]

if settings_df.empty:
    current_settings = {
        "hours_week": 43.0,
        "shifts": 1,
        "availability": 1.0,
        "efficiency": 1.0,
        "days_open_year": 250,
        "days_open_week": 5,
    }
else:
    row = settings_df.iloc[0]
    current_settings = {
        "hours_week": float(row["hours_week"]),
        "shifts": int(row["shifts"]),
        "availability": float(row["availability"]),
        "efficiency": float(row["efficiency"]),
        "days_open_year": int(row["days_open_year"]),
        "days_open_week": int(row["days_open_week"]),
    }

hours_week = st.sidebar.number_input(
    t("param_hours_week"),
    min_value=0.0,
    value=current_settings["hours_week"],
    step=0.5
)

shifts = st.sidebar.number_input(
    t("param_shifts"),
    min_value=1,
    value=current_settings["shifts"],
    step=1
)

availability = st.sidebar.slider(
    t("param_availability"),
    0.0, 1.0,
    current_settings["availability"],
    0.01
)

efficiency = st.sidebar.slider(
    t("param_efficiency"),
    0.0, 1.0,
    current_settings["efficiency"],
    0.01
)

st.sidebar.divider()

days_open_year = st.sidebar.number_input(
    t("param_days_year"),
    min_value=1,
    value=current_settings["days_open_year"],
    step=1
)

days_open_week = st.sidebar.number_input(
    t("param_days_week"),
    min_value=1,
    max_value=7,
    value=current_settings["days_open_week"],
    step=1
)

weeks_equiv = days_open_year / max(days_open_week, 1)
hours_eff = hours_week * shifts * availability * efficiency

st.sidebar.caption(f"{t('caption_hours_eff')} **{_fmt_num(hours_eff)} {t('unit_week')}**")
st.sidebar.caption(f"{t('caption_weeks_equiv')} **{_fmt_num(weeks_equiv)} {t('unit_year')}**")

# ---------------------------------------------------------
# GUARDAR PARÁMETROS DE ESTA PLANTA
# ---------------------------------------------------------

if st.sidebar.button(t("btn_save_params")):

    new_settings = pd.DataFrame([{
        "plant_id": plant_id,
        "hours_week": hours_week,
        "shifts": shifts,
        "availability": availability,
        "efficiency": efficiency,
        "days_open_year": days_open_year,
        "days_open_week": days_open_week
    }])

    save_table(new_settings, "settings")

    st.sidebar.success(t("msg_params_saved"))

# =========================================================
# CARGA DATOS
# =========================================================
_pd = load_plant_data(plant_id)

models_df     = _pd["models_df"]
times_df      = _pd["times_df"]
stations_df   = _pd["stations_df"]
compat_df     = _pd["compat_df"]
active_models = _pd["active_models"]
line_ids_nave = _pd["line_ids_nave"]

# Líneas disponibles (derivadas de stations_df)
lines = sorted(stations_df["line"].unique().tolist())

# Compatibilidad activa — calculada una sola vez, compartida por Tab 1 y Tab 4
compat_active = compat_df[(compat_df["compatible"] == 1) & (compat_df["model"].isin(active_models))]
allowed_by_line = compat_active.groupby("line_id")["model"].apply(list).to_dict()

# =========================================================
# Inicialización segura de session_state de planificación
# Debe estar a nivel de módulo (fuera de cualquier tab) para que Resultados
# y Capacidad según mix no rompan al entrar sin pasar por Planificación.
# =========================================================
if "line_model" not in st.session_state:
    st.session_state.line_model = {}
if "line_demand" not in st.session_state:
    st.session_state.line_demand = {}
# Variante de prueba por línea D&A — fase 2 bancos de prueba
# Estructura: {plant_id: {line_id: 'LV' | 'MV' | ''}}
# Vacío ('') significa "usar regla general de la familia".
if "line_bench_variant" not in st.session_state:
    st.session_state.line_bench_variant = {}
_pid_init = st.session_state["plant_id"]
if _pid_init != st.session_state.get("_last_pid"):
    # Plant changed (first visit or return) — always reload active scenario from DB
    _scenario = load_active_scenario(_pid_init)
    if _scenario:
        st.session_state.line_model[_pid_init] = _scenario["line_model"]
        st.session_state.line_demand[_pid_init] = _scenario["line_demand"]
        st.session_state.line_bench_variant[_pid_init] = _scenario["line_bench_variant"]
        st.session_state[f"_session_sc_id_{_pid_init}"] = _scenario["scenario_id"]
        _scenario["_msg"] = None
        st.session_state[f"_pending_scenario_{_pid_init}"] = _scenario
    else:
        if _pid_init not in st.session_state.line_model:
            st.session_state.line_model[_pid_init] = {}
        if _pid_init not in st.session_state.line_demand:
            st.session_state.line_demand[_pid_init] = {}
        if _pid_init not in st.session_state.line_bench_variant:
            st.session_state.line_bench_variant[_pid_init] = {}
    # Populate name map on plant change — ensures Resultados and Simulación
    # always have the scenario name available, even before Planificación renders.
    if _has_db():
        _init_sc_list = list_scenarios(_pid_init)
        st.session_state[f"_sc_name_map_{_pid_init}"] = {s["id"]: s["name"] for s in _init_sc_list}
    st.session_state["_last_pid"] = _pid_init
if _pid_init not in st.session_state.line_demand:
    st.session_state.line_demand[_pid_init] = {}
if _pid_init not in st.session_state.line_bench_variant:
    st.session_state.line_bench_variant[_pid_init] = {}

# =========================================================
# Capacidad por modelo y planta (cacheada a nivel de módulo)
# =========================================================
@st.cache_data(show_spinner=False)
def compute_model_capacity_by_plant(model_name: str, p_id: int, all_data: dict, shifts_override: int = None) -> dict:
    """Calcula capacidad de un modelo específico en una planta."""
    p_settings = all_data["settings"][all_data["settings"]["plant_id"] == p_id]
    if p_settings.empty:
        p_hours_eff = 43.0
        p_weeks_equiv = 50.0
    else:
        row = p_settings.iloc[0]
        p_hours_week = float(row.get("hours_week", 43.0))
        p_shifts = int(row.get("shifts", 1))
        p_availability = float(row.get("availability", 1.0))
        p_efficiency = float(row.get("efficiency", 1.0))
        p_days_year = int(row.get("days_open_year", 250))
        p_days_week = int(row.get("days_open_week", 5))
        effective_shifts = shifts_override if shifts_override else p_shifts
        p_hours_eff = p_hours_week * effective_shifts * p_availability * p_efficiency
        p_weeks_equiv = p_days_year / max(p_days_week, 1)

    p_times = all_data["times"][all_data["times"]["plant_id"] == p_id].copy()
    p_stations = all_data["stations"][all_data["stations"]["plant_id"] == p_id].copy()
    p_compat = all_data["compat"][all_data["compat"]["plant_id"] == p_id].copy()

    p_times["model"] = p_times["model"].astype(str).str.strip()
    p_compat["model"] = p_compat["model"].astype(str).str.strip()
    p_compat["line"] = p_compat["line"].astype(str).str.strip()
    p_compat["nave"] = p_compat["nave"].astype(str).str.strip()
    p_compat["line_id"] = p_compat["nave"] + "-" + p_compat["line"]
    p_compat = ensure_int(p_compat, ["compatible"])

    # Líneas compatibles con este modelo (por line_id)
    compat_line_ids = p_compat[(p_compat["model"] == model_name) & (p_compat["compatible"] == 1)]["line_id"].tolist()

    if not compat_line_ids:
        return {"cap_u_sem": 0.0, "cap_h_sem": 0.0}

    model_times = p_times[p_times["model"] == model_name].copy()
    model_times["cycle_time"] = pd.to_numeric(model_times["cycle_time"], errors="coerce").fillna(0.0)
    total_cycle = float(model_times["cycle_time"].sum())

    line_ids = p_stations["line_id"].astype(str).str.strip().unique().tolist()

    total_cap_u = 0.0
    total_cap_h = 0.0

    for line_id in line_ids:
        if line_id not in compat_line_ids:
            continue

        t = model_times.copy()
        s = p_stations[p_stations["line_id"] == line_id].copy()

        merged = pd.merge(s, t, on="process", how="inner")

        if merged.empty:
            continue

        merged["stations"] = pd.to_numeric(merged["stations"], errors="coerce").fillna(0)
        merged["operators_per_station"] = pd.to_numeric(merged["operators_per_station"], errors="coerce").fillna(0).clip(lower=1.0)
        for _tc in ("machine_time", "labor_time"):
            if _tc in merged.columns:
                merged[_tc] = pd.to_numeric(merged[_tc], errors="coerce").fillna(0.0)
            else:
                merged[_tc] = 0.0

        productive = merged[merged["stations"] > 0].copy()
        if productive.empty:
            continue

        labor_per_op = productive["labor_time"] / productive["operators_per_station"]
        ctr = pd.concat([productive["machine_time"], labor_per_op], axis=1).max(axis=1)
        valid = ctr > 0
        if not valid.any():
            continue

        cap_final = (p_hours_eff * productive.loc[valid, "stations"]) / ctr[valid]

        if cap_final.empty or cap_final.min() <= 0:
            continue

        cap_u = float(cap_final.min())
        cap_h = cap_u * total_cycle

        total_cap_u += cap_u
        total_cap_h += cap_h

    return {
        "cap_u_sem": total_cap_u,
        "cap_h_sem": total_cap_h,
        "weeks_equiv": p_weeks_equiv,
    }


# =========================================================
# Carga global multiplanta (cacheada a nivel de módulo)
# =========================================================
@st.cache_data(ttl=300)
def load_all_plants_data():
    """Carga datos de todas las plantas en UNA conexión en lugar de 6.
    Antes: load_table() × 6 → 6 conexiones TCP (~3 s en miss).
    Ahora: get_connection() × 1 + 6 queries (~600 ms en miss)."""
    if not _has_db():
        return {
            "plants":   load_table("plants"),
            "settings": load_table("settings"),
            "models":   load_table("models"),
            "times":    load_table("models_process_times"),
            "stations": load_table("lines_process_stations"),
            "compat":   load_table("compatibility"),
        }
    _TABLES = (
        "plants", "settings", "models",
        "models_process_times", "lines_process_stations", "compatibility",
    )
    c = get_connection()
    _dfs = {}
    try:
        with c.cursor() as cur:
            for tname in _TABLES:
                cur.execute(f'SELECT * FROM "{tname}"')
                cols = [d[0] for d in cur.description]
                df = pd.DataFrame(cur.fetchall(), columns=cols)
                for col in df.columns:
                    if df[col].dtype == "object":
                        df[col] = df[col].astype(str).str.strip()
                _dfs[tname] = df
    finally:
        c.close()
    return {
        "plants":   _dfs["plants"],
        "settings": _dfs["settings"],
        "models":   _dfs["models"],
        "times":    _dfs["models_process_times"],
        "stations": _dfs["lines_process_stations"],
        "compat":   _dfs["compatibility"],
    }


# Track previous active tab — used by Planificación to detect tab-return without syncing on every render.
_prev_active_tab = st.session_state.get("_active_tab_prev", st.session_state.active_tab)
st.session_state["_active_tab_prev"] = st.session_state.active_tab

# =========================================================
# 0) GLOBAL - VISIÓN MULTIPLANTA
# =========================================================

if st.session_state.active_tab == "🌐 Global":
    st.subheader(t("tab_global_header"))
    st.info(t("global_info"))
    
    # --- Cargar TODOS los datos de TODAS las plantas ---
    all_data = load_all_plants_data()

    # --- Calcular capacidad para TODAS las plantas ---
    # Selector global de turnos (solo afecta a esta pestaña)
    st.markdown(t("global_scenario_header"))
    
    col_esc1, col_esc2, col_esc3 = st.columns([1, 1, 2])
    with col_esc1:
        escenario = st.radio(
            t("global_escenario_label"),
            ["Máximo", "Promedio", "Mínimo"],
            index=1,  # Promedio por defecto
            horizontal=True,
            key="global_escenario"
        )
    with col_esc2:
        turnos_option = st.radio(
            t("global_turnos_label"),
            ["Config. actual", "1 turno", "2 turnos", "3 turnos"],
            index=0,
            horizontal=True,
            key="global_turnos"
        )
    
    # Mapear selección de turnos a valor numérico (None = usar config de cada planta)
    _turnos_map = {"Config. actual": None, "1 turno": 1, "2 turnos": 2, "3 turnos": 3}
    shifts_override = _turnos_map[turnos_option]
    
    global_results = compute_all_plants_structural_capacity(all_data, shifts_override)

    # Mapear escenario a columnas
    esc_map = {
        "Máximo": ("max_u_sem", "max_u_year", "max_h_sem", "max_h_year"),
        "Promedio": ("prom_u_sem", "prom_u_year", "prom_h_sem", "prom_h_year"),
        "Mínimo": ("min_u_sem", "min_u_year", "min_h_sem", "min_h_year"),
    }
    u_sem_col, u_year_col, h_sem_col, h_year_col = esc_map[escenario]
    
    if shifts_override:
        st.caption(t("global_sim_turns").format(n=shifts_override))
    
    st.divider()
    
    # =====================================================
    # 2️⃣ RESUMEN GLOBAL DE CAPACIDAD
    # =====================================================
    st.markdown(t("global_capacity_header"))
    st.caption(f"{t('global_caption_esc')} **{escenario}**" + (f" | {t('global_caption_turns')} **{shifts_override}**" if shifts_override else ""))
    
    # Construir DataFrame de resumen
    resumen_rows = []
    for r in global_results:
        resumen_rows.append({
            "PLANTA": r["plant_name"],
            "LÍNEAS": r["lines_count"],
            "CAP. LÍNEA (uds/sem)": r[u_sem_col],
            "CAP. LÍNEA (uds/año)": r[u_year_col],
            "CAPACIDAD (h/sem)": r[h_sem_col],
            "CAPACIDAD (h/año)": r[h_year_col],
        })
    
    resumen_df = pd.DataFrame(resumen_rows)
    
    # Añadir fila TOTAL
    total_row = {
        "PLANTA": "TOTAL",
        "LÍNEAS": int(resumen_df["LÍNEAS"].sum()),
        "CAP. LÍNEA (uds/sem)": resumen_df["CAP. LÍNEA (uds/sem)"].sum(),
        "CAP. LÍNEA (uds/año)": resumen_df["CAP. LÍNEA (uds/año)"].sum(),
        "CAPACIDAD (h/sem)": resumen_df["CAPACIDAD (h/sem)"].sum(),
        "CAPACIDAD (h/año)": resumen_df["CAPACIDAD (h/año)"].sum(),
    }
    resumen_df = pd.concat([resumen_df, pd.DataFrame([total_row])], ignore_index=True)
    
    # Estilo para la tabla
    def style_resumen(df):
        def highlight_total(row):
            if row["PLANTA"] == "TOTAL":
                return ["font-weight: bold; background-color: #f0f0f0;"] * len(row)
            return [""] * len(row)
        return df.style.apply(highlight_total, axis=1).format({
            "CAP. LÍNEA (uds/sem)": "{:.1f}",
            "CAP. LÍNEA (uds/año)": "{:.1f}",
            "CAPACIDAD (h/sem)": "{:.1f}",
            "CAPACIDAD (h/año)": "{:.1f}",
        })
    
    st.dataframe(style_resumen(resumen_df), use_container_width=True, hide_index=True)
    
    st.divider()
    
    # =====================================================
    # 3️⃣ DISPONIBILIDAD Y % USO
    # =====================================================
    st.markdown(t("global_avail_header"))
    st.caption(t("global_avail_caption"))
    
    # Inicializar disponibilidad en session_state
    if "global_disponibilidad" not in st.session_state:
        st.session_state.global_disponibilidad = {}
    
    # Crear inputs de disponibilidad
    disp_cols = st.columns(min(len(global_results), 4))
    for i, r in enumerate(global_results):
        col_idx = i % len(disp_cols)
        with disp_cols[col_idx]:
            default_disp = st.session_state.global_disponibilidad.get(r["plant_name"], r["max_h_year"] * 1.2)
            disp_value = st.number_input(
                f"{t('global_disp_prefix')} {r['plant_name']} (h/año)",
                min_value=0.0,
                value=float(default_disp),
                step=1000.0,
                key=f"disp_{r['plant_name']}"
            )
            st.session_state.global_disponibilidad[r["plant_name"]] = disp_value
    
    # Tabla Capacidad vs Disponibilidad
    cap_disp_rows = []
    for r in global_results:
        cap_h_year = r[h_year_col]
        disp_h_year = st.session_state.global_disponibilidad.get(r["plant_name"], 0.0)
        pct_uso = (cap_h_year / disp_h_year * 100) if disp_h_year > 0 else 0.0
        
        cap_disp_rows.append({
            "PLANTA": r["plant_name"],
            "Capacidad (h/año)": cap_h_year,
            "Disponibilidad (h/año)": disp_h_year,
            "% USO CAPACIDAD": pct_uso,
        })
    
    cap_disp_df = pd.DataFrame(cap_disp_rows)
    
    # Añadir total
    total_cap = cap_disp_df["Capacidad (h/año)"].sum()
    total_disp = cap_disp_df["Disponibilidad (h/año)"].sum()
    total_pct = (total_cap / total_disp * 100) if total_disp > 0 else 0.0
    
    cap_disp_df = pd.concat([cap_disp_df, pd.DataFrame([{
        "PLANTA": "TOTAL",
        "Capacidad (h/año)": total_cap,
        "Disponibilidad (h/año)": total_disp,
        "% USO CAPACIDAD": total_pct,
    }])], ignore_index=True)
    
    def style_cap_disp(df):
        def color_pct(val):
            try:
                v = float(val)
                if v >= 100:
                    return "color: red; font-weight: bold;"
                elif v >= 80:
                    return "color: orange; font-weight: bold;"
                else:
                    return "color: green; font-weight: bold;"
            except:
                return ""
        
        def highlight_total(row):
            if row["PLANTA"] == "TOTAL":
                return ["font-weight: bold; background-color: #f0f0f0;"] * len(row)
            return [""] * len(row)
        
        return df.style.apply(highlight_total, axis=1).map(
            color_pct, subset=["% USO CAPACIDAD"]
        ).format({
            "Capacidad (h/año)": "{:,.1f}",
            "Disponibilidad (h/año)": "{:,.1f}",
            "% USO CAPACIDAD": "{:.1f}%",
        })
    
    st.dataframe(style_cap_disp(cap_disp_df), use_container_width=True, hide_index=True)
    
    # Gráfico de barras
    fig_cap_disp = go.Figure()
    plants_no_total = [r["PLANTA"] for r in cap_disp_rows]
    caps = [r["Capacidad (h/año)"] for r in cap_disp_rows]
    disps = [r["Disponibilidad (h/año)"] for r in cap_disp_rows]
    
    fig_cap_disp.add_bar(x=plants_no_total, y=caps, name=t("global_chart_legend_cap"), marker_color="#A6192E")
    fig_cap_disp.add_bar(x=plants_no_total, y=disps, name=t("global_chart_legend_disp"), marker_color="#2E75B6")
    fig_cap_disp.update_layout(
        barmode="group",
        title=t("global_chart_capvsdisp"),
        height=400,
    )
    st.plotly_chart(fig_cap_disp, use_container_width=True, key="chart_cap_disp_global")
    
    st.divider()
    
    # =====================================================
    # 4️⃣ RESUMEN POR MODELO
    # =====================================================
    st.markdown(t("global_model_header"))

    # Obtener todos los modelos activos de todas las plantas
    all_active_models = set()
    for _, plant_row in all_data["plants"].iterrows():
        p_id = int(plant_row["id"])
        p_models = all_data["models"][all_data["models"]["plant_id"] == p_id]
        p_models = ensure_int(p_models.copy(), ["active"])
        active_m = p_models.loc[p_models["active"] == 1, "model"].astype(str).str.strip().tolist()
        all_active_models.update(active_m)
    
    all_active_models = sorted(list(all_active_models))
    
    if all_active_models:
        selected_model = st.selectbox(
            t("global_model_select"),
            ["Todos los modelos"] + all_active_models,
            key="global_model_filter"
        )
        
        if selected_model != "Todos los modelos":
            # Mostrar capacidad del modelo seleccionado en todas las plantas
            model_rows = []
            for _, plant_row in all_data["plants"].iterrows():
                p_id = int(plant_row["id"])
                p_name = str(plant_row["name"])
                
                cap_data = compute_model_capacity_by_plant(selected_model, p_id, all_data, shifts_override=shifts_override)
                
                model_rows.append({
                    "PLANTA": p_name,
                    "Capacidad (uds/sem)": cap_data["cap_u_sem"],
                    "Capacidad (uds/año)": cap_data["cap_u_sem"] * cap_data.get("weeks_equiv", 50),
                    "Capacidad (h/sem)": cap_data["cap_h_sem"],
                    "Capacidad (h/año)": cap_data["cap_h_sem"] * cap_data.get("weeks_equiv", 50),
                })
            
            model_df = pd.DataFrame(model_rows)
            
            # Añadir total
            model_df = pd.concat([model_df, pd.DataFrame([{
                "PLANTA": "TOTAL",
                "Capacidad (uds/sem)": model_df["Capacidad (uds/sem)"].sum(),
                "Capacidad (uds/año)": model_df["Capacidad (uds/año)"].sum(),
                "Capacidad (h/sem)": model_df["Capacidad (h/sem)"].sum(),
                "Capacidad (h/año)": model_df["Capacidad (h/año)"].sum(),
            }])], ignore_index=True)
            
            st.markdown(f"{t('global_model_selected')} {selected_model}")
            st.dataframe(
                model_df.style.format({
                    "Capacidad (uds/sem)": "{:.1f}",
                    "Capacidad (uds/año)": "{:.1f}",
                    "Capacidad (h/sem)": "{:.1f}",
                    "Capacidad (h/año)": "{:.1f}",
                }),
                use_container_width=True,
                hide_index=True
            )
        else:
            st.info(t("global_model_info"))
    else:
        st.warning(t("global_no_models"))
    
    st.divider()
    
    # =====================================================
    # 5️⃣ PANEL DE MODIFICACIONES
    # =====================================================
    col_left, col_right = st.columns([2, 1])
    
    with col_right:
        st.markdown(t("global_mods_header"))
        st.caption(t("global_mods_caption"))
        
        # Inicializar modificaciones en session_state
        if "global_modificaciones" not in st.session_state:
            st.session_state.global_modificaciones = []
        
        # Mostrar modificaciones existentes
        for i, mod in enumerate(st.session_state.global_modificaciones):
            with st.expander(f"📌 {mod['nombre']}", expanded=False):
                st.write(f"{t('global_mod_detail_plant')} {mod['planta']}")
                st.write(f"{t('global_mod_detail_hours')} {mod['horas']} h")
                if st.button(t("global_mod_delete"), key=f"del_mod_{i}"):
                    st.session_state.global_modificaciones.pop(i)
                    st.rerun()
        
        # Añadir nueva modificación
        with st.expander(t("global_mod_expander"), expanded=False):
            new_mod_nombre = st.text_input(t("global_mod_name"), key="new_mod_nombre")
            new_mod_planta = st.selectbox(
                t("global_mod_plant"),
                [p["name"] for _, p in all_data["plants"].iterrows()],
                key="new_mod_planta"
            )
            new_mod_horas = st.number_input(t("global_mod_hours"), min_value=0, value=10, key="new_mod_horas")

            if st.button(t("global_mod_btn"), key="btn_add_mod"):
                if new_mod_nombre.strip():
                    st.session_state.global_modificaciones.append({
                        "nombre": new_mod_nombre.strip(),
                        "planta": new_mod_planta,
                        "horas": new_mod_horas,
                    })
                    st.rerun()
        
        # Resumen de hitos
        st.markdown("---")
        st.markdown(t("global_milestones_header"))
        total_mods = len(st.session_state.global_modificaciones)
        total_horas_mods = sum(m["horas"] for m in st.session_state.global_modificaciones)

        col_h1, col_h2 = st.columns(2)
        with col_h1:
            st.metric(t("global_metric_count"), total_mods)
        with col_h2:
            st.metric(t("global_metric_hours"), f"{total_horas_mods} h")
    
    with col_left:
        # =====================================================
        # 6️⃣ % USO LÍNEAS POR PLANTA
        # =====================================================
        st.markdown(t("global_distrib_header"))
        
        # Gráfico de tarta
        plant_names_chart = [r["plant_name"] for r in global_results if r[h_year_col] > 0]
        plant_caps_chart = [r[h_year_col] for r in global_results if r[h_year_col] > 0]
        
        if plant_caps_chart:
            fig_pie = go.Figure(go.Pie(
                labels=plant_names_chart,
                values=plant_caps_chart,
                hole=0.4,
                textinfo="label+percent",
                marker=dict(colors=px.colors.qualitative.Set2),
            ))
            fig_pie.update_layout(
                title=t("global_chart_distrib").format(esc=escenario),
                height=400,
            )
            st.plotly_chart(fig_pie, use_container_width=True, key="chart_pie_global")
        
        # Tabla de % uso líneas
        st.markdown(t("global_usage_header"))
        
        uso_lineas_rows = []
        total_lineas = sum(r["lines_count"] for r in global_results)
        
        for r in global_results:
            pct_lineas = (r["lines_count"] / total_lineas * 100) if total_lineas > 0 else 0
            uso_lineas_rows.append({
                "PLANTA": r["plant_name"],
                "Nº Líneas": r["lines_count"],
                "% del Total": pct_lineas,
                "Cap. Media/Línea (h/sem)": r[h_sem_col] / r["lines_count"] if r["lines_count"] > 0 else 0,
            })
        
        uso_lineas_df = pd.DataFrame(uso_lineas_rows)
        st.dataframe(
            uso_lineas_df.style.format({
                "% del Total": "{:.1f}%",
                "Cap. Media/Línea (h/sem)": "{:.1f}",
            }),
            use_container_width=True,
            hide_index=True
        )


# =========================================================
# 1) PLANIFICACIÓN
# =========================================================

# Valores D&A actuales en el motor. Definida aquí (nivel de módulo) para que
# esté disponible tanto en Planificación como en la función compute_bench_analysis.
_DA_VALUES = {"SL", "SD", "LL", "LD", "XD", "XL"}

if st.session_state.active_tab == "📊 Planificación":
    st.subheader(t("tab_plan_header"))

    # Session state anidado por planta: line_model[plant_id][line_id], line_demand[plant_id][line_id]
    if "line_model" not in st.session_state:
        st.session_state.line_model = {}
    if "line_demand" not in st.session_state:
        st.session_state.line_demand = {}
    if "line_bench_variant" not in st.session_state:
        st.session_state.line_bench_variant = {}

    _pid = st.session_state["plant_id"]
    if _pid not in st.session_state.line_model:
        st.session_state.line_model[_pid] = {}
    if _pid not in st.session_state.line_demand:
        st.session_state.line_demand[_pid] = {}
    if _pid not in st.session_state.line_bench_variant:
        st.session_state.line_bench_variant[_pid] = {}

    # Una fila por línea con 2 columnas: Selección combinada | Demanda
    # Para D&A el selector combina familia + variante en una sola opción.
    # Para no D&A muestra solo el modelo, sin variantes.
    _VARIANT_LABELS = ["LV", "MV"]  # etiquetas visibles para D&A

    # Lookup de la variante por defecto de cada familia D&A según su regla general
    # (fila da_variant='' en da_bench_type). Usada cuando la variante almacenada está vacía.
    # bench_type LV → "LV";  MV / XL_MV → "MV"
    _da_default_variant: dict[str, str] = {}
    _bmap_plan = load_table("da_bench_type")
    if "plant_id" in _bmap_plan.columns:
        _bmap_plan = _bmap_plan[
            pd.to_numeric(_bmap_plan["plant_id"], errors="coerce") == _pid
        ].copy()
    if not _bmap_plan.empty and "da_variant" in _bmap_plan.columns:
        _gen_rows = _bmap_plan[
            _bmap_plan["da_variant"].astype(str).str.strip() == ""
        ]
        for _, _gr in _gen_rows.iterrows():
            _dv = str(_gr["da_value"]).strip()
            _bt = str(_gr.get("bench_type", "")).strip()
            _da_default_variant[_dv] = "LV" if _bt == "LV" else "MV"

    # Pending scenario load — apply BEFORE widgets are instantiated (Streamlit requirement)
    _pending_key = f"_pending_scenario_{_pid}"
    if _pending_key in st.session_state:
        _pending = st.session_state.pop(_pending_key)
        st.session_state.line_model[_pid] = _pending["line_model"]
        st.session_state.line_demand[_pid] = _pending["line_demand"]
        st.session_state.line_bench_variant[_pid] = _pending["line_bench_variant"]
        if "scenario_id" in _pending:
            st.session_state[f"scenario_select_{_pid}"] = _pending["scenario_id"]
            st.session_state[f"_session_sc_id_{_pid}"] = _pending["scenario_id"]
        if "scenario_name" in _pending:
            st.session_state[f"scenario_name_input_{_pid}"] = _pending["scenario_name"]
        for _lid, _mdl in _pending["line_model"].items():
            _var = _pending["line_bench_variant"].get(_lid, "")
            if _mdl in _DA_VALUES:
                if _var not in _VARIANT_LABELS:
                    _var = _da_default_variant.get(_mdl, _VARIANT_LABELS[0])
                st.session_state[f"sel_combined_{_pid}_{_lid}"] = f"{_mdl} · {_var}"
            elif _mdl:
                st.session_state[f"sel_combined_{_pid}_{_lid}"] = _mdl
        for _lid, _dem in _pending["line_demand"].items():
            st.session_state[f"demand_{_pid}_{_lid}"] = float(_dem)
        _pending_msg = _pending.get("_msg", "plan_load_ok")
        if _pending_msg:
            st.success(t(_pending_msg))

    for nave in sorted(stations_df["nave"].astype(str).str.strip().unique().tolist()):
        st.markdown(f"#### {t('cfg_filter_nave')} {nave}")

        _ch0, _ch1, _ch2 = st.columns([0.7, 1.8, 1.0])
        _ch0.caption(t("cfg_line_label"))
        _ch1.caption(t("plan_col_model"))
        _ch2.caption(t("plan_col_demand"))

        nave_line_ids = sorted(
            stations_df.loc[stations_df["nave"] == nave, "line_id"]
            .astype(str)
            .str.strip()
            .unique()
            .tolist()
        )

        for line_id in nave_line_ids:
            allowed = allowed_by_line.get(line_id, [])
            if not allowed:
                _cn0, _cn1, _cn2 = st.columns([0.7, 1.8, 1.0])
                _cn0.markdown(f"**{line_id}**")
                _cn1.caption(t("plan_no_models"))
                continue

            # Construir lista de opciones combinadas
            # D&A → "LL · LV", "LL · MV"
            # No D&A → "PT0163"
            _combined_opts = []
            for _mod in allowed:
                if _mod in _DA_VALUES:
                    for _vlbl in _VARIANT_LABELS:
                        _combined_opts.append(f"{_mod} · {_vlbl}")
                else:
                    _combined_opts.append(_mod)

            # Recuperar selección actual y traducirla a opción combinada
            _cur_model = (
                st.session_state.line_model.get(_pid, {}).get(line_id)
                or allowed[0]
            )
            if _cur_model not in allowed:
                _cur_model = allowed[0]

            _cur_variant = st.session_state.line_bench_variant.get(_pid, {}).get(line_id, "")

            if _cur_model in _DA_VALUES:
                if _cur_variant in ("LV", "MV"):
                    _vlbl_cur = _cur_variant
                else:
                    # Sin variante explícita: usar la regla general real de la familia
                    # (da_bench_type donde da_variant=''); si no está configurada, LV
                    _vlbl_cur = _da_default_variant.get(_cur_model, _VARIANT_LABELS[0])
                _cur_opt = f"{_cur_model} · {_vlbl_cur}"
            else:
                _cur_opt = _cur_model

            if _cur_opt not in _combined_opts:
                _cur_opt = _combined_opts[0]

            _sel_key = f"sel_combined_{_pid}_{line_id}"
            if _sel_key not in st.session_state:
                st.session_state[_sel_key] = _cur_opt

            _c0, _c1, _c2 = st.columns([0.7, 1.8, 1.0])
            _c0.markdown(f"**{line_id}**")

            with _c1:
                _sel = st.selectbox(
                    f"Selección ({line_id})",
                    options=_combined_opts,
                    key=_sel_key,
                    label_visibility="collapsed",
                    help="Para familias D&A elige también la variante de prueba: LV o MV.",
                )

            # Descomponer la selección en modelo + variante
            if " · " in _sel:
                _m, _vlbl = _sel.split(" · ", 1)
                _variant = _vlbl  # always LV or MV now
            else:
                _m = _sel
                _variant = ""
                st.session_state.line_bench_variant.get(_pid, {}).pop(line_id, None)

            st.session_state.line_model[_pid][line_id] = _m
            if _m in _DA_VALUES:
                st.session_state.line_bench_variant[_pid][line_id] = _variant

            with _c2:
                # Garantizar que la clave existe antes de renderizar para evitar el warning
                # "widget created with default value but also set via Session State API"
                _dem_key = f"demand_{_pid}_{line_id}"
                if _dem_key not in st.session_state:
                    st.session_state[_dem_key] = float(
                        st.session_state.line_demand.get(_pid, {}).get(line_id, 0.0)
                    )
                d = st.number_input(
                    f"Demanda ({line_id} – {_m})",
                    min_value=0.0,
                    step=1.0,
                    key=_dem_key,
                    label_visibility="collapsed",
                )
                st.session_state.line_demand[_pid][line_id] = d

    # --- Scenario management ---
    st.divider()
    _sc_sel_id = None
    _sc_name_map: dict = {}
    if _has_db():
        _sc_list = list_scenarios(_pid)
        if _sc_list:
            _sc_name_map = {s["id"]: s["name"] for s in _sc_list}
            st.markdown(f"**{t('plan_scenarios_header')}**")
            _sc_label_map = {
                s["id"]: (f"{s['name']}  {t('plan_active_marker')}" if s["is_active"] else s["name"])
                for s in _sc_list
            }
            _sc_is_active_map = {s["id"]: s["is_active"] for s in _sc_list}

            # Make name map available to the on_change callback (read by _on_scenario_select_change)
            st.session_state[f"_sc_name_map_{_pid}"] = _sc_name_map

            # Apply deferred selection (set by delete/duplicate/save-as-new on previous run)
            _defer_sel_key = f"_deferred_sc_select_{_pid}"
            if _defer_sel_key in st.session_state:
                _deferred_id = st.session_state.pop(_defer_sel_key)
                st.session_state[f"scenario_select_{_pid}"] = _deferred_id

            # Apply deferred name (set by duplicate/save-as-new/save-changes on previous run)
            _defer_name_key = f"_deferred_sc_name_{_pid}"
            if _defer_name_key in st.session_state:
                st.session_state[f"scenario_name_input_{_pid}"] = st.session_state.pop(_defer_name_key)
            elif f"scenario_name_input_{_pid}" not in st.session_state:
                # First render only: initialize field from selected scenario's name
                _init_sel = st.session_state.get(f"scenario_select_{_pid}")
                if _init_sel in _sc_name_map:
                    st.session_state[f"scenario_name_input_{_pid}"] = _sc_name_map[_init_sel]
                else:
                    st.session_state[f"scenario_name_input_{_pid}"] = t("plan_save_name_default")

            # Tab-return sync: align dropdown and name to _session_sc_id_ only when
            # coming back from another tab (not on every render — avoids dropdown rebound).
            if _prev_active_tab != "📊 Planificación":
                _trusted_sc_id = st.session_state.get(f"_session_sc_id_{_pid}")
                if _trusted_sc_id in _sc_label_map:
                    st.session_state[f"scenario_select_{_pid}"] = _trusted_sc_id
                    st.session_state[f"scenario_name_input_{_pid}"] = _sc_name_map.get(_trusted_sc_id, "")

            _sce1, _sce2, _sce3, _sce4, _sce5 = st.columns([3, 1, 1, 1, 1], vertical_alignment="bottom")
            _sc_sel_id = _sce1.selectbox(
                t("plan_scenarios_header"),
                options=list(_sc_label_map.keys()),
                format_func=lambda sid: _sc_label_map[sid],
                key=f"scenario_select_{_pid}",
                label_visibility="collapsed",
                on_change=_on_scenario_select_change,
                args=(_pid,),
            )
            if _sce2.button(t("plan_load_btn"), use_container_width=True, key="btn_load_sc"):
                _loaded = load_scenario_by_id(_sc_sel_id)
                if _loaded:
                    st.session_state[f"_pending_scenario_{_pid}"] = _loaded
                    st.rerun()
            if _sce3.button(t("plan_activate_btn"), use_container_width=True, key="btn_activate_sc"):
                activate_scenario(_pid, _sc_sel_id)
                _activated = load_scenario_by_id(_sc_sel_id)
                if _activated:
                    _activated["_msg"] = "plan_activate_ok"
                    st.session_state[f"_pending_scenario_{_pid}"] = _activated
                st.rerun()
            if _sce4.button(t("plan_duplicate_btn"), use_container_width=True, key="btn_duplicate_sc"):
                _dup_id, _dup_name = duplicate_scenario(_sc_sel_id, _pid)
                if _dup_id is not None:
                    st.session_state[f"_deferred_sc_select_{_pid}"] = _dup_id
                    st.session_state[f"_deferred_sc_name_{_pid}"] = _dup_name
                st.rerun()
            if _sce5.button(t("plan_delete_btn"), use_container_width=True, key="btn_delete_sc"):
                if _sc_is_active_map.get(_sc_sel_id, False):
                    st.warning(t("plan_delete_blocked"))
                else:
                    delete_scenario(_sc_sel_id)
                    _remaining = [s for s in _sc_list if s["id"] != _sc_sel_id]
                    if _remaining:
                        _next_id = next((s["id"] for s in _remaining if s["is_active"]), _remaining[0]["id"])
                        st.session_state[f"_deferred_sc_select_{_pid}"] = _next_id
                    sc_key = f"scenario_select_{_pid}"
                    if st.session_state.get(sc_key) == _sc_sel_id:
                        st.session_state.pop(sc_key, None)
                    st.rerun()
        else:
            st.caption(t("plan_no_scenarios"))

    # --- Save ---
    _sc_col1, _sc_col2, _sc_col3 = st.columns([2, 1, 1], vertical_alignment="bottom")
    _sc_name = _sc_col1.text_input(
        t("plan_save_name_label"),
        key=f"scenario_name_input_{_pid}",
    )
    if _sc_col2.button(t("plan_save_changes_btn"), use_container_width=True):
        if not _has_db():
            st.info(t("plan_save_no_db"))
        elif _sc_sel_id is None:
            st.warning(t("plan_save_changes_no_sel"))
        else:
            _name_to_save = _sc_name.strip() or t("plan_save_name_default")
            update_scenario_lines(
                _sc_sel_id,
                st.session_state.line_model.get(_pid, {}),
                st.session_state.line_demand.get(_pid, {}),
                st.session_state.line_bench_variant.get(_pid, {}),
            )
            rename_scenario(_sc_sel_id, _name_to_save)
            st.session_state[f"_deferred_sc_name_{_pid}"] = _name_to_save
            # Pop the selectbox key so Streamlit treats the next render as a fresh
            # initialization — this forces the frontend to re-render the header label
            # with the new name from format_func, even though the selected id is the same.
            st.session_state.pop(f"scenario_select_{_pid}", None)
            st.session_state[f"_deferred_sc_select_{_pid}"] = _sc_sel_id
            st.toast(t("plan_save_changes_ok"))
            st.rerun()
    if _sc_col3.button(t("plan_save_new_btn"), use_container_width=True):
        if not _has_db():
            st.info(t("plan_save_no_db"))
        else:
            _used_name = _sc_name.strip() or t("plan_save_name_default")
            _new_id = create_scenario_inactive(
                _pid,
                _used_name,
                st.session_state.line_model.get(_pid, {}),
                st.session_state.line_demand.get(_pid, {}),
                st.session_state.line_bench_variant.get(_pid, {}),
            )
            if _new_id is not None:
                st.session_state[f"_deferred_sc_select_{_pid}"] = _new_id
                st.session_state[f"_deferred_sc_name_{_pid}"] = _used_name
            st.success(t("plan_save_ok"))
            st.rerun()


# =========================================================
# 2) CONFIGURACIÓN (POWER USER)
# =========================================================
if st.session_state.active_tab == "⚙️ Configuración (Power User)":
    st.subheader(t("tab_cfg_header"))
    st.caption(t("tab_cfg_caption"))

    # --- A) Gestión de modelos (checkbox)
    st.markdown(t("cfg_models_header"))

    models_editor = models_df.copy()
    models_editor["active"] = models_editor["active"].astype(int).clip(0, 1).astype(bool)

    _mf1, _mf2 = st.columns(2)
    _filter_m_name = _mf1.text_input(t("cfg_filter_model"), key="filter_models_name", placeholder="")
    _status_opts = [t("cfg_status_all"), t("cfg_status_active"), t("cfg_status_inactive")]
    if st.session_state.get("filter_models_status", "") not in _status_opts:
        st.session_state["filter_models_status"] = t("cfg_status_all")
    _filter_m_status = _mf2.selectbox(t("cfg_filter_status"), _status_opts, key="filter_models_status")

    _mask_m = pd.Series([True] * len(models_editor), index=models_editor.index)
    if _filter_m_name:
        _mask_m &= models_editor["model"].astype(str).str.contains(_filter_m_name, case=False, na=False)
    if _filter_m_status == t("cfg_status_active"):
        _mask_m &= models_editor["active"] == True
    elif _filter_m_status == t("cfg_status_inactive"):
        _mask_m &= models_editor["active"] == False

    _models_visible = models_editor[_mask_m].copy()
    _models_hidden  = models_editor[~_mask_m].copy()
    if _filter_m_name or _filter_m_status != t("cfg_status_all"):
        st.caption(t("cfg_showing_rows").format(shown=len(_models_visible), total=len(models_editor)))

    edited_models = st.data_editor(
        _models_visible,
        use_container_width=True,
        num_rows="dynamic",
        column_config={
            "active": st.column_config.CheckboxColumn("active", help="Modelo activo (aparece en la app).")
        }
    )

    if st.button(t("btn_save_models")):
        out = pd.concat([_models_hidden, edited_models], ignore_index=True)
        out = out.reset_index(drop=True)
        out["model"] = out["model"].astype(str).str.strip()
        out["description"] = out["description"].astype(str).str.strip()
        out["active"] = out["active"].astype(bool).astype(int)
        out["plant_id"] = plant_id
        save_table(out, "models")
        st.session_state["models_saved"] = True

    if st.session_state.get("models_saved"):
        st.success(t("msg_models_saved"))
        st.session_state["models_saved"] = False

    st.divider()

    # --- B) Tiempos por modelo y proceso
    st.markdown(t("cfg_times_header"))
    st.info(t("cfg_times_info"))

    with st.expander("📐 ¿Dos métricas de tiempo distintas? Ejemplo real: PT0163 · proceso ML"):
        st.markdown(
            """
**Capacidad en unidades — lógica de cuello de botella**

El motor calcula `cycle_time_real = max(machine_time, labor_time / operarios)` por proceso.
Ese valor es el tiempo mínimo entre dos unidades consecutivas: lo que tarda la máquina en liberar la pieza,
o lo que tarda cada operario individualmente, lo que sea mayor.

Proceso **ML** en PT0163 — 5 estaciones, 2 operarios/estación, proceso manual:

| | Valor |
|---|---|
| machine_time | 0 h (proceso manual, sin ciclo de máquina) |
| labor_time | 48,03 HH |
| Tiempo por operario | 48,03 / 2 = 24,015 h |
| cycle_time_real | max(0 ; 24,015) = **24,015 h** |
| Capacidad ML | (43 h/sem × 5 estaciones) / 24,015 ≈ **8,95 uds/sem** |

---

**Horas-hombre visibles — contenido de trabajo, no ritmo**

Tener 2 operarios no reduce las horas: las distribuye.
Producir 1 unidad en ML cuesta **48,03 HH** — lo ejecutan 2 personas en paralelo,
pero el contenido total de trabajo no desaparece.

Por eso las horas visibles usan `cycle_time` directamente, sin dividir entre operarios:

`h/sem = capacidad (uds/sem) × Σ cycle_time por proceso`

Una cifra mide **ritmo de salida** (cuello de botella).
La otra mide **carga real de planta** (horas-hombre totales comprometidas).
Dividir cycle_time entre operarios para calcular horas sería un error: reduciría artificialmente
la carga visible sin que el trabajo desaparezca de la planta.
"""
        )

    _fc1, _fc2 = st.columns(2)
    _filter_t_model = _fc1.text_input(t("cfg_filter_model"), key="filter_times_model", placeholder="ej. PT0163")
    _filter_t_proc  = _fc2.text_input(t("cfg_filter_process"), key="filter_times_proc", placeholder="ej. ML")
    _mask_t = pd.Series([True] * len(times_df), index=times_df.index)
    if _filter_t_model:
        _mask_t &= times_df["model"].astype(str).str.contains(_filter_t_model, case=False, na=False)
    if _filter_t_proc:
        _mask_t &= times_df["process"].astype(str).str.contains(_filter_t_proc, case=False, na=False)
    _times_visible = times_df[_mask_t].copy()
    _times_hidden  = times_df[~_mask_t].copy()
    if _filter_t_model or _filter_t_proc:
        st.caption(t("cfg_showing_rows").format(shown=len(_times_visible), total=len(times_df)))

    edited_times = st.data_editor(
        _times_visible,
        use_container_width=True,
        num_rows="dynamic",
        column_config={
            "cycle_time": st.column_config.NumberColumn(
                "cycle_time",
                help="Tiempo de ciclo legacy (HH/ud). Se mantiene por compatibilidad.",
                min_value=0.0,
                step=0.01,
                format="%.2f"
            ),
            "machine_time": st.column_config.NumberColumn(
                "machine_time",
                help="Tiempo automático fijo no reducible (HH/ud). Puede ser 0 en procesos manuales.",
                min_value=0.0,
                step=0.01,
                format="%.2f"
            ),
            "labor_time": st.column_config.NumberColumn(
                "labor_time",
                help="Horas-hombre secuenciales necesarias por unidad (HH/ud).",
                min_value=0.0,
                step=0.01,
                format="%.2f"
            ),
        }
    )

    if st.button(t("btn_save_times")):
        out = pd.concat([_times_hidden, edited_times], ignore_index=True)
        out["model"] = out["model"].astype(str).str.strip()
        out["process"] = out["process"].astype(str).str.strip()
        out["cycle_time"] = pd.to_numeric(out["cycle_time"], errors="coerce").fillna(0.0)

        # machine_time y labor_time: si vienen vacíos, convertir a 0.0
        # IMPORTANTE: machine_time = 0 es válido en procesos manuales
        if "machine_time" in out.columns:
            out["machine_time"] = pd.to_numeric(out["machine_time"], errors="coerce").fillna(0.0)
        else:
            out["machine_time"] = 0.0

        if "labor_time" in out.columns:
            out["labor_time"] = pd.to_numeric(out["labor_time"], errors="coerce").fillna(0.0)
        else:
            out["labor_time"] = 0.0

        # evitar duplicados modelo-proceso
        out = out.drop_duplicates(subset=["model", "process"])
        out["plant_id"] = plant_id

        # Validación mínima: avisar si algún proceso tiene ambos tiempos a 0
        if "machine_time" in out.columns and "labor_time" in out.columns:
            bad = out[(out["machine_time"] == 0.0) & (out["labor_time"] == 0.0)]
            if not bad.empty:
                st.session_state["times_warning"] = (
                    f"⚠️ {len(bad)} proceso(s) con machine_time=0 y labor_time=0. "
                    "Esos procesos tendrán capacidad calculada = 0."
                )

        save_table(out, "models_process_times")

        st.session_state["times_saved"] = True

    if st.session_state.get("times_warning"):
        st.warning(st.session_state["times_warning"])
        st.session_state["times_warning"] = None

    if st.session_state.get("times_saved"):
        st.success(t("msg_times_saved"))
        st.session_state["times_saved"] = False

    st.divider()

    # --- C) Estaciones / operarios por línea y proceso
    st.markdown(t("cfg_stations_header"))

    _sf1, _sf2, _sf3 = st.columns(3)
    _nave_opts = [t("cfg_filter_da_all")] + sorted(stations_df["nave"].astype(str).str.strip().unique().tolist())
    if st.session_state.get("filter_stations_nave", "") not in _nave_opts:
        st.session_state["filter_stations_nave"] = t("cfg_filter_da_all")
    _filter_st_nave = _sf1.selectbox(t("cfg_filter_nave"), _nave_opts, key="filter_stations_nave")

    _line_opts = [t("cfg_filter_da_all")] + sorted(stations_df["line"].astype(str).str.strip().unique().tolist())
    if st.session_state.get("filter_stations_line", "") not in _line_opts:
        st.session_state["filter_stations_line"] = t("cfg_filter_da_all")
    _filter_st_line = _sf2.selectbox(t("cfg_line_label"), _line_opts, key="filter_stations_line")

    _proc_opts = [t("cfg_filter_da_all")] + sorted(stations_df["process"].astype(str).str.strip().unique().tolist())
    if st.session_state.get("filter_stations_proc", "") not in _proc_opts:
        st.session_state["filter_stations_proc"] = t("cfg_filter_da_all")
    _filter_st_proc = _sf3.selectbox(t("cfg_filter_proc_label"), _proc_opts, key="filter_stations_proc")

    _mask_st = pd.Series([True] * len(stations_df), index=stations_df.index)
    _all_token = t("cfg_filter_da_all")
    if _filter_st_nave != _all_token:
        _mask_st &= stations_df["nave"].astype(str).str.strip() == _filter_st_nave
    if _filter_st_line != _all_token:
        _mask_st &= stations_df["line"].astype(str).str.strip() == _filter_st_line
    if _filter_st_proc != _all_token:
        _mask_st &= stations_df["process"].astype(str).str.strip() == _filter_st_proc

    _stations_visible = stations_df[_mask_st].copy()
    _stations_hidden  = stations_df[~_mask_st].copy()
    if any(f != _all_token for f in [_filter_st_nave, _filter_st_line, _filter_st_proc]):
        st.caption(t("cfg_showing_rows").format(shown=len(_stations_visible), total=len(stations_df)))

    edited_stations = st.data_editor(
        _stations_visible,
        use_container_width=True,
        num_rows="dynamic",
        column_config={
            "stations": st.column_config.NumberColumn("stations", min_value=0.0, step=0.1, format="%.2f"),
            "operators_per_station": st.column_config.NumberColumn("operators_per_station", min_value=0.0, step=0.1, format="%.2f"),
        }
    )

    if st.button(t("btn_save_stations")):
        out = pd.concat([_stations_hidden, edited_stations], ignore_index=True)
        out = out.reset_index(drop=True)
        out["line"] = out["line"].astype(str).str.strip()
        out["nave"] = out["nave"].astype(str).str.strip()
        out["process"] = out["process"].astype(str).str.strip()
        out["stations"] = pd.to_numeric(out["stations"], errors="coerce").fillna(0.0)
        out["operators_per_station"] = pd.to_numeric(out["operators_per_station"], errors="coerce").fillna(0.0)
        # Reconstruir line_id para mantener coherencia
        out["line_id"] = out["nave"] + "-" + out["line"]
        out["plant_id"] = plant_id

        # Validación mínima: avisar si hay filas con stations=0
        zero_st = out[out["stations"] == 0.0]
        if not zero_st.empty:
            st.session_state["stations_warning"] = (
                f"⚠️ {len(zero_st)} fila(s) con stations=0. "
                "Esas filas serán ignoradas en el cálculo de capacidad."
            )

        save_table(out, "lines_process_stations")

        st.session_state["stations_saved"] = True

    if st.session_state.get("stations_warning"):
        st.warning(st.session_state["stations_warning"])
        st.session_state["stations_warning"] = None

    if st.session_state.get("stations_saved"):
        st.success(t("msg_stations_saved"))
        st.session_state["stations_saved"] = False

    st.divider()

    
    # --- D) Compatibilidad modelo ↔ línea (checkbox)
    st.markdown(t("cfg_compat_header"))

    # Sólo mostramos modelos existentes en models.csv (da igual activos o no: compat se define aquí)
    all_models = sorted(models_df["model"].astype(str).str.strip().unique().tolist())
    all_line_ids = sorted(stations_df["line_id"].astype(str).str.strip().unique().tolist())

    # Inicializar estado de expanders por planta (persistente entre reruns)
    for _lid in all_line_ids:
        _ek = f"exp_compat_{plant_id}_{_lid}"
        if _ek not in st.session_state:
            st.session_state[_ek] = True

    # Botones de control global
    _gc1, _gc2, _gc3 = st.columns([1, 1, 5])
    if _gc1.button(t("cfg_compat_expand_all"), key="btn_expand_all_compat"):
        for _lid in all_line_ids:
            st.session_state[f"exp_compat_{plant_id}_{_lid}"] = True
        st.rerun()
    if _gc2.button(t("cfg_compat_collapse_all"), key="btn_collapse_all_compat"):
        for _lid in all_line_ids:
            st.session_state[f"exp_compat_{plant_id}_{_lid}"] = False
        st.rerun()

    # Matriz editable por línea real (nave + línea)
    edited_rows = []

    for line_id in all_line_ids:

        parts = line_id.split("-", 1)

        if len(parts) == 2:
            nave, base_line = parts
        else:
            nave = "N1"
            base_line = parts[0]

        _expanded_now = st.session_state.get(f"exp_compat_{plant_id}_{line_id}", True)
        with st.expander(f"{t('cfg_line_label')} {line_id}", expanded=_expanded_now):
            cols = st.columns(5)
            for i, m in enumerate(all_models):
                current = compat_df[
                    (compat_df["line"] == base_line) &
                    (compat_df["model"] == m) &
                    (compat_df["nave"] == nave)
                ]
                cur_val = 0
                if not current.empty:
                    cur_val = int(current.iloc[0]["compatible"])

                checked = cols[i % 5].checkbox(
                    m,
                    value=bool(cur_val),
                    key=f"compat_{plant_id}_{line_id}_{m}"
                )

                edited_rows.append({
                    "nave": nave,
                    "line": base_line,
                    "model": m,
                    "compatible": 1 if checked else 0
                })

    if edited_rows:
        import io as _io
        from datetime import date as _date
        from openpyxl import Workbook as _Workbook
        from openpyxl.styles import Font as _Font, PatternFill as _PFill, Alignment as _Align, Border as _Border, Side as _Side

        def _build_compat_xlsx(rows: list, p_id: int, p_name: str) -> bytes:
            _df_all = pd.DataFrame(rows).assign(plant_id=p_id)
            _df_all["line_id"] = _df_all["nave"] + "-" + _df_all["line"]

            _all_lines = sorted(_df_all["line_id"].unique().tolist())
            _all_mods  = sorted(_df_all["model"].unique().tolist())

            _wb = _Workbook()

            # ---------- HOJA 1: MATRIZ ----------
            _ws_m = _wb.active
            _ws_m.title = "MATRIZ"

            _hdr_fill  = _PFill("solid", fgColor="FF1F4E79")
            _ok_fill   = _PFill("solid", fgColor="FF70AD47")
            _no_fill   = _PFill("solid", fgColor="FFFFFFFF")
            _row_fill  = _PFill("solid", fgColor="FFD6E4F0")
            _bold_w    = _Font(bold=True, color="FFFFFFFF")
            _bold_b    = _Font(bold=True)
            _center    = _Align(horizontal="center", vertical="center")
            _thin_side = _Side(style="thin", color="FFBFBFBF")
            _thin_brd  = _Border(left=_thin_side, right=_thin_side, top=_thin_side, bottom=_thin_side)

            # fila 0: título
            _ws_m.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(_all_mods))
            _tc = _ws_m.cell(1, 1, f"Compatibilidad modelo ↔ línea  |  Planta: {p_name}  |  {_date.today().strftime('%d/%m/%Y')}")
            _tc.font = _Font(bold=True, size=13, color="FFFFFFFF")
            _tc.fill = _PFill("solid", fgColor="FF1F4E79")
            _tc.alignment = _center
            _ws_m.row_dimensions[1].height = 22

            # fila 1: cabecera columnas (línea | modelo1 | modelo2 …)
            _ws_m.cell(2, 1, "Línea").font = _bold_w
            _ws_m.cell(2, 1).fill = _hdr_fill
            _ws_m.cell(2, 1).alignment = _center
            _ws_m.cell(2, 1).border = _thin_brd
            for _ci, _mod in enumerate(_all_mods, start=2):
                _c = _ws_m.cell(2, _ci, _mod)
                _c.font = _bold_w
                _c.fill = _hdr_fill
                _c.alignment = _center
                _c.border = _thin_brd
            _ws_m.row_dimensions[2].height = 18

            # lookup rápido
            _compat_set = set(
                _df_all.loc[_df_all["compatible"] == 1, ["line_id", "model"]]
                .apply(lambda r: (r["line_id"], r["model"]), axis=1)
            )

            for _ri, _lid in enumerate(_all_lines, start=3):
                _row_bg = _row_fill if _ri % 2 == 1 else _PFill("solid", fgColor="FFFFFFFF")
                _lc = _ws_m.cell(_ri, 1, _lid)
                _lc.font = _bold_b
                _lc.fill = _row_bg
                _lc.alignment = _Align(horizontal="left", vertical="center")
                _lc.border = _thin_brd
                for _ci, _mod in enumerate(_all_mods, start=2):
                    _ok = (_lid, _mod) in _compat_set
                    _vc = _ws_m.cell(_ri, _ci, "✓" if _ok else "")
                    _vc.fill = _ok_fill if _ok else _row_bg
                    _vc.alignment = _center
                    _vc.border = _thin_brd
                    if _ok:
                        _vc.font = _Font(bold=True, color="FFFFFFFF")

            # anchos
            _ws_m.column_dimensions["A"].width = 18
            for _ci in range(2, 2 + len(_all_mods)):
                _ws_m.column_dimensions[
                    _ws_m.cell(2, _ci).column_letter
                ].width = max(len(_all_mods[_ci - 2]) + 2, 7)
            _ws_m.freeze_panes = "B3"

            # ---------- HOJA 2: DETALLE ----------
            _ws_d = _wb.create_sheet("DETALLE")
            _df_det = (
                _df_all[_df_all["compatible"] == 1]
                [["plant_id", "nave", "line", "line_id", "model"]]
                .sort_values(["nave", "line", "model"])
                .reset_index(drop=True)
            )

            _det_cols = ["plant_id", "nave", "line", "line_id", "model"]
            _det_hdr_fills = {
                1: _PFill("solid", fgColor="FF1F4E79"),
                2: _PFill("solid", fgColor="FF1F4E79"),
                3: _PFill("solid", fgColor="FF1F4E79"),
                4: _PFill("solid", fgColor="FF1F4E79"),
                5: _PFill("solid", fgColor="FF1F4E79"),
            }

            for _ci, _col in enumerate(_det_cols, start=1):
                _hc = _ws_d.cell(1, _ci, _col)
                _hc.font = _bold_w
                _hc.fill = _det_hdr_fills[_ci]
                _hc.alignment = _center
                _hc.border = _thin_brd

            for _ri, _row in _df_det.iterrows():
                _bg = _row_fill if _ri % 2 == 0 else _PFill("solid", fgColor="FFFFFFFF")
                for _ci, _col in enumerate(_det_cols, start=1):
                    _dc = _ws_d.cell(_ri + 2, _ci, str(_row[_col]))
                    _dc.alignment = _Align(horizontal="left", vertical="center")
                    _dc.fill = _bg
                    _dc.border = _thin_brd

            _ws_d.freeze_panes = "A2"
            _ws_d.auto_filter.ref = _ws_d.dimensions
            for _ci, _col in enumerate(_det_cols, start=1):
                _ws_d.column_dimensions[
                    _ws_d.cell(1, _ci).column_letter
                ].width = max(len(_col) + 4, 12)

            _buf = _io.BytesIO()
            _wb.save(_buf)
            return _buf.getvalue()

        _plant_name = selected_plant_name if isinstance(selected_plant_name, str) else str(plant_id)
        _today_str  = _date.today().strftime("%Y%m%d")
        _xlsx_bytes = _build_compat_xlsx(edited_rows, plant_id, _plant_name)
        st.download_button(
            label=t("cfg_compat_export"),
            data=_xlsx_bytes,
            file_name=f"compatibilidad_{_plant_name}_{_today_str}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="btn_download_compat",
        )

    if st.button(t("btn_save_compat")):
        out = pd.DataFrame(edited_rows)
        out["plant_id"] = plant_id
        save_table(out, "compatibility")
        st.session_state["compat_saved"] = True

    if st.session_state.get("compat_saved"):
        st.success(t("msg_compat_saved"))
        st.session_state["compat_saved"] = False

    # --- E) Bancos de prueba ----------------------------
    st.divider()
    st.markdown(t("cfg_benches_section"))
    st.caption(t("cfg_benches_caption"))

    # E1) Bancos disponibles por tipo
    st.markdown(t("cfg_benches_by_type"))
    _bcfg_ed = load_table("test_bench_config")
    if "plant_id" in _bcfg_ed.columns:
        _bcfg_ed = _bcfg_ed[
            pd.to_numeric(_bcfg_ed["plant_id"], errors="coerce") == plant_id
        ].copy()
    else:
        _bcfg_ed = pd.DataFrame(columns=["bench_type", "quantity", "hours_per_unit_override"])

    _bcfg_show = _bcfg_ed[
        [c for c in ["bench_type", "quantity", "hours_per_unit_override"] if c in _bcfg_ed.columns]
    ].copy()

    edited_bench_cfg = st.data_editor(
        _bcfg_show,
        use_container_width=True,
        num_rows="dynamic",
        column_config={
            "bench_type": st.column_config.TextColumn(
                t("cfg_col_bench_type"), help="LV, MV o XL_MV"
            ),
            "quantity": st.column_config.NumberColumn(
                t("cfg_col_bench_qty"), min_value=0, step=1, format="%d"
            ),
            "hours_per_unit_override": st.column_config.NumberColumn(
                "Tiempo manual h/ud (solo si PARA no existe)",
                min_value=0.0, step=0.01, format="%.2f",
                help="Dejar vacío para usar automáticamente el proceso PARA del motor."
            ),
        },
    )

    if st.button(t("btn_save_benches")):
        _out = edited_bench_cfg.copy()
        _out["plant_id"] = plant_id
        save_table(_out, "test_bench_config")
        st.session_state["bench_cfg_saved"] = True
    if st.session_state.get("bench_cfg_saved"):
        st.success(t("msg_benches_saved"))
        st.session_state["bench_cfg_saved"] = False

    # E2) Asignación valor D&A → tipo de banco
    st.markdown(t("cfg_da_header"))
    st.caption(t("cfg_da_caption"))

    _bmap_ed = load_table("da_bench_type")
    if "plant_id" in _bmap_ed.columns:
        _bmap_ed = _bmap_ed[
            pd.to_numeric(_bmap_ed["plant_id"], errors="coerce") == plant_id
        ].copy()
    else:
        _bmap_ed = pd.DataFrame(columns=["da_value", "da_variant", "bench_type"])

    # Asegurar que da_variant existe aunque la tabla venga sin esa columna (pre-migración)
    if "da_variant" not in _bmap_ed.columns:
        _bmap_ed["da_variant"] = ""

    _bmap_show = _bmap_ed[
        [c for c in ["da_value", "da_variant", "bench_type"] if c in _bmap_ed.columns]
    ].copy()

    _da_opts = [t("cfg_filter_da_all"), "SL", "SD", "LL", "LD", "XD", "XL"]
    if st.session_state.get("filter_da_bench", "") not in _da_opts:
        st.session_state["filter_da_bench"] = t("cfg_filter_da_all")
    _filter_da = st.selectbox(
        t("cfg_filter_da"),
        options=_da_opts,
        key="filter_da_bench",
    )
    if _filter_da in _DA_VALUES:
        _bmap_visible = _bmap_show[_bmap_show["da_value"] == _filter_da].copy()
        _bmap_hidden  = _bmap_show[_bmap_show["da_value"] != _filter_da].copy()
        st.caption(t("cfg_showing_rows").format(shown=len(_bmap_visible), total=len(_bmap_show)))
    else:
        _bmap_visible = _bmap_show.copy()
        _bmap_hidden  = _bmap_show.iloc[0:0].copy()

    edited_bench_map = st.data_editor(
        _bmap_visible,
        use_container_width=True,
        num_rows="dynamic",
        column_config={
            "da_value": st.column_config.SelectboxColumn(
                t("cfg_col_da_value"),
                options=["SL", "SD", "LL", "LD", "XD", "XL"],
                help="Familia D&A tal como existe hoy en el motor.",
                required=True,
            ),
            "da_variant": st.column_config.SelectboxColumn(
                t("cfg_col_da_variant"),
                options=["", "LV", "MV"],
                help=(
                    "Vacío = regla general de la familia. "
                    "'LV' o 'MV' = regla específica para esa variante."
                ),
                required=False,
            ),
            "bench_type": st.column_config.SelectboxColumn(
                t("cfg_col_bench_apply"),
                options=["LV", "MV", "XL_MV"],
                help="Banco que se asigna a esta combinación.",
                required=True,
            ),
        },
    )

    if st.button(t("btn_save_da")):
        _out = pd.concat([_bmap_hidden, edited_bench_map], ignore_index=True)
        _out["plant_id"] = plant_id
        # Normalizar da_variant: None → ''
        if "da_variant" in _out.columns:
            _out["da_variant"] = _out["da_variant"].fillna("").astype(str).str.strip()
        save_table(_out, "da_bench_type")
        st.session_state["bench_map_saved"] = True
    if st.session_state.get("bench_map_saved"):
        st.success(t("msg_da_saved"))
        st.session_state["bench_map_saved"] = False

# =========================================================
# 3) RESULTADOS
# =========================================================

# =========================================================
# BANCOS DE PRUEBA — fase 1 informativa
# =========================================================

def compute_bench_analysis(
    line_id: str,
    da_value: str,
    demand_week: float,
    bench_cfg: "pd.DataFrame",
    bench_map: "pd.DataFrame",
    times_df_b: "pd.DataFrame",
    stations_df_b: "pd.DataFrame",
    hours_eff: float,
    da_variant: str = "",
) -> dict:
    """
    Capa informativa de bancos de prueba — fase 2.

    Calcula si los bancos de prueba disponibles podrían limitar la capacidad
    de una línea D&A. NO modifica ningún valor oficial del motor (capacidad,
    saturación, déficit, cuello de botella).

    da_variant (fase 2): variante de prueba informada por el planificador
    para esta línea ('LV', 'MV', o '' para usar la regla general).
    Si da_variant es '' el comportamiento es idéntico a la fase 1.

    Búsqueda en bench_map (dos pasos):
      1. Coincidencia exacta: da_value + da_variant
      2. Si no existe: regla general da_value + da_variant=''
      3. Si tampoco: "Datos insuficientes"

    Devuelve un dict con:
        da_value       — valor tal como existe en el motor
        da_variant     — variante informada ('' si no se especificó)
        test_type      — tipo de prueba (LV / MV / No aplica)
        bench_type     — banco aplicable (LV / MV / XL_MV / No aplica)
        bench_capacity — UDS/SEM máximas que soportan los bancos (0 si no aplica)
        tiempo_para    — horas/ud del proceso PARA calculadas (0.0 si no aplica o sin datos)
        observation    — texto del aviso para el usuario
    """
    da_variant = (da_variant or "").strip()

    # 1. ¿Es D&A?
    if da_value not in _DA_VALUES:
        return {
            "da_value": da_value,
            "da_variant": da_variant,
            "test_type": "No aplica",
            "bench_type": "No aplica",
            "bench_capacity": 0.0,
            "tiempo_para": 0.0,
            "observation": "No aplica",
        }

    # 2. Buscar asignación de banco — dos pasos
    # Si la tabla no tiene la columna da_variant aún (pre-migración), tratar
    # todas las filas como da_variant='', compatibilidad hacia atrás.
    _has_variant_col = "da_variant" in bench_map.columns

    row_map = pd.DataFrame()
    if da_variant and _has_variant_col:
        # Paso 1: coincidencia exacta familia + variante
        row_map = bench_map[
            (bench_map["da_value"] == da_value) &
            (bench_map["da_variant"].astype(str).str.strip() == da_variant)
        ]
    if row_map.empty:
        # Paso 2: regla general de la familia (da_variant vacío)
        if _has_variant_col:
            row_map = bench_map[
                (bench_map["da_value"] == da_value) &
                (bench_map["da_variant"].astype(str).str.strip() == "")
            ]
        else:
            row_map = bench_map[bench_map["da_value"] == da_value]

    if row_map.empty:
        return {
            "da_value": da_value,
            "da_variant": da_variant,
            "test_type": "—",
            "bench_type": "—",
            "bench_capacity": 0.0,
            "tiempo_para": 0.0,
            "observation": "Datos insuficientes: sin asignación de banco",
        }
    bench_type = str(row_map.iloc[0]["bench_type"]).strip()

    # bench_type == test_type salvo XL_MV, que es banco específico pero prueba MV
    test_type = "MV" if bench_type == "XL_MV" else bench_type

    # 3. ¿Cuántos bancos hay?
    row_cfg = bench_cfg[bench_cfg["bench_type"] == bench_type]
    if row_cfg.empty:
        return {
            "da_value": da_value,
            "da_variant": da_variant,
            "test_type": test_type,
            "bench_type": bench_type,
            "bench_capacity": 0.0,
            "tiempo_para": 0.0,
            "observation": "Datos insuficientes: sin configuración de bancos para este tipo",
        }
    quantity = int(pd.to_numeric(row_cfg.iloc[0]["quantity"], errors="coerce") or 0)
    override_raw = row_cfg.iloc[0].get("hours_per_unit_override", None)

    # 4. Tiempo PARA — misma lógica que el motor:
    #    cycle_time_real = max(machine_time, labor_time / operators_per_station)
    t_row = times_df_b[
        (times_df_b["model"] == da_value) & (times_df_b["process"] == "PARA")
    ]
    s_row = stations_df_b[
        (stations_df_b["line_id"] == line_id) & (stations_df_b["process"] == "PARA")
    ]

    tiempo_para = None
    if not t_row.empty and not s_row.empty:
        machine_t = float(pd.to_numeric(t_row.iloc[0].get("machine_time", 0), errors="coerce") or 0)
        labor_t   = float(pd.to_numeric(t_row.iloc[0].get("labor_time",   0), errors="coerce") or 0)
        operators = float(pd.to_numeric(s_row.iloc[0].get("operators_per_station", 0), errors="coerce") or 0)

        # Protección técnica (no regla de negocio): si operators_per_station es
        # nulo, 0 o inválido, se usa 1 como valor de seguridad para evitar
        # división por cero en esta capa informativa. No implica que 1 sea el
        # valor real de la planta.
        if operators <= 0:
            operators = 1.0

        labor_per_op = labor_t / operators
        calc = max(machine_t, labor_per_op)
        if calc > 0:
            tiempo_para = calc

    # 5. Fallback al override manual (solo si PARA no pudo calcularse)
    if tiempo_para is None:
        if override_raw is not None:
            try:
                v = float(override_raw)
                if v > 0:
                    tiempo_para = v
            except (TypeError, ValueError):
                pass

    if tiempo_para is None or tiempo_para <= 0:
        return {
            "da_value": da_value,
            "da_variant": da_variant,
            "test_type": test_type,
            "bench_type": bench_type,
            "bench_capacity": 0.0,
            "tiempo_para": 0.0,
            "observation": "Datos insuficientes: sin tiempo PARA calculable",
        }

    # 6. Capacidad máxima por bancos
    bench_capacity = (hours_eff * quantity) / tiempo_para

    # 7. Observación
    if bench_capacity > demand_week:
        observation = "El banco no limita"
    else:
        observation = f"⚠️ Atención: los bancos {bench_type} podrían limitar"

    return {
        "da_value":       da_value,
        "da_variant":     da_variant,
        "test_type":      test_type,
        "bench_type":     bench_type,
        "bench_capacity": round(bench_capacity, 2),
        "tiempo_para":    tiempo_para,
        "observation":    observation,
    }


def _apply_capacity(merged: pd.DataFrame, h_eff: float) -> pd.DataFrame:
    """Calcula capacidad usando cycle_time_real = max(machine_time, labor_time/operators).
    capacity = (h_eff * stations) / cycle_time_real"""
    m = merged.copy()
    m["stations"] = pd.to_numeric(m["stations"], errors="coerce").fillna(0)
    m["operators_per_station"] = pd.to_numeric(m["operators_per_station"], errors="coerce").fillna(0).clip(lower=1.0)

    for col in ("machine_time", "labor_time"):
        if col in m.columns:
            m[col] = pd.to_numeric(m[col], errors="coerce").fillna(0.0)
        else:
            m[col] = 0.0

    m["labor_per_operator"] = m["labor_time"] / m["operators_per_station"]
    m["cycle_time_real"] = m[["machine_time", "labor_per_operator"]].max(axis=1)

    m["capacity"] = 0.0
    mask = (m["cycle_time_real"] > 0) & (m["stations"] > 0)
    m.loc[mask, "capacity"] = (h_eff * m.loc[mask, "stations"]) / m.loc[mask, "cycle_time_real"]

    return m


@st.cache_data(show_spinner=False)
def _compute_line_base(
    line_id: str,
    model: str,
    times_df_param: pd.DataFrame,
    stations_df_param: pd.DataFrame,
) -> pd.DataFrame:
    """Merge + normalización + cycle_time_real, sin hours_eff.
    Caché estable ante cambios de sliders (disponibilidad, eficiencia, turnos).
    Llamada internamente por compute_line_detail."""
    t = times_df_param[times_df_param["model"] == model].copy()
    s = stations_df_param[stations_df_param["line_id"] == line_id].copy()
    merged = pd.merge(s, t, on="process", how="inner")
    if merged.empty:
        return merged
    m = merged.copy()
    m["stations"] = pd.to_numeric(m["stations"], errors="coerce").fillna(0)
    m["operators_per_station"] = pd.to_numeric(m["operators_per_station"], errors="coerce").fillna(0).clip(lower=1.0)
    for col in ("machine_time", "labor_time"):
        if col in m.columns:
            m[col] = pd.to_numeric(m[col], errors="coerce").fillna(0.0)
        else:
            m[col] = 0.0
    m["labor_per_operator"] = m["labor_time"] / m["operators_per_station"]
    m["cycle_time_real"] = m[["machine_time", "labor_per_operator"]].max(axis=1)
    return m


@st.cache_data(show_spinner=False)
def compute_line_detail(
    line_id: str,
    model: str,
    times_df_param: pd.DataFrame,
    stations_df_param: pd.DataFrame,
    hours_eff_param: float,
) -> tuple[pd.DataFrame, str, float]:
    """
    Devuelve:
    - merged detail DF con capacity por proceso
    - bottleneck process (min capacity)
    - capacity_total_week (uds/sem) (cap del cuello)

    El trabajo pesado (merge + normalización) lo hace _compute_line_base,
    que cachea sin hours_eff. Solo la multiplicación final depende de hours_eff_param.
    """
    base = _compute_line_base(line_id, model, times_df_param, stations_df_param)

    if base.empty:
        return base, "", 0.0

    merged = base.copy()
    merged["capacity"] = 0.0
    mask = (merged["cycle_time_real"] > 0) & (merged["stations"] > 0)
    merged.loc[mask, "capacity"] = (hours_eff_param * merged.loc[mask, "stations"]) / merged.loc[mask, "cycle_time_real"]

    # Solo considerar procesos productivos para el cuello de botella:
    productive = merged[mask].copy()

    if productive.empty or productive["capacity"].dropna().empty:
        return merged, "", 0.0

    bottleneck_row = productive.loc[productive["capacity"].idxmin()]
    bottleneck_proc = str(bottleneck_row["process"])
    cap_week = float(bottleneck_row["capacity"])

    return merged, bottleneck_proc, cap_week


def compute_line_detail_v2(
    line_id: str,
    model: str,
    times_df_param: pd.DataFrame,
    stations_df_param: pd.DataFrame,
    hours_eff_default: float,
    hours_eff_by_process: dict,
) -> tuple[pd.DataFrame, str, float]:
    """Per-process hours_eff variant of compute_line_detail.
    hours_eff_by_process = {process: float}. Missing process → hours_eff_default.
    No cache decorator: dict arg is unhashable; _compute_line_base (cached) absorbs the heavy work."""
    base = _compute_line_base(line_id, model, times_df_param, stations_df_param)
    if base.empty:
        return base, "", 0.0
    merged = base.copy()
    merged["capacity"] = 0.0
    mask = (merged["cycle_time_real"] > 0) & (merged["stations"] > 0)
    merged.loc[mask, "capacity"] = merged[mask].apply(
        lambda r: (
            hours_eff_by_process.get(str(r["process"]), hours_eff_default) * r["stations"]
        ) / r["cycle_time_real"],
        axis=1,
    )
    productive = merged[mask].copy()
    if productive.empty or productive["capacity"].dropna().empty:
        return merged, "", 0.0
    bottleneck_row = productive.loc[productive["capacity"].idxmin()]
    bottleneck_proc = str(bottleneck_row["process"])
    cap_week = float(bottleneck_row["capacity"])
    return merged, bottleneck_proc, cap_week


def capacity_hours_for_output(merged: pd.DataFrame, output_units: float) -> float:
    """
    Capacidad (h/SEM) y (h/AÑO) = **horas-hombre (HH)** necesarias para producir `output_units`.

    En TU modelo:
    - `cycle_time` está en **HH/ud**
    - NO se multiplica por operarios (si no, duplicas HH)

    Regla:
        HH_total = sum_procesos( output_units * cycle_time )
    """
    if merged is None or merged.empty or output_units <= 0:
        return 0.0

    m = merged.copy()
    if "cycle_time" in m.columns:
        m["cycle_time"] = pd.to_numeric(m["cycle_time"], errors="coerce").fillna(0.0)
    else:
        m["cycle_time"] = 0.0

    hours_proc = output_units * m["cycle_time"]
    hours_proc = pd.to_numeric(hours_proc, errors="coerce").fillna(0.0)

    return float(hours_proc.sum())


@st.cache_data(show_spinner=False)
def _precompute_all_bases_for_tab4(
    times_df: pd.DataFrame,
    stations_df: pd.DataFrame,
    line_ids_tuple: tuple,
    allowed_by_line_tuple: tuple,
) -> dict:
    """Pre-computa merge+normalize para todas las combinaciones (line_id, model) de Tab 4.
    times_df y stations_df se hashean UNA SOLA VEZ en lugar de N×M veces.
    Sin hours_eff → caché estable ante cambios de sliders (disponibilidad, eficiencia, turnos)."""
    allowed_dict = {k: list(v) for k, v in allowed_by_line_tuple}
    result = {}
    for line_id in line_ids_tuple:
        models = allowed_dict.get(line_id, [])
        if not models:
            continue
        s = stations_df[stations_df["line_id"] == line_id].copy()
        if s.empty:
            continue
        s["stations"] = pd.to_numeric(s["stations"], errors="coerce").fillna(0)
        s["operators_per_station"] = pd.to_numeric(s["operators_per_station"], errors="coerce").fillna(0).clip(lower=1.0)
        for m in models:
            t = times_df[times_df["model"] == m].copy()
            merged = pd.merge(s, t, on="process", how="inner")
            if merged.empty:
                continue
            for col in ("machine_time", "labor_time"):
                if col in merged.columns:
                    merged[col] = pd.to_numeric(merged[col], errors="coerce").fillna(0.0)
                else:
                    merged[col] = 0.0
            merged["labor_per_operator"] = merged["labor_time"] / merged["operators_per_station"]
            merged["cycle_time_real"] = merged[["machine_time", "labor_per_operator"]].max(axis=1)
            result[(line_id, m)] = merged
    return result


if st.session_state.active_tab == "📈 Resultados":
    # ── State init + resolvers (antes de cualquier render) ───────────────────
    #
    # ESTRUCTURA DE SESIÓN
    #   session_state["line_params_override"][plant_id][scenario_key][line_id]
    #   = {"enabled": bool, "shifts": int, "availability": float, "efficiency": float}
    #
    # PRIORIDAD DE RESOLUCIÓN
    #   1. scenario_line_overrides del escenario activo  (persistencia principal)
    #   2. line_overrides de la planta                   (fallback si escenario vacío)
    #   3. parámetros globales de la planta              (si enabled=False o sin override)
    _active_sc_id = st.session_state.get(f"_session_sc_id_{plant_id}")
    _ov_sc_key = _active_sc_id if _active_sc_id is not None else 0
    st.session_state.setdefault("line_params_override", {})
    st.session_state["line_params_override"].setdefault(plant_id, {})
    st.session_state["line_params_override"][plant_id].setdefault(_ov_sc_key, {})
    _plant_ov = st.session_state["line_params_override"][plant_id][_ov_sc_key]

    # Carga desde DB una vez por (plant_id, scenario_key) por sesión
    _ov_load_key = f"_line_ov_loaded_{plant_id}_{_ov_sc_key}"
    if not st.session_state.get(_ov_load_key, False):
        if _active_sc_id:
            _db_ov = load_scenario_line_overrides(_active_sc_id)
            if not _db_ov:
                _db_ov = load_line_overrides(plant_id)
        else:
            _db_ov = load_line_overrides(plant_id)
        for _lid, _ov_row in _db_ov.items():
            _plant_ov[_lid] = _ov_row
        st.session_state[_ov_load_key] = True

    def _resolve_hours_eff(lid: str) -> float:
        _ov = _plant_ov.get(lid, {})
        if not _ov.get("enabled", False):
            return hours_eff  # hereda global completo
        _s = int(_ov.get("shifts", shifts))
        _a = float(_ov.get("availability", availability))
        _e = float(_ov.get("efficiency", efficiency))
        return hours_week * _s * _a * _e

    # ── Process shift overrides ───────────────────────────────────────────────
    # Structure: proc_shift_override[plant_id][ov_sc_key][line_id] = {process: shifts}
    # Only personalized processes are stored; absence means "hereda turno de línea".
    # Only meaningful when a scenario is active (_active_sc_id is not None).
    st.session_state.setdefault("proc_shift_override", {})
    st.session_state["proc_shift_override"].setdefault(plant_id, {})
    st.session_state["proc_shift_override"][plant_id].setdefault(_ov_sc_key, {})
    _proc_ov = st.session_state["proc_shift_override"][plant_id][_ov_sc_key]

    _proc_sh_load_key = f"_proc_sh_loaded_{plant_id}_{_ov_sc_key}"
    if not st.session_state.get(_proc_sh_load_key, False):
        if _active_sc_id:
            for _lid2, _pd2 in load_scenario_process_shifts(_active_sc_id).items():
                _proc_ov[_lid2] = _pd2
        st.session_state[_proc_sh_load_key] = True

    def _resolve_process_hours_eff(lid: str) -> dict | None:
        """Returns per-process hours_eff dict, or None if no process overrides.
        Line-level availability/efficiency apply; only shifts vary per process."""
        _line_pd = _proc_ov.get(lid)
        if not _line_pd:
            return None
        _ov_l = _plant_ov.get(lid, {})
        _a = float(_ov_l.get("availability", availability)) if _ov_l.get("enabled") else availability
        _e = float(_ov_l.get("efficiency", efficiency)) if _ov_l.get("enabled") else efficiency
        return {proc: hours_week * float(sh) * _a * _e for proc, sh in _line_pd.items()}

    summary_rows = []
    detail_by_line = {}

    all_line_ids = sorted(
        stations_df["line_id"]
        .astype(str)
        .str.strip()
        .unique()
       .tolist()
    )

    _tc = times_df.copy()
    _tc["cycle_time"] = pd.to_numeric(_tc["cycle_time"], errors="coerce").fillna(0.0)
    cycle_time_by_model = _tc.groupby("model")["cycle_time"].sum().to_dict()

    # ── UI: override de parámetros por línea (Bloques 8A + 10) ───────────────
    _planned_line_ids = [
        lid for lid in all_line_ids
        if st.session_state.line_model.get(st.session_state["plant_id"], {}).get(lid)
    ]

    # Fase 1 — Hidratación: inicializa widget keys desde line_params_override (fuente de verdad)
    # Solo escribe si la clave NO existe aún → respeta interacciones recientes del usuario.
    # Las claves de valor se inicializan según el estado ACTUAL del checkbox (no el guardado),
    # para cubrir el caso en que el usuario activa el override en este mismo run.
    for _lid in _planned_line_ids:
        _ov_saved = _plant_ov.get(_lid, {})
        _ov_saved_en = bool(_ov_saved.get("enabled", False))
        _en_key = f"ov_en_{plant_id}_{_ov_sc_key}_{_lid}"
        if _en_key not in st.session_state:
            st.session_state[_en_key] = _ov_saved_en
        # Leer estado actual (refleja interacción del usuario en este run)
        _cur_en = bool(st.session_state[_en_key])
        if _cur_en:
            for _wk, _wv in [
                (f"shifts_ov_{plant_id}_{_ov_sc_key}_{_lid}",  int(_ov_saved.get("shifts", shifts))),
                (f"avail_ov_{plant_id}_{_ov_sc_key}_{_lid}",   float(_ov_saved.get("availability", availability))),
                (f"eff_ov_{plant_id}_{_ov_sc_key}_{_lid}",     float(_ov_saved.get("efficiency", efficiency))),
            ]:
                if _wk not in st.session_state:
                    st.session_state[_wk] = _wv

    # Fase 2 — Limpieza: borra claves de valor para líneas con override OFF
    # Opera sobre widget state ya hidratado → garantiza que los sliders muestren global
    for _lid in _planned_line_ids:
        if not st.session_state.get(f"ov_en_{plant_id}_{_ov_sc_key}_{_lid}", False):
            for _wk in [f"shifts_ov_{plant_id}_{_ov_sc_key}_{_lid}", f"avail_ov_{plant_id}_{_ov_sc_key}_{_lid}", f"eff_ov_{plant_id}_{_ov_sc_key}_{_lid}"]:
                st.session_state.pop(_wk, None)

    _active_ov_lines = [
        lid for lid in _planned_line_ids
        if st.session_state.get(f"ov_en_{plant_id}_{_ov_sc_key}_{lid}", False)
    ]

    # ── Título + botón guardar (misma fila) ───────────────────────────────────
    _ex_sc_name = st.session_state.get(f"_sc_name_map_{plant_id}", {}).get(_active_sc_id, "—")


    _ex_header_info = (
        f"**{_ex_sc_name}**"
        f"  ·  {len(_planned_line_ids)} líneas"
        + (f"  ·  ✏ {len(_active_ov_lines)} ajuste(s)" if _active_ov_lines else "")
    )
    _hdr_title_col, _hdr_save_col = st.columns([3, 1])
    _hdr_title_col.subheader(t("tab_res_header"))
    if _has_db() and _planned_line_ids:
        if _hdr_save_col.button(
            t("res_save_line_btn"),
            key="btn_save_line_overrides",
            use_container_width=True,
        ):
            if _active_sc_id:
                _saved_ok = save_scenario_line_overrides(_active_sc_id, _plant_ov)
            else:
                _saved_ok = save_line_overrides(plant_id, _plant_ov)
            if _saved_ok:
                st.toast(t("res_save_line_ok"), icon="✅")
            else:
                st.toast(t("res_save_line_err"), icon="🚨")

    # ── Bloque azul escenario (debajo del título) ─────────────────────────────
    st.info(_ex_header_info)

    _expander_label = t("res_params_expander")
    if _active_ov_lines:
        _expander_label += f"  —  ✏ {len(_active_ov_lines)} activo(s)"
    with st.expander(_expander_label, expanded=False):
        if _planned_line_ids:
            import math as _math
            _chunk = _math.ceil(len(_planned_line_ids) / 3)
            _groups = [
                _planned_line_ids[0:_chunk],
                _planned_line_ids[_chunk:_chunk * 2],
                _planned_line_ids[_chunk * 2:],
            ]
            _col_a, _col_b, _col_c = st.columns(3)
            for _grp_col, _grp_lids in zip((_col_a, _col_b, _col_c), _groups):
                if not _grp_lids:
                    continue
                with _grp_col:
                    for _lid in _grp_lids:
                        _en = st.session_state.get(f"ov_en_{plant_id}_{_ov_sc_key}_{_lid}", False)
                        if _en:
                            # Línea con override: cabecera checkbox | nombre+indicador | [⚙ Procesos]
                            _proc_line_ov = _proc_ov.get(_lid, {})
                            _proc_c = len(_proc_line_ov)
                            _proc_indicator = f" · proc: {_proc_c}" if _proc_c else ""
                            _rc, _rn, _rp = st.columns([0.4, 2.8, 1.4], gap="small", vertical_alignment="center")
                            _rc.checkbox(
                                "ov",
                                key=f"ov_en_{plant_id}_{_ov_sc_key}_{_lid}",
                                label_visibility="collapsed",
                            )
                            _rn.markdown(
                                f'<p style="margin:0;line-height:1;font-weight:600">✏ {_lid}{_proc_indicator}</p>',
                                unsafe_allow_html=True,
                            )
                            # Botón ⚙ Procesos: solo disponible con escenario activo
                            _lid_model = st.session_state.line_model.get(plant_id, {}).get(_lid, "")
                            if _active_sc_id and _lid_model:
                                with _rp.popover("⚙ Procesos", use_container_width=True):
                                    _base_df = _compute_line_base(_lid, _lid_model, times_df, stations_df)
                                    _procs_list = _base_df["process"].tolist() if not _base_df.empty else []
                                    _line_global_sh = int(st.session_state.get(f"shifts_ov_{plant_id}_{_ov_sc_key}_{_lid}", shifts))
                                    if not _procs_list:
                                        st.caption(t("res_proc_no_procs"))
                                    else:
                                        st.caption(f"{t('res_proc_global_shifts')} **{_line_global_sh}**")
                                        for _proc in _procs_list:
                                            _pen_key = f"psh_en_{plant_id}_{_ov_sc_key}_{_lid}_{_proc}"
                                            _pval_key = f"psh_val_{plant_id}_{_ov_sc_key}_{_lid}_{_proc}"
                                            if _pen_key not in st.session_state:
                                                st.session_state[_pen_key] = _proc in _proc_line_ov
                                            if st.session_state[_pen_key] and _pval_key not in st.session_state:
                                                st.session_state[_pval_key] = int(_proc_line_ov.get(_proc, _line_global_sh))
                                            _pc, _pn, _pt = st.columns([0.5, 1.6, 1.4], vertical_alignment="center")
                                            _pc.checkbox("p", key=_pen_key, label_visibility="collapsed")
                                            _pn.markdown(
                                                f'<p style="margin:0;line-height:1;font-size:0.85rem;position:relative;top:-3px">{_proc}</p>',
                                                unsafe_allow_html=True,
                                            )
                                            if st.session_state[_pen_key]:
                                                if _pval_key not in st.session_state:
                                                    st.session_state[_pval_key] = int(_proc_line_ov.get(_proc, _line_global_sh))
                                                _pt.number_input(
                                                    "t", min_value=1, max_value=5, step=1,
                                                    key=_pval_key, label_visibility="collapsed",
                                                )
                                            else:
                                                st.session_state.pop(_pval_key, None)
                                                _pt.markdown(
                                                    f'<p style="margin:0;line-height:1;font-size:0.8rem;color:gray">hereda ({_line_global_sh})</p>',
                                                    unsafe_allow_html=True,
                                                )
                                        # Sync proc_ov from widget state
                                        _new_proc_ov: dict = {}
                                        for _proc in _procs_list:
                                            _pen_key = f"psh_en_{plant_id}_{_ov_sc_key}_{_lid}_{_proc}"
                                            _pval_key = f"psh_val_{plant_id}_{_ov_sc_key}_{_lid}_{_proc}"
                                            if st.session_state.get(_pen_key, False):
                                                _new_proc_ov[_proc] = int(st.session_state.get(_pval_key, _line_global_sh))
                                        _proc_ov[_lid] = _new_proc_ov
                                        st.divider()
                                        if st.button(
                                            "💾 Guardar turnos por proceso",
                                            key=f"btn_save_psh_{plant_id}_{_ov_sc_key}_{_lid}",
                                        ):
                                            _psh_ok = save_scenario_process_shifts(_active_sc_id, _lid, _new_proc_ov)
                                            if _psh_ok:
                                                st.success(t("res_proc_save_ok"))
                                            else:
                                                st.error(t("res_proc_save_err"))
                            else:
                                _rp.button(
                                    "⚙ Procesos", disabled=True, use_container_width=True,
                                    key=f"psh_btn_dis_{plant_id}_{_ov_sc_key}_{_lid}",
                                )
                            # Controles: turnos | disponib. | eficiencia
                            _rs, _ra, _re = st.columns([1, 2, 2])
                            _rs.number_input(
                                t("param_shifts"), min_value=1, max_value=5, step=1,
                                key=f"shifts_ov_{plant_id}_{_ov_sc_key}_{_lid}",
                                help=f"Global: {shifts}",
                            )
                            _ra.slider(
                                t("param_availability"), min_value=0.0, max_value=1.0, step=0.01,
                                key=f"avail_ov_{plant_id}_{_ov_sc_key}_{_lid}",
                                help=f"Global: {availability}",
                            )
                            _re.slider(
                                t("param_efficiency"), min_value=0.0, max_value=1.0, step=0.01,
                                key=f"eff_ov_{plant_id}_{_ov_sc_key}_{_lid}",
                                help=f"Global: {efficiency}",
                            )
                        else:
                            # Línea sin override: fila compacta, sin sliders
                            _rc, _rn = st.columns([0.4, 3.6], gap="small", vertical_alignment="center")
                            _rc.checkbox(
                                "ov",
                                key=f"ov_en_{plant_id}_{_ov_sc_key}_{_lid}",
                                label_visibility="collapsed",
                            )
                            _rn.markdown(
                                f'<p style="margin:0;line-height:1;font-size:0.85rem">{_lid}</p>',
                                unsafe_allow_html=True,
                            )
            if _active_ov_lines:
                st.caption(
                    f"✏ Override activo en: **{', '.join(_active_ov_lines)}**. "
                    f"Resto hereda global (turnos: {shifts}, disp.: {availability}, ef.: {efficiency})."
                )
            else:
                st.caption(
                    f"Todas las líneas heredan el global "
                    f"(turnos: {shifts}, disponibilidad: {availability}, eficiencia: {efficiency})."
                )
            # Caption horas efectivas — al borde inferior del expander
            if _active_ov_lines:
                st.caption(
                    f"{t('res_hours_eff')} {_fmt_num(hours_eff)} {t('unit_week')} "
                    f"(referencia global). Parámetros pueden variar por línea."
                )
            else:
                st.caption(f"{t('res_hours_eff')} {_fmt_num(hours_eff)} {t('unit_week')}")
        else:
            st.caption("Sin líneas planificadas. Selecciona modelos en Planificación.")

    # Sincronizar line_params_override con estado explícito enabled/disabled
    for _lid in _planned_line_ids:
        _en = st.session_state.get(f"ov_en_{plant_id}_{_ov_sc_key}_{_lid}", False)
        if _en:
            _plant_ov[_lid] = {
                "enabled":      True,
                "shifts":       int(st.session_state.get(f"shifts_ov_{plant_id}_{_ov_sc_key}_{_lid}", shifts)),
                "availability": float(st.session_state.get(f"avail_ov_{plant_id}_{_ov_sc_key}_{_lid}", availability)),
                "efficiency":   float(st.session_state.get(f"eff_ov_{plant_id}_{_ov_sc_key}_{_lid}", efficiency)),
            }
        else:
            _plant_ov[_lid] = {"enabled": False}

    for line_id in all_line_ids:

        parts = line_id.split("-", 1)

        if len(parts) == 2:
            nave, base_line = parts
        else:
            nave = "N1"
            base_line = parts[0]

        model = (
            st.session_state.line_model
            .get(st.session_state["plant_id"], {})
            .get(line_id)
        )
        if not model:
            continue

        demand_week = float(
            st.session_state.line_demand
            .get(st.session_state["plant_id"], {})
            .get(line_id, 0.0)
        )

        _line_hours_eff = _resolve_hours_eff(line_id)
        _proc_he = _resolve_process_hours_eff(line_id)
        if _proc_he is not None:
            merged, bottleneck_proc, cap_week = compute_line_detail_v2(
                line_id, model, times_df, stations_df, _line_hours_eff, _proc_he
            )
        else:
            merged, bottleneck_proc, cap_week = compute_line_detail(
                line_id, model, times_df, stations_df, _line_hours_eff
            )

        saturation = 0.0
        deficit = 0.0
        if cap_week > 0:
            saturation = (demand_week / cap_week) * 100.0
            deficit = max(0.0, demand_week - cap_week)

        demand_year = demand_week * weeks_equiv
        cap_year = cap_week * weeks_equiv

        total_cycle_time_model = cycle_time_by_model.get(model, 0.0)

        cap_hours_week = cap_week * total_cycle_time_model
        cap_hours_year = cap_hours_week * weeks_equiv

        dem_hours_week = demand_week * total_cycle_time_model
        dem_hours_year = dem_hours_week * weeks_equiv

        # Bloque 9/10 — personas equivalentes con disponibilidad/eficiencia por línea
        # Denominador: hours_week × availability × efficiency (SIN shifts — una persona no trabaja dos turnos)
        _ov_loop = _plant_ov.get(line_id, {})
        if _ov_loop.get("enabled", False):
            _a_line = float(_ov_loop.get("availability", availability))
            _e_line = float(_ov_loop.get("efficiency", efficiency))
        else:
            _a_line = availability
            _e_line = efficiency
        _fte_denom = hours_week * _a_line * _e_line
        people_eq = dem_hours_week / _fte_denom if _fte_denom > 0 else 0.0

        summary_rows.append({
            "nave": nave,
            "line": base_line,
            "line_id": line_id,
            "model": model,
            "Demanda (UDS/SEM)": demand_week,
            "Capacidad (UDS/SEM)": cap_week,
            "Saturación (%)": saturation,
            "Déficit (UDS/SEM)": deficit,
            "bottleneck": bottleneck_proc,
            "Demanda (UDS/AÑO)": demand_year,
            "Capacidad (UDS/AÑO)": cap_year,
            "Demanda (h/SEM)": dem_hours_week,
            "Capacidad (h/SEM)": cap_hours_week,
            "Demanda (h/AÑO)": dem_hours_year,
            "Capacidad (h/AÑO)": cap_hours_year,
            "Personas eq.": people_eq,
        })

        detail_by_line[line_id] = (nave, base_line, model, demand_week, bottleneck_proc, merged)

    summary_df = pd.DataFrame(summary_rows)

    if not summary_df.empty:
        _sum_cols = [
            "Demanda (UDS/SEM)", "Capacidad (UDS/SEM)",
            "Demanda (UDS/AÑO)", "Capacidad (UDS/AÑO)",
            "Demanda (h/SEM)", "Capacidad (h/SEM)",
            "Demanda (h/AÑO)", "Capacidad (h/AÑO)",
            "Personas eq.",
        ]
        for c in _sum_cols:
            if c in summary_df.columns:
                summary_df[c] = pd.to_numeric(summary_df[c], errors="coerce")

        # Ordenar por criticidad antes de añadir TOTAL
        _def_s = pd.to_numeric(summary_df.get("Déficit (UDS/SEM)", 0), errors="coerce").fillna(0.0)
        _sat_s = pd.to_numeric(summary_df.get("Saturación (%)", 0), errors="coerce").fillna(0.0)
        summary_df["_prio"] = np.where(
            _def_s > 0, 0,
            np.where(_sat_s >= 90, 1, np.where(_sat_s >= 70, 2, 3))
        )
        summary_df = (
            summary_df
            .sort_values(["_prio", "Déficit (UDS/SEM)", "Saturación (%)"], ascending=[True, False, False])
            .drop(columns=["_prio"])
            .reset_index(drop=True)
        )

        total_row = {c: float(summary_df[c].sum(skipna=True)) if c in summary_df.columns else 0.0 for c in _sum_cols}
        total_row.update({
            "nave": "",
            "line": "",
            "line_id": "TOTAL",
            "model": "",
            "Saturación (%)": float("nan"),
            "Déficit (UDS/SEM)": float("nan"),
            "bottleneck": "",
        })

        for c in ["nave", "line", "line_id", "model", "bottleneck"]:
            if c not in summary_df.columns:
                summary_df[c] = ""

        summary_df = pd.concat([summary_df, pd.DataFrame([total_row])], ignore_index=True)

        # ── Resumen ejecutivo — cómputo (render tras la tabla) ───────────────
        _ex = summary_df.iloc[:-1]  # excluye fila TOTAL
        _ex_cap  = pd.to_numeric(_ex["Capacidad (UDS/SEM)"], errors="coerce").fillna(0).sum()
        _ex_dem  = pd.to_numeric(_ex["Demanda (UDS/SEM)"],   errors="coerce").fillna(0).sum()
        _ex_def  = pd.to_numeric(_ex["Déficit (UDS/SEM)"],   errors="coerce").fillna(0).sum()
        _ex_sat  = pd.to_numeric(_ex["Saturación (%)"],      errors="coerce").fillna(0).max()
        _ex_crit = _ex[pd.to_numeric(_ex["Déficit (UDS/SEM)"], errors="coerce").fillna(0) > 0]["line_id"].tolist()

    def style_summary(df: pd.DataFrame):
        styled = df.copy()

        fmt_cols_1 = [
            "Demanda (UDS/SEM)", "Capacidad (UDS/SEM)", "Déficit (UDS/SEM)",
            "Demanda (UDS/AÑO)", "Capacidad (UDS/AÑO)",
            "Demanda (h/SEM)", "Capacidad (h/SEM)",
            "Demanda (h/AÑO)", "Capacidad (h/AÑO)"
        ]
        for c in fmt_cols_1:
            if c in styled.columns:
                styled[c] = pd.to_numeric(styled[c], errors="coerce")

        styled["Saturación (%)"] = pd.to_numeric(styled["Saturación (%)"], errors="coerce")

        _fmt_dict = {c: "{:.1f}" for c in fmt_cols_1 if c in styled.columns}
        _fmt_dict["Saturación (%)"] = "{:.1f} %"
        if "Personas eq." in styled.columns:
            styled["Personas eq."] = pd.to_numeric(styled["Personas eq."], errors="coerce")
            _fmt_dict["Personas eq."] = "{:.1f}"
        s = styled.style.format(_fmt_dict)

        def sat_color(val):
            try:
                v = float(val)
            except Exception:
                return ""
            return "color: red; font-weight: 700;" if v >= 100 else "color: green; font-weight: 700;"

        s = s.map(sat_color, subset=["Saturación (%)"])
        s = s.map(lambda _: "color: red; font-weight: 700;", subset=["bottleneck"])

        def _style_row(row):
            if str(row.get("line", "")) == "TOTAL":
                return ["font-weight: bold; font-size: 16px; background-color: #f0f0f0;"] * len(row)
            _dv = pd.to_numeric(row.get("Déficit (UDS/SEM)", 0), errors="coerce")
            _sv = pd.to_numeric(row.get("Saturación (%)", 0), errors="coerce")
            if pd.notna(_dv) and _dv > 0:
                return ["background-color: #FFE4E4;"] * len(row)
            if pd.notna(_sv) and _sv >= 90:
                return ["background-color: #FFF3CD;"] * len(row)
            return [""] * len(row)
        s = s.apply(_style_row, axis=1)

        return s

    if summary_df.empty:
        st.info(t("res_no_results_yet"))
    else:
        display_cols = [
            "nave", "line", "model",
            "Demanda (UDS/SEM)", "Capacidad (UDS/SEM)", "Saturación (%)", "Déficit (UDS/SEM)",
            "bottleneck",
            "Demanda (UDS/AÑO)", "Capacidad (UDS/AÑO)",
            "Demanda (h/SEM)", "Personas eq.", "Capacidad (h/SEM)",
            "Demanda (h/AÑO)", "Capacidad (h/AÑO)",
        ]
        display_cols = [c for c in display_cols if c in summary_df.columns]

        # ── Cómputo panel de lectura rápida (sin render — va tras la tabla) ───
        _lo = summary_df[summary_df["line_id"] != "TOTAL"]
        _def_v = pd.to_numeric(_lo["Déficit (UDS/SEM)"], errors="coerce").fillna(0.0)
        _sat_v = pd.to_numeric(_lo["Saturación (%)"], errors="coerce").fillna(0.0)
        _n_def = int((_def_v > 0).sum())
        _n_crit = int((_sat_v >= 90).sum())
        if not _sat_v.empty:
            _mx_idx = _sat_v.idxmax()
            _mx_sat = _sat_v[_mx_idx]
            _mx_line = str(_lo.loc[_mx_idx, "line"]) if "line" in _lo.columns else "—"
        else:
            _mx_sat, _mx_line = 0.0, "—"
        _crit_lo = _lo[(_def_v > 0) | (_sat_v >= 90)]
        _main_bn = (
            str(_crit_lo.iloc[0]["bottleneck"])
            if not _crit_lo.empty and "bottleneck" in _crit_lo.columns and pd.notna(_crit_lo.iloc[0]["bottleneck"])
            else "—"
        )

        # ── Cómputo FTE (sin render) ──────────────────────────────────────────
        _fte_show = False
        _total_fte = 0.0
        _top_fte_line, _top_fte_val, _top_fte_pct = "—", 0.0, 0.0
        if "Personas eq." in _lo.columns:
            _fte_v = pd.to_numeric(_lo["Personas eq."], errors="coerce").fillna(0.0)
            _total_fte = float(_fte_v.sum())
            if _total_fte > 0 and not _fte_v.empty:
                _top_fte_idx = _fte_v.idxmax()
                _top_fte_line = str(_lo.loc[_top_fte_idx, "line"]) if "line" in _lo.columns else "—"
                _top_fte_val = float(_fte_v[_top_fte_idx])
                _top_fte_pct = _top_fte_val / _total_fte * 100.0
                _fte_show = True

        # ── Cómputo buffer Excel (sin render) ────────────────────────────────
        total_display_df = summary_df.copy()
        total_display_df.loc[total_display_df["line_id"] == "TOTAL", ["nave", "line"]] = ["", "TOTAL"]
        _export_df = total_display_df[display_cols].copy()
        _export_df = _export_df.rename(columns={
            "nave": "Nave", "line": "Línea", "model": "Modelo", "bottleneck": "Cuello de botella"
        })
        _sel_sc_id = st.session_state.get(f"_session_sc_id_{plant_id}")
        _sc_name_export = st.session_state.get(f"_sc_name_map_{plant_id}", {}).get(_sel_sc_id, "")
        _export_ts = datetime.now().strftime("%Y-%m-%d %H:%M")
        _export_filename = f"capacidad_{selected_plant_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        _buf = io.BytesIO()
        with pd.ExcelWriter(_buf, engine="openpyxl") as _writer:
            _meta = pd.DataFrame([
                {"Campo": "Planta", "Valor": selected_plant_name},
                {"Campo": "Escenario", "Valor": _sc_name_export},
                {"Campo": "Exportado", "Valor": _export_ts},
            ])
            _meta.to_excel(_writer, sheet_name="Info", index=False)
            _export_df.to_excel(_writer, sheet_name="Resultados", index=False)
        _buf.seek(0)

        # ── UI: 4 métricas en banda compacta ─────────────────────────────────
        st.markdown('<div style="margin-top:-2rem;margin-bottom:-1rem">', unsafe_allow_html=True)
        _pm1, _pm2, _pm3, _pm4 = st.columns(4, vertical_alignment="bottom")
        _pm1.metric(t("res_panel_n_deficit"), _n_def)
        _pm2.metric(t("res_panel_max_sat"), f"{_fmt_num(_mx_sat)} %", help=_mx_line)
        _pm3.metric(t("res_panel_n_critical"), _n_crit)
        _pm4.metric(t("res_panel_bottleneck"), _main_bn)
        st.markdown('</div>', unsafe_allow_html=True)

        # ── UI: tabla ────────────────────────────────────────────────────────
        st.dataframe(style_summary(total_display_df[display_cols]), use_container_width=True, hide_index=True)

        # ── UI: 5 métricas resumen (bajo la tabla) ────────────────────────────

        _mc1, _mc2, _mc3, _mc4, _mc5 = st.columns(5)
        _mc1.metric(t("res_metric_cap_total"), _fmt_num(_ex_cap))
        _mc2.metric(t("res_metric_dem_total"), _fmt_num(_ex_dem))
        _mc3.metric(t("res_metric_deficit"),   _fmt_num(_ex_def),
                    delta=None if _ex_def == 0 else f"−{_fmt_num(_ex_def)}",
                    delta_color="inverse")
        _mc4.metric(t("res_metric_sat_max"),   _fmt_num(_ex_sat))
        _mc5.metric(t("res_metric_crit_lines"),
                    str(len(_ex_crit)),
                    help=", ".join(_ex_crit) if _ex_crit else "Ninguna")

        # ── UI: bloque azul FTE ───────────────────────────────────────────────
        if _fte_show:
            st.info(
                t("res_fte_info").format(
                    total_fte=_fmt_num(_total_fte),
                    top_line=_top_fte_line,
                    top_fte=_fmt_num(_top_fte_val),
                    top_pct=_fmt_num(_top_fte_pct),
                )
            )

        # ── Export simple — salida secundaria del escenario activo ──────────
        st.download_button(
            label=t("res_export_btn"),
            data=_buf,
            file_name=_export_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # --- Comparativa entre escenarios ---
        if _has_db():
            _cmp_sc_list = list_scenarios(plant_id)
            if _cmp_sc_list and len(_cmp_sc_list) > 1:
                st.markdown(t("res_compare_header"))

                _cmp_id_map = {s["id"]: s["name"] for s in _cmp_sc_list if s["id"] != _sel_sc_id}
                _cmp_options = [None] + list(_cmp_id_map.keys())
                _cmp_labels = {None: t("res_compare_none")}
                _cmp_labels.update(_cmp_id_map)

                _cmp_sel = st.selectbox(
                    t("res_compare_select"),
                    options=_cmp_options,
                    format_func=lambda sid: _cmp_labels.get(sid, ""),
                    key=f"res_compare_sc_{plant_id}",
                    label_visibility="collapsed",
                )

                if _cmp_sel is not None:
                    _cmp_sc_data = load_scenario_by_id(_cmp_sel)
                    if _cmp_sc_data is None:
                        st.warning(t("res_compare_no_data"))
                    else:
                        def _build_summary_df(line_model_d: dict, line_demand_d: dict) -> pd.DataFrame:
                            _rows = []
                            for _lid in all_line_ids:
                                _parts = _lid.split("-", 1)
                                _nave = _parts[0] if len(_parts) == 2 else "N1"
                                _bline = _parts[1] if len(_parts) == 2 else _parts[0]
                                _mdl = line_model_d.get(_lid)
                                if not _mdl:
                                    continue
                                _dem_w = float(line_demand_d.get(_lid, 0.0))
                                _, _, _cap_w = compute_line_detail(_lid, _mdl, times_df, stations_df, hours_eff)
                                _sat = (_dem_w / _cap_w * 100.0) if _cap_w > 0 else 0.0
                                _def = max(0.0, _dem_w - _cap_w)
                                _ctm = cycle_time_by_model.get(_mdl, 0.0)
                                _rows.append({
                                    "nave": _nave, "line": _bline, "line_id": _lid, "model": _mdl,
                                    "Demanda (UDS/SEM)": _dem_w,
                                    "Capacidad (UDS/SEM)": _cap_w,
                                    "Saturación (%)": _sat,
                                    "Déficit (UDS/SEM)": _def,
                                    "Demanda (UDS/AÑO)": _dem_w * weeks_equiv,
                                    "Capacidad (UDS/AÑO)": _cap_w * weeks_equiv,
                                    "Demanda (h/SEM)": _dem_w * _ctm,
                                    "Capacidad (h/SEM)": _cap_w * _ctm,
                                    "Demanda (h/AÑO)": _dem_w * _ctm * weeks_equiv,
                                    "Capacidad (h/AÑO)": _cap_w * _ctm * weeks_equiv,
                                })
                            return pd.DataFrame(_rows)

                        from openpyxl.styles import Font, PatternFill, Alignment

                        # ── Helpers de formato ────────────────────────────────
                        def _fnes(v) -> str:
                            """Número formateado para Excel español: entero sin decimal, float máx 2 con coma."""
                            s = _fmt_num(v).replace(".", ",")
                            return s.rstrip(",")

                        def _fnes_signed(v) -> str:
                            """Como _fnes pero con + explícito en positivos."""
                            s = _fnes(v)
                            try:
                                if v > 0:
                                    return "+" + s
                            except TypeError:
                                pass
                            return s

                        def _lectura_linea(def_a, def_b, sat_a, sat_b):
                            """Devuelve (texto, orden_impacto). Menor orden = más crítico."""
                            if def_a == 0 and def_b > 0:
                                return "aparece déficit", 0
                            if def_a > 0 and def_b > def_a:
                                return "el déficit crece", 1
                            if def_a > 0 and 0 < def_b < def_a:
                                return "el déficit se reduce", 2
                            if def_a > 0 and def_b == 0:
                                return "desaparece el déficit", 3
                            _ds = sat_b - sat_a
                            if _ds >= 3.0:
                                return "sube la carga", 4
                            if _ds <= -3.0:
                                return "reduce la carga", 5
                            return "sin cambio relevante", 6

                        def _xl_hdr(ws, hdr_fills: dict, freeze=True, autofilter=True):
                            """Aplica estilo de cabecera. hdr_fills = {col_1based: PatternFill}"""
                            _bold = Font(bold=True)
                            _center = Alignment(horizontal="center", vertical="center", wrap_text=True)
                            _default_fill = PatternFill("solid", fgColor="FFD9D9D9")
                            for cell in ws[1]:
                                cell.font = _bold
                                cell.fill = hdr_fills.get(cell.column, _default_fill)
                                cell.alignment = _center
                            if freeze:
                                ws.freeze_panes = "A2"
                            if autofilter:
                                ws.auto_filter.ref = ws.dimensions
                            for col in ws.columns:
                                _mw = max((len(str(c.value or "")) for c in col), default=8)
                                ws.column_dimensions[col[0].column_letter].width = min(max(_mw + 3, 10), 32)

                        # ── Datos de comparación ──────────────────────────────
                        _cmp_model_d = _cmp_sc_data["line_model"]
                        _cmp_demand_d = _cmp_sc_data["line_demand"]
                        _cmp_df = _build_summary_df(_cmp_model_d, _cmp_demand_d)

                        _cols_base = ["Demanda (UDS/SEM)", "Capacidad (UDS/SEM)", "Saturación (%)", "Déficit (UDS/SEM)"]
                        _base_lines_df = summary_df[summary_df["line_id"] != "TOTAL"]
                        _base_idx = _base_lines_df[["line_id"] + _cols_base].set_index("line_id")
                        _cmp_idx = (
                            _cmp_df[["line_id"] + _cols_base].set_index("line_id")
                            if not _cmp_df.empty else pd.DataFrame(columns=_cols_base)
                        )
                        _union_idx = _base_idx.index.union(_cmp_idx.index)
                        _base_al = _base_idx.reindex(_union_idx, fill_value=0.0)
                        _cmp_al = _cmp_idx.reindex(_union_idx, fill_value=0.0)
                        _bm = (_base_lines_df.set_index("line_id")["model"].to_dict()
                               if "model" in _base_lines_df.columns else {})
                        _cm = (_cmp_df.set_index("line_id")["model"].to_dict()
                               if not _cmp_df.empty and "model" in _cmp_df.columns else {})

                        # ── Filas de la COMPARATIVA ───────────────────────────
                        _comp_rows = []
                        for _lid in _union_idx:
                            _pts = _lid.split("-", 1)
                            _nave = _pts[0] if len(_pts) == 2 else "N1"
                            _bl = _pts[1] if len(_pts) == 2 else _pts[0]
                            _da = _base_al.at[_lid, "Demanda (UDS/SEM)"]
                            _ca = _base_al.at[_lid, "Capacidad (UDS/SEM)"]
                            _sa = _base_al.at[_lid, "Saturación (%)"]
                            _fa = _base_al.at[_lid, "Déficit (UDS/SEM)"]
                            _db = _cmp_al.at[_lid, "Demanda (UDS/SEM)"]
                            _cb = _cmp_al.at[_lid, "Capacidad (UDS/SEM)"]
                            _sb = _cmp_al.at[_lid, "Saturación (%)"]
                            _fb = _cmp_al.at[_lid, "Déficit (UDS/SEM)"]
                            _ma = _bm.get(_lid, "")
                            _mb = _cm.get(_lid, "")
                            _modelo = _ma if _ma == _mb else f"{_ma or '—'} → {_mb or '—'}"
                            _lec, _ord = _lectura_linea(_fa, _fb, _sa, _sb)
                            _comp_rows.append({
                                "Nave": _nave, "Línea": _bl, "Modelo": _modelo,
                                "Dem. Partida": _fnes(_da), "Cap. Partida": _fnes(_ca),
                                "Sat. Partida (%)": _fnes(_sa), "Déf. Partida": _fnes(_fa),
                                "Dem. Evaluado": _fnes(_db), "Cap. Evaluado": _fnes(_cb),
                                "Sat. Evaluado (%)": _fnes(_sb), "Déf. Evaluado": _fnes(_fb),
                                "Cambio cap.": _fnes_signed(_cb - _ca),
                                "Cambio sat. (pts)": _fnes_signed(_sb - _sa),
                                "Cambio déf.": _fnes_signed(_fb - _fa),
                                "Lectura": _lec,
                                "_ord": _ord, "_dsat": _sb - _sa, "_fa": _fa, "_fb": _fb,
                            })

                        _comp_rows.sort(key=lambda r: (r["_ord"], -abs(r["_dsat"])))

                        # Fila TOTAL para COMPARATIVA
                        _tca = _base_al["Capacidad (UDS/SEM)"].sum()
                        _tcb = _cmp_al["Capacidad (UDS/SEM)"].sum()
                        _tda = _base_al["Demanda (UDS/SEM)"].sum()
                        _tdb = _cmp_al["Demanda (UDS/SEM)"].sum()
                        _tfa = _base_al["Déficit (UDS/SEM)"].sum()
                        _tfb = _cmp_al["Déficit (UDS/SEM)"].sum()
                        _total_comp_row = {
                            "Nave": "", "Línea": "TOTAL", "Modelo": "",
                            "Dem. Partida": _fnes(_tda), "Cap. Partida": _fnes(_tca),
                            "Sat. Partida (%)": "", "Déf. Partida": _fnes(_tfa),
                            "Dem. Evaluado": _fnes(_tdb), "Cap. Evaluado": _fnes(_tcb),
                            "Sat. Evaluado (%)": "", "Déf. Evaluado": _fnes(_tfb),
                            "Cambio cap.": _fnes_signed(_tcb - _tca),
                            "Cambio sat. (pts)": "", "Cambio déf.": _fnes_signed(_tfb - _tfa),
                            "Lectura": "", "_ord": 99, "_dsat": 0, "_fa": 0, "_fb": 0,
                        }

                        _export_cols = [
                            "Nave", "Línea", "Modelo",
                            "Dem. Partida", "Cap. Partida", "Sat. Partida (%)", "Déf. Partida",
                            "Dem. Evaluado", "Cap. Evaluado", "Sat. Evaluado (%)", "Déf. Evaluado",
                            "Cambio cap.", "Cambio sat. (pts)", "Cambio déf.", "Lectura",
                        ]
                        _comp_df_out = pd.DataFrame(_comp_rows + [_total_comp_row])[_export_cols].rename(columns={
                            "Dem. Partida":    "Demanda partida (uds/sem)",
                            "Cap. Partida":    "Capacidad partida (uds/sem)",
                            "Sat. Partida (%)":"Saturación partida (%)",
                            "Déf. Partida":    "Déficit partida (uds/sem)",
                            "Dem. Evaluado":   "Demanda evaluado (uds/sem)",
                            "Cap. Evaluado":   "Capacidad evaluado (uds/sem)",
                            "Sat. Evaluado (%)":"Saturación evaluado (%)",
                            "Déf. Evaluado":   "Déficit evaluado (uds/sem)",
                            "Cambio cap.":     "Cambio capacidad (uds/sem) [+más/-menos]",
                            "Cambio sat. (pts)":"Cambio saturación (pts) [+más presión/-menos]",
                            "Cambio déf.":     "Cambio déficit (uds/sem) [+empeora/-mejora]",
                        })

                        # Hojas PARTIDA y EVALUADO
                        def _fmt_simple_df(df_in):
                            _df = df_in.copy()
                            for _c in _cols_base:
                                if _c in _df.columns:
                                    _df[_c] = _df[_c].apply(
                                        lambda v: "" if (v is None or (isinstance(v, float) and pd.isna(v)))
                                        else _fnes(v)
                                    )
                            return _df

                        _pcols = ["nave", "line", "model"] + _cols_base
                        _partida_out = _fmt_simple_df(
                            _base_lines_df[_pcols].rename(
                                columns={"nave": "Nave", "line": "Línea", "model": "Modelo"}
                            )
                        )
                        _eval_out = pd.DataFrame()
                        if not _cmp_df.empty:
                            _eval_out = _fmt_simple_df(
                                _cmp_df[_pcols].rename(
                                    columns={"nave": "Nave", "line": "Línea", "model": "Modelo"}
                                )
                            )

                        # ── Datos para RESUMEN ────────────────────────────────
                        _cmp_sc_name = _cmp_id_map.get(_cmp_sel, "")
                        _n_tot = len(_comp_rows)
                        _n_emp = sum(1 for r in _comp_rows if r["_ord"] in (0, 1, 4))
                        _n_mej = sum(1 for r in _comp_rows if r["_ord"] in (2, 3, 5))
                        _n_nc = sum(1 for r in _comp_rows if r["_ord"] == 6)
                        _frase_res = (
                            f"Al pasar de '{_sc_name_export}' a '{_cmp_sc_name}': "
                            f"{_n_emp} línea(s) empeoran y {_n_mej} línea(s) mejoran."
                        )
                        _mejora_prio = {3: 0, 2: 1, 5: 2}
                        _emp_rows = sorted(
                            [r for r in _comp_rows if r["_ord"] in (0, 1, 4)],
                            key=lambda r: (r["_ord"], -abs(r["_dsat"]))
                        )
                        _mej_rows = sorted(
                            [r for r in _comp_rows if r["_ord"] in (2, 3, 5)],
                            key=lambda r: (_mejora_prio.get(r["_ord"], 9), -abs(r["_dsat"]))
                        )
                        _nc_rows = [r for r in _comp_rows if r["_ord"] == 6]

                        # ── Construir Excel ───────────────────────────────────
                        _cmp_filename = f"comparativa_{selected_plant_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                        _cmp_buf = io.BytesIO()

                        _BLUE = PatternFill("solid", fgColor="FFBDD7EE")
                        _GREEN = PatternFill("solid", fgColor="FFC6EFCE")
                        _YELL = PatternFill("solid", fgColor="FFFFEB9C")
                        _PURP = PatternFill("solid", fgColor="FFD9D2E9")
                        _GRAY = PatternFill("solid", fgColor="FFD9D9D9")
                        _RED_ROW = PatternFill("solid", fgColor="FFFFC7CE")
                        _GRN_ROW = PatternFill("solid", fgColor="FFE2EFDA")
                        _EMP_SEC = PatternFill("solid", fgColor="FFFFC7CE")
                        _MEJ_SEC = PatternFill("solid", fgColor="FFC6EFCE")
                        _NC_SEC = PatternFill("solid", fgColor="FFFFF2CC")
                        _BOLD14 = Font(bold=True, size=14)
                        _BOLD11 = Font(bold=True, size=11)
                        _BOLD = Font(bold=True)
                        _RIGHT = Alignment(horizontal="right")
                        _CENTER = Alignment(horizontal="center", vertical="center")

                        with pd.ExcelWriter(_cmp_buf, engine="openpyxl") as _cw:

                            # ── Hoja COMPARATIVA ──────────────────────────────
                            _comp_df_out.to_excel(_cw, sheet_name="COMPARATIVA", index=False)
                            _ws_comp = _cw.sheets["COMPARATIVA"]
                            _fills_comp = {}
                            for _ci in range(1, 16):
                                if _ci <= 3:
                                    _fills_comp[_ci] = _GRAY
                                elif _ci <= 7:
                                    _fills_comp[_ci] = _BLUE
                                elif _ci <= 11:
                                    _fills_comp[_ci] = _GREEN
                                elif _ci <= 14:
                                    _fills_comp[_ci] = _YELL
                                else:
                                    _fills_comp[_ci] = _PURP
                            _xl_hdr(_ws_comp, _fills_comp)
                            # Fila TOTAL en negrita
                            for _cell in _ws_comp[_ws_comp.max_row]:
                                _cell.font = _BOLD
                            # Resaltar filas críticas y alinear números
                            for _row_cells in _ws_comp.iter_rows(min_row=2, max_row=_ws_comp.max_row - 1):
                                _lec_val = _row_cells[14].value if len(_row_cells) > 14 else ""
                                if _lec_val == "aparece déficit":
                                    for _c in _row_cells:
                                        _c.fill = _RED_ROW
                                elif _lec_val == "desaparece el déficit":
                                    for _c in _row_cells:
                                        _c.fill = _GRN_ROW
                                for _c in _row_cells[3:14]:  # columnas numéricas
                                    if _c.value not in (None, ""):
                                        _c.alignment = _RIGHT

                            # ── Hoja PARTIDA ──────────────────────────────────
                            _partida_out.to_excel(_cw, sheet_name="PARTIDA", index=False)
                            _xl_hdr(_cw.sheets["PARTIDA"],
                                    {i: _BLUE for i in range(1, len(_partida_out.columns) + 1)})

                            # ── Hoja EVALUADO ─────────────────────────────────
                            if not _eval_out.empty:
                                _eval_out.to_excel(_cw, sheet_name="EVALUADO", index=False)
                                _xl_hdr(_cw.sheets["EVALUADO"],
                                        {i: _GREEN for i in range(1, len(_eval_out.columns) + 1)})

                            # ── Hoja RESUMEN (primera, vía openpyxl directo) ──
                            if "Sheet" in _cw.book.sheetnames:
                                del _cw.book["Sheet"]
                            _ws_res = _cw.book.create_sheet("RESUMEN", 0)

                            def _rc(r, c, val, font=None, fill=None, align=None):
                                _cell = _ws_res.cell(row=r, column=c, value=val)
                                if font:
                                    _cell.font = font
                                if fill:
                                    _cell.fill = fill
                                if align:
                                    _cell.alignment = align
                                return _cell

                            _r = 1
                            # Frase narrativa
                            _rc(_r, 1, _frase_res, font=_BOLD14)
                            _ws_res.merge_cells(start_row=_r, start_column=1, end_row=_r, end_column=6)
                            _r += 2

                            # Bloque de contexto
                            for _lbl, _val in [
                                ("Planta", selected_plant_name),
                                ("Escenario de partida", _sc_name_export),
                                ("Escenario evaluado", _cmp_sc_name),
                                ("Exportado", _export_ts),
                            ]:
                                _rc(_r, 1, _lbl, font=_BOLD, fill=_GRAY)
                                _rc(_r, 2, _val)
                                _r += 1

                            _r += 1
                            # Lectura ejecutiva
                            _rc(_r, 1, "LECTURA EJECUTIVA", font=_BOLD11)
                            _r += 1
                            _max_imp = max(_comp_rows, key=lambda r: abs(r["_dsat"]), default=None)
                            if _max_imp and abs(_max_imp["_dsat"]) >= 1.0:
                                _li = f"{_max_imp['Nave']}-{_max_imp['Línea']}"
                                _dir = "subida" if _max_imp["_dsat"] > 0 else "bajada"
                                _msg1 = f"La mayor {_dir} de saturación se concentra en la línea {_li}."
                            else:
                                _msg1 = "No hay cambios de saturación relevantes en ninguna línea."
                            _n_aparece = sum(1 for r in _comp_rows if r["_ord"] == 0)
                            _n_crece = sum(1 for r in _comp_rows if r["_ord"] == 1)
                            if _n_aparece > 0:
                                _msg2 = f"Aparece déficit nuevo en {_n_aparece} línea(s)."
                            elif _n_crece > 0:
                                _msg2 = f"El déficit crece en {_n_crece} línea(s) ya afectadas."
                            else:
                                _msg2 = "No aparece déficit nuevo."
                            if _n_emp == 0 and _n_mej == 0:
                                _msg3 = "El cambio no tiene impacto relevante en ninguna línea."
                            elif _n_emp > _n_mej:
                                _msg3 = f"El escenario evaluado empeora {_n_emp} línea(s) y mejora {_n_mej}. Balance negativo."
                            elif _n_mej > _n_emp:
                                _msg3 = f"El escenario evaluado mejora {_n_mej} línea(s) y empeora {_n_emp}. Balance positivo."
                            else:
                                _msg3 = f"Empeoran {_n_emp} línea(s) y mejoran {_n_mej}. Balance neutro."
                            for _tag, _msg in [
                                ("Impacto principal", _msg1),
                                ("Riesgo clave", _msg2),
                                ("Lectura global", _msg3),
                            ]:
                                _rc(_r, 1, _tag, font=_BOLD, fill=_GRAY)
                                _rc(_r, 2, _msg)
                                _r += 1

                            _r += 1
                            # Síntesis numérica
                            _rc(_r, 1, "SÍNTESIS DEL CAMBIO", font=_BOLD11)
                            _r += 1
                            for _lbl, _val in [
                                ("Líneas analizadas", str(_n_tot)),
                                ("Líneas que empeoran", str(_n_emp)),
                                ("Líneas que mejoran", str(_n_mej)),
                                ("Sin cambio relevante", str(_n_nc)),
                                ("Cambio en capacidad total (uds/sem)", _fnes_signed(_tcb - _tca)),
                                ("Cambio en déficit total (uds/sem)", _fnes_signed(_tfb - _tfa)),
                            ]:
                                _rc(_r, 1, _lbl, font=_BOLD, fill=_GRAY)
                                _rc(_r, 2, _val)
                                _r += 1

                            # Leyenda de interpretación
                            _r += 1
                            _rc(_r, 1, "CÓMO LEER ESTA COMPARACIÓN", font=_BOLD11)
                            _r += 1
                            for _lbl, _exp in [
                                ("Líneas que empeoran",
                                 "aparece o crece el déficit, o sube la saturación más de 3 puntos"),
                                ("Líneas que mejoran",
                                 "desaparece o baja el déficit, o reduce la saturación más de 3 puntos"),
                                ("Sin cambio relevante",
                                 "la variación es pequeña y no hay déficit nuevo"),
                                ("Cambio en capacidad",
                                 "positivo = más capacidad en el evaluado · negativo = menos capacidad"),
                                ("Cambio en déficit",
                                 "positivo = más déficit (empeora) · negativo = menos déficit (mejora)"),
                                ("Cambio en saturación",
                                 "positivo = más presión sobre la línea · negativo = menos presión"),
                            ]:
                                _rc(_r, 1, _lbl, font=_BOLD, fill=_GRAY)
                                _rc(_r, 2, _exp)
                                _r += 1

                            # Tabla líneas que empeoran
                            if _emp_rows:
                                _r += 1
                                _rc(_r, 1, "LÍNEAS QUE EMPEORAN", font=_BOLD11, fill=_EMP_SEC)
                                _ws_res.merge_cells(start_row=_r, start_column=1,
                                                    end_row=_r, end_column=7)
                                _r += 1
                                for _ci, _h in enumerate(
                                    ["Nave", "Línea", "Sat. partida (%)", "Sat. evaluado (%)",
                                     "Déf. partida", "Déf. evaluado", "Qué ocurre"], start=1
                                ):
                                    _rc(_r, _ci, _h, font=_BOLD, fill=_EMP_SEC)
                                _r += 1
                                for _rw in _emp_rows:
                                    for _ci, _v in enumerate([
                                        _rw["Nave"], _rw["Línea"],
                                        _rw["Sat. Partida (%)"], _rw["Sat. Evaluado (%)"],
                                        _rw["Déf. Partida"], _rw["Déf. Evaluado"],
                                        _rw["Lectura"],
                                    ], start=1):
                                        _rc(_r, _ci, _v)
                                    _r += 1

                            # Tabla líneas que mejoran
                            if _mej_rows:
                                _r += 1
                                _rc(_r, 1, "LÍNEAS QUE MEJORAN", font=_BOLD11, fill=_MEJ_SEC)
                                _ws_res.merge_cells(start_row=_r, start_column=1,
                                                    end_row=_r, end_column=7)
                                _r += 1
                                for _ci, _h in enumerate(
                                    ["Nave", "Línea", "Sat. partida (%)", "Sat. evaluado (%)",
                                     "Déf. partida", "Déf. evaluado", "Qué ocurre"], start=1
                                ):
                                    _rc(_r, _ci, _h, font=_BOLD, fill=_MEJ_SEC)
                                _r += 1
                                for _rw in _mej_rows:
                                    for _ci, _v in enumerate([
                                        _rw["Nave"], _rw["Línea"],
                                        _rw["Sat. Partida (%)"], _rw["Sat. Evaluado (%)"],
                                        _rw["Déf. Partida"], _rw["Déf. Evaluado"],
                                        _rw["Lectura"],
                                    ], start=1):
                                        _rc(_r, _ci, _v)
                                    _r += 1

                            # Lista sin cambio relevante
                            if _nc_rows:
                                _r += 1
                                _rc(_r, 1, "SIN CAMBIO RELEVANTE", font=_BOLD11, fill=_NC_SEC)
                                _r += 1
                                for _rw in _nc_rows:
                                    _rc(_r, 1, _rw["Nave"])
                                    _rc(_r, 2, _rw["Línea"])
                                    _rc(_r, 3, _rw["Modelo"])
                                    _r += 1

                            # Anchos columna RESUMEN
                            _ws_res.column_dimensions["A"].width = 32
                            _ws_res.column_dimensions["B"].width = 28
                            for _cl in ["C", "D", "E", "F", "G"]:
                                _ws_res.column_dimensions[_cl].width = 18

                        _cmp_buf.seek(0)

                        st.download_button(
                            label=t("res_compare_export_btn"),
                            data=_cmp_buf,
                            file_name=_cmp_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )

        # -----------------------------------------------------------------
        # SECCIÓN: Análisis de bancos de prueba (fase 1 informativa)
        # No modifica ningún valor oficial del motor.
        # -----------------------------------------------------------------
        st.divider()
        st.markdown(t("res_bench_header"))
        st.info(t("res_bench_info"))

        _bench_cfg_r = load_table("test_bench_config")
        _bench_map_r = load_table("da_bench_type")

        if "plant_id" in _bench_cfg_r.columns:
            _bench_cfg_r = _bench_cfg_r[
                pd.to_numeric(_bench_cfg_r["plant_id"], errors="coerce") == plant_id
            ].copy()
        else:
            _bench_cfg_r = pd.DataFrame()

        if "plant_id" in _bench_map_r.columns:
            _bench_map_r = _bench_map_r[
                pd.to_numeric(_bench_map_r["plant_id"], errors="coerce") == plant_id
            ].copy()
        else:
            _bench_map_r = pd.DataFrame()

        if _bench_cfg_r.empty or _bench_map_r.empty:
            st.warning(t("res_bench_warning"))
        elif detail_by_line:
            _bench_rows = []
            for _lid, (_nave, _bline, _model, _dem_w, _bot, _mrg) in detail_by_line.items():
                _da_variant = (
                    st.session_state.line_bench_variant
                    .get(plant_id, {})
                    .get(_lid, "")
                )
                _res = compute_bench_analysis(
                    line_id=_lid,
                    da_value=_model,
                    demand_week=_dem_w,
                    bench_cfg=_bench_cfg_r,
                    bench_map=_bench_map_r,
                    times_df_b=times_df,
                    stations_df_b=stations_df,
                    hours_eff=hours_eff,
                    da_variant=_da_variant,
                )
                # Solo mostramos líneas D&A que entran en este análisis
                if _res["bench_type"] == "No aplica":
                    continue
                _variant_label = _res["da_variant"] if _res["da_variant"] else "—"
                _bench_rows.append({
                    "Línea":                                   f"{_nave}-{_bline}",
                    "Valor D&A":                              _res["da_value"],
                    "Variante":                               _variant_label,
                    "Tipo de prueba":                         _res["test_type"],
                    "Banco aplicable":                        _res["bench_type"],
                    "Capacidad máxima por bancos (UDS/SEM)":  _fmt_num(_res["bench_capacity"]) if _res["bench_capacity"] > 0 else "—",
                    "Demanda (UDS/SEM)":                      _fmt_num(_dem_w),
                    "Observación":                            _res["observation"],
                })

            if _bench_rows:
                _bench_df = pd.DataFrame(_bench_rows)
                st.dataframe(_bench_df, use_container_width=True, hide_index=True)
            else:
                st.info(t("res_no_da_lines"))

            # -----------------------------------------------------------------
            # RESUMEN AGREGADO POR TIPO DE BANCO
            # -----------------------------------------------------------------
            st.markdown("---")
            st.markdown(t("res_bench_agg_header"))
            st.markdown(t("res_bench_agg_text"))

            # Recolectar datos numéricos por banco — solo líneas D&A con tiempo_para válido
            _agg_input: dict[str, dict] = {}
            for _lid, (_nave, _bline, _model, _dem_w, _bot, _mrg) in detail_by_line.items():
                _da_variant_agg = (
                    st.session_state.line_bench_variant
                    .get(plant_id, {})
                    .get(_lid, "")
                )
                _r = compute_bench_analysis(
                    line_id=_lid,
                    da_value=_model,
                    demand_week=_dem_w,
                    bench_cfg=_bench_cfg_r,
                    bench_map=_bench_map_r,
                    times_df_b=times_df,
                    stations_df_b=stations_df,
                    hours_eff=hours_eff,
                    da_variant=_da_variant_agg,
                )
                _bt = _r["bench_type"]
                _tp = _r["tiempo_para"]
                if _bt in ("No aplica", "—") or _tp <= 0:
                    continue
                if _bt not in _agg_input:
                    _agg_input[_bt] = {"demanda_total": 0.0, "horas_demandadas": 0.0}
                _agg_input[_bt]["demanda_total"]   += _dem_w
                _agg_input[_bt]["horas_demandadas"] += _dem_w * _tp

            _agg_rows = []
            for _bt, _vals in _agg_input.items():
                _cfg_row = _bench_cfg_r[_bench_cfg_r["bench_type"] == _bt]
                if _cfg_row.empty:
                    continue
                _qty = int(pd.to_numeric(_cfg_row.iloc[0]["quantity"], errors="coerce") or 0)
                _horas_disp = hours_eff * _qty
                _dem_total  = _vals["demanda_total"]
                _horas_dem  = _vals["horas_demandadas"]

                if _horas_dem > 0 and _dem_total > 0:
                    _cap_max = (_horas_disp * _dem_total) / _horas_dem
                else:
                    _cap_max = 0.0

                _sat = (_horas_dem / _horas_disp * 100) if _horas_disp > 0 else 0.0

                if _sat <= 85.0:
                    _estado = "✅ OK"
                elif _sat <= 100.0:
                    _estado = "⚠️ Cerca del límite"
                else:
                    _estado = "🔴 Saturado"

                _agg_rows.append({
                    "Tipo de banco":              _bt,
                    "Horas disponibles/sem":      round(_horas_disp, 2),
                    "Horas demandadas/sem":        round(_horas_dem,  2),
                    "Capacidad máxima (UDS/SEM)":  round(_cap_max,    2),
                    "Demanda total (UDS/SEM)":     round(_dem_total,  2),
                    "Saturación (%)":              round(_sat,        2),
                    "Estado":                      _estado,
                })

            if _agg_rows:
                _agg_df = pd.DataFrame(_agg_rows)

                def _style_estado(row):
                    if "Saturado" in str(row.get("Estado", "")):
                        return ["background-color: #ffe6e6; font-weight: 700;"] * len(row)
                    if "Cerca" in str(row.get("Estado", "")):
                        return ["background-color: #fff8e1;"] * len(row)
                    return [""] * len(row)

                _agg_num_cols = [
                    "Horas disponibles/sem", "Horas demandadas/sem",
                    "Capacidad máxima (UDS/SEM)", "Demanda total (UDS/SEM)", "Saturación (%)",
                ]
                st.dataframe(
                    _agg_df.style
                        .apply(_style_estado, axis=1)
                        .format({c: _fmt_num for c in _agg_num_cols if c in _agg_df.columns}),
                    use_container_width=True,
                    hide_index=True,
                )

                # ---------------------------------------------------------
                # FASE 3 — Simulación estratégica de escenarios de banco
                # Proyección hipotética basada en el resumen agregado.
                # No modifica ningún valor oficial del motor.
                # ---------------------------------------------------------
                st.markdown("---")
                st.markdown(t("res_bench_sim_header"))
                st.warning(t("res_bench_sim_warning"))

                import math as _math

                # ── Bloque A: proyección con bancos adicionales ────────────
                st.markdown(t("res_bench_q1"))

                _sim_a_rows = []
                for _row in _agg_rows:
                    _bt       = _row["Tipo de banco"]
                    _qty      = int(_bench_cfg_r.loc[
                        _bench_cfg_r["bench_type"] == _bt, "quantity"
                    ].iloc[0]) if not _bench_cfg_r[_bench_cfg_r["bench_type"] == _bt].empty else 1
                    _cap_now  = _row["Capacidad máxima (UDS/SEM)"]
                    _dem      = _row["Demanda total (UDS/SEM)"]
                    _def_now  = round(_cap_now - _dem, 2)

                    # Proyección lineal: cap × (qty + N) / qty
                    if _qty > 0:
                        _cap_p1 = round(_cap_now * (_qty + 1) / _qty, 2)
                        _cap_p2 = round(_cap_now * (_qty + 2) / _qty, 2)
                    else:
                        _cap_p1 = _cap_now
                        _cap_p2 = _cap_now

                    _def_p1 = round(_cap_p1 - _dem, 2)
                    _def_p2 = round(_cap_p2 - _dem, 2)

                    # Columna de diagnóstico rápido
                    if _def_now >= 0:
                        _diag = "✅ Cubierto con configuración actual"
                    elif _def_p1 >= 0:
                        _diag = "➕ Se cubriría con +1 banco"
                    elif _def_p2 >= 0:
                        _diag = "➕➕ Se cubriría con +2 bancos"
                    else:
                        _diag = "🔴 Insuficiente incluso con +2 bancos"

                    _sim_a_rows.append({
                        "Tipo de banco":             _bt,
                        "Bancos actuales":            _qty,
                        "Cap. actual (UDS/SEM)":      _cap_now,
                        "Demanda (UDS/SEM)":           _dem,
                        "Déficit actual":              _def_now,
                        "Cap. con +1 banco":          _cap_p1,
                        "Déficit con +1 banco":       _def_p1,
                        "Cap. con +2 bancos":         _cap_p2,
                        "Déficit con +2 bancos":      _def_p2,
                        "Diagnóstico":                _diag,
                    })

                def _style_sim_a(row):
                    def_now = row.get("Déficit actual", 0)
                    try:
                        def_now = float(def_now)
                    except (TypeError, ValueError):
                        def_now = 0
                    if def_now < 0:
                        return ["background-color: #fff3cd;"] * len(row)
                    return [""] * len(row)

                def _fmt_deficit(v):
                    try:
                        f = float(v)
                        base = _fmt_num(abs(f))
                        return f"+{base}" if f >= 0 else f"-{base}"
                    except (TypeError, ValueError):
                        return str(v)

                _sim_a_df = pd.DataFrame(_sim_a_rows)
                _deficit_cols_a = [
                    "Déficit actual", "Déficit con +1 banco", "Déficit con +2 bancos"
                ]
                _cap_cols_a = [
                    "Cap. actual (UDS/SEM)", "Demanda (UDS/SEM)",
                    "Cap. con +1 banco", "Cap. con +2 bancos",
                ]
                _fmt_a = {c: _fmt_num for c in _cap_cols_a if c in _sim_a_df.columns}
                _fmt_a.update({c: _fmt_deficit for c in _deficit_cols_a if c in _sim_a_df.columns})
                st.dataframe(
                    _sim_a_df.style
                        .apply(_style_sim_a, axis=1)
                        .format(_fmt_a),
                    use_container_width=True,
                    hide_index=True,
                )

                # ── Bloque B: bancos mínimos necesarios ───────────────────
                st.markdown(t("res_bench_q2"))

                _sim_b_rows = []
                for _row in _agg_rows:
                    _bt      = _row["Tipo de banco"]
                    _qty     = int(_bench_cfg_r.loc[
                        _bench_cfg_r["bench_type"] == _bt, "quantity"
                    ].iloc[0]) if not _bench_cfg_r[_bench_cfg_r["bench_type"] == _bt].empty else 1
                    _cap_now = _row["Capacidad máxima (UDS/SEM)"]
                    _dem     = _row["Demanda total (UDS/SEM)"]

                    # Capacidad unitaria por banco
                    _cap_por_banco = (_cap_now / _qty) if _qty > 0 else 0.0

                    # Bancos mínimos para cubrir demanda (ceil)
                    if _cap_por_banco > 0 and _dem > 0:
                        _min_banks = _math.ceil(_dem / _cap_por_banco)
                    else:
                        _min_banks = _qty  # sin datos suficientes, no cambiar

                    _extra = max(0, _min_banks - _qty)

                    if _extra == 0:
                        _estado_b = "✅ Sin necesidad adicional"
                    else:
                        _estado_b = f"➕ Faltan {_extra} banco{'s' if _extra > 1 else ''}"

                    _sim_b_rows.append({
                        "Tipo de banco":                         _bt,
                        "Bancos actuales":                       _qty,
                        "Cap. por banco (UDS/SEM)":              round(_cap_por_banco, 2),
                        "Demanda actual (UDS/SEM)":              _dem,
                        "Bancos mínimos para cubrir demanda":    _min_banks,
                        "Bancos adicionales necesarios":         _extra,
                        "Estado":                                _estado_b,
                    })

                def _style_sim_b(row):
                    extra = row.get("Bancos adicionales necesarios", 0)
                    try:
                        extra = int(extra)
                    except (TypeError, ValueError):
                        extra = 0
                    if extra > 0:
                        return ["background-color: #ffe6e6;"] * len(row)
                    return ["background-color: #e8f5e9;"] * len(row)

                _sim_b_df = pd.DataFrame(_sim_b_rows)
                _fmt_b = {
                    c: _fmt_num for c in ["Cap. por banco (UDS/SEM)", "Demanda actual (UDS/SEM)"]
                    if c in _sim_b_df.columns
                }
                st.dataframe(
                    _sim_b_df.style
                        .apply(_style_sim_b, axis=1)
                        .format(_fmt_b),
                    use_container_width=True,
                    hide_index=True,
                )

            else:
                st.info(t("res_no_da_data"))
        else:
            st.info(t("res_no_lines"))

        st.divider()
        st.markdown(t("res_detail_header"))
        st.caption(t("res_detail_caption"))

        for line_id, (nave, base_line, model, demand_week, bottleneck_proc, merged) in detail_by_line.items():
            cap_week = 0.0
            if merged is not None and not merged.empty:
                productive_m = merged[(merged["cycle_time_real"] > 0) & (merged["stations"] > 0)]
                if not productive_m.empty:
                    cap_week = float(productive_m["capacity"].min())

            header = f"{nave}-{base_line} — Modelo: {model} | Capacidad máx: {_fmt_num(cap_week)} uds/sem | Cuello: {bottleneck_proc} | Demanda: {_fmt_num(demand_week)} uds/sem"
            with st.expander(header, expanded=False):
                if merged is None or merged.empty:
                    st.warning(t("res_no_data"))
                else:
                    _detail_cols = ["process", "stations", "operators_per_station"]
                    if "machine_time" in merged.columns:
                        _detail_cols.append("machine_time")
                    if "labor_time" in merged.columns:
                        _detail_cols.append("labor_time")
                    if "labor_per_operator" in merged.columns:
                        _detail_cols.append("labor_per_operator")
                    if "cycle_time_real" in merged.columns:
                        _detail_cols.append("cycle_time_real")
                    _detail_cols.append("capacity")
                    show = merged[[c for c in _detail_cols if c in merged.columns]].copy()
                    show["capacity"] = pd.to_numeric(show["capacity"], errors="coerce").fillna(0.0)

                    def hl_bottleneck(row):
                        if str(row["process"]) == str(bottleneck_proc):
                            return ["background-color: #ffe6e6; font-weight: 700; color: #b00000;"] * len(row)
                        return [""] * len(row)

                    st.dataframe(
                        show.style.format({"capacity": "{:.3f}"}).apply(hl_bottleneck, axis=1),
                        use_container_width=True,
                        hide_index=True
                    )

    st.divider()
    st.markdown(t("res_chart_header"))

    if not summary_df.empty:
        df_plot = summary_df.copy()
        df_plot = df_plot[df_plot["line_id"] != "TOTAL"].copy()

        line_order = sorted(df_plot["line_id"].astype(str).unique().tolist())

        if "TOTAL" in set(summary_df["line_id"].astype(str).tolist()):
            line_order.append("TOTAL")

        df_plot["line_id"] = pd.Categorical(df_plot["line_id"].astype(str), categories=line_order, ordered=True)
        df_plot = df_plot.sort_values("line_id")

        st.markdown("### 🔹 Demanda vs Capacidad – UDS/SEM")
        fig1 = go.Figure()
        fig1.add_bar(x=df_plot["line_id"], y=df_plot["Demanda (UDS/SEM)"], name="Demanda", marker_color="#A6192E")
        fig1.add_bar(x=df_plot["line_id"], y=df_plot["Capacidad (UDS/SEM)"], name="Capacidad", marker_color="green")
        fig1.update_layout(barmode="group")
        st.plotly_chart(fig1, use_container_width=True, key="fig1")

        st.markdown("### 🔴 Demanda vs Capacidad – h/SEM (CRÍTICO)")
        fig2 = go.Figure()
        fig2.add_bar(x=df_plot["line_id"], y=df_plot["Demanda (h/SEM)"], name="Demanda", marker_color="#A6192E")
        fig2.add_bar(x=df_plot["line_id"], y=df_plot["Capacidad (h/SEM)"], name="Capacidad", marker_color="green")
        fig2.update_layout(barmode="group")
        st.plotly_chart(fig2, use_container_width=True, key="fig2")

        st.markdown("### 🔹 Demanda vs Capacidad – UDS/AÑO")
        fig3 = go.Figure()
        fig3.add_bar(x=df_plot["line_id"], y=df_plot["Demanda (UDS/AÑO)"], name="Demanda", marker_color="#A6192E")
        fig3.add_bar(x=df_plot["line_id"], y=df_plot["Capacidad (UDS/AÑO)"], name="Capacidad", marker_color="green")
        fig3.update_layout(barmode="group")
        st.plotly_chart(fig3, use_container_width=True, key="fig3")

        st.markdown("### 🔴 Demanda vs Capacidad – h/AÑO (CRÍTICO)")
        fig4 = go.Figure()
        fig4.add_bar(x=df_plot["line_id"], y=df_plot["Demanda (h/AÑO)"], name="Demanda", marker_color="#A6192E")
        fig4.add_bar(x=df_plot["line_id"], y=df_plot["Capacidad (h/AÑO)"], name="Capacidad", marker_color="green")
        fig4.update_layout(barmode="group")
        st.plotly_chart(fig4, use_container_width=True, key="fig4")

        st.markdown("## 🧠 Visión global planta – h/SEM")
        total_demand_h_sem = float(summary_df.loc[summary_df["line_id"] != "TOTAL", "Demanda (h/SEM)"].sum())
        total_capacity_h_sem = float(summary_df.loc[summary_df["line_id"] != "TOTAL", "Capacidad (h/SEM)"].sum())

        fig_total = go.Figure()
        fig_total.add_bar(x=["TOTAL"], y=[total_demand_h_sem], name="Demanda Total", marker_color="#A6192E")
        fig_total.add_bar(x=["TOTAL"], y=[total_capacity_h_sem], name="Capacidad Total", marker_color="green")
        fig_total.update_layout(barmode="group")
        st.plotly_chart(fig_total, use_container_width=True, key="chart_total")


if st.session_state.active_tab == "🧭 Capacidad según mix":
    st.subheader(t("tab_mix_header"))
    st.info(t("mix_info"))

    _t = times_df.copy()
    _t["cycle_time"] = pd.to_numeric(_t["cycle_time"], errors="coerce").fillna(0.0)
    cycle_by_model = _t.groupby("model")["cycle_time"].sum().to_dict()

    line_stats_rows = []
    capH_line_model = {}
    capU_line_model = {}
    max_h_week_by_line = {}

    # Pre-computar bases UNA VEZ: times_df y stations_df hasheados 1 vez en lugar de N×M
    _all_bases = _precompute_all_bases_for_tab4(
        times_df,
        stations_df,
        tuple(sorted(line_ids_nave)),
        tuple(sorted((k, tuple(v)) for k, v in allowed_by_line.items())),
    )

    for line_id in line_ids_nave:

        parts = line_id.split("-", 1)

        if len(parts) == 2:
            nave, base_line = parts
        else:
            nave = "N1"
            base_line = parts[0]

        models_allowed = allowed_by_line.get(line_id, [])
        if not models_allowed:
            continue

        capU_vals = []
        capH_vals = []
        model_for_capH = []

        for m in models_allowed:
            _base = _all_bases.get((line_id, m))
            if _base is None or _base.empty:
                continue
            _prod = _base[(_base["cycle_time_real"] > 0) & (_base["stations"] > 0)]
            if _prod.empty:
                continue
            _cap_vals = (hours_eff * _prod["stations"]) / _prod["cycle_time_real"]
            cap_week = float(_cap_vals.min()) if not _cap_vals.empty and _cap_vals.min() > 0 else 0.0
            if cap_week <= 0:
                continue

            w_m = float(cycle_by_model.get(m, 0.0))
            cap_h_week = cap_week * w_m

            capH_line_model[(line_id, m)] = float(cap_h_week)
            capU_line_model[(line_id, m)] = float(cap_week)

            capU_vals.append(cap_week)
            capH_vals.append(cap_h_week)
            model_for_capH.append(m)

        if not capU_vals:
            continue

        max_u = float(np.max(capU_vals))
        min_u = float(np.min(capU_vals))
        avg_u = float(np.mean(capU_vals))

        max_h = float(np.max(capH_vals))
        min_h = float(np.min(capH_vals))
        avg_h = float(np.mean(capH_vals))
        max_h_week_by_line[line_id] = max_h

        idx_max_h = int(np.argmax(capH_vals))
        idx_min_h = int(np.argmin(capH_vals))
        model_max_h = model_for_capH[idx_max_h] if capH_vals else ""
        model_min_h = model_for_capH[idx_min_h] if capH_vals else ""

        line_stats_rows.append({
            "nave": nave,
            "line": base_line,
            "line_id": line_id,
            "Modelo Máx (h/SEM)": model_max_h,
            "Modelo Mín (h/SEM)": model_min_h,
            "Max UDS/SEM": max_u,
            "Prom UDS/SEM": avg_u,
            "Min UDS/SEM": min_u,
            "Max UDS/AÑO": max_u * weeks_equiv,
            "Prom UDS/AÑO": avg_u * weeks_equiv,
            "Min UDS/AÑO": min_u * weeks_equiv,
            "Max h/SEM": max_h,
            "Prom h/SEM": avg_h,
            "Min h/SEM": min_h,
            "Max h/AÑO": max_h * weeks_equiv,
            "Prom h/AÑO": avg_h * weeks_equiv,
            "Min h/AÑO": min_h * weeks_equiv,
        })

    if not line_stats_rows:
        st.warning(t("mix_no_combos"))
    else:
        line_stats_df = pd.DataFrame(line_stats_rows)

        plant_max_u_sem = float(line_stats_df["Max UDS/SEM"].sum())
        plant_avg_u_sem = float(line_stats_df["Prom UDS/SEM"].sum())
        plant_min_u_sem = float(line_stats_df["Min UDS/SEM"].sum())

        plant_max_h_sem = float(line_stats_df["Max h/SEM"].sum())
        plant_avg_h_sem = float(line_stats_df["Prom h/SEM"].sum())
        plant_min_h_sem = float(line_stats_df["Min h/SEM"].sum())

        plant_rows = [
            {"Escenario": "Máximo", "UDS/SEM": plant_max_u_sem, "UDS/AÑO": plant_max_u_sem * weeks_equiv, "h/SEM": plant_max_h_sem, "h/AÑO": plant_max_h_sem * weeks_equiv},
            {"Escenario": "Promedio", "UDS/SEM": plant_avg_u_sem, "UDS/AÑO": plant_avg_u_sem * weeks_equiv, "h/SEM": plant_avg_h_sem, "h/AÑO": plant_avg_h_sem * weeks_equiv},
            {"Escenario": "Mínimo", "UDS/SEM": plant_min_u_sem, "UDS/AÑO": plant_min_u_sem * weeks_equiv, "h/SEM": plant_min_h_sem, "h/AÑO": plant_min_h_sem * weeks_equiv},
        ]
        plant_df = pd.DataFrame(plant_rows)

        st.markdown(t("mix_level1"))
        st.dataframe(
            plant_df.style.format({"UDS/SEM": "{:.1f}", "UDS/AÑO": "{:.1f}", "h/SEM": "{:.1f}", "h/AÑO": "{:.1f}"}),
            use_container_width=True,
            hide_index=True
        )

        st.divider()

        st.markdown(t("mix_level2"))
        col_order = [
            "nave", "line",
            "Modelo Máx (h/SEM)", "Modelo Mín (h/SEM)",
            "Max UDS/SEM", "Prom UDS/SEM", "Min UDS/SEM",
            "Max UDS/AÑO", "Prom UDS/AÑO", "Min UDS/AÑO",
            "Max h/SEM", "Prom h/SEM", "Min h/SEM",
            "Max h/AÑO", "Prom h/AÑO", "Min h/AÑO",
        ]
        for c in col_order:
            if c not in line_stats_df.columns:
                line_stats_df[c] = ""

        line_stats_df = line_stats_df[col_order].copy()

        st.dataframe(
            line_stats_df.style.format({
                "Max UDS/SEM": "{:.1f}", "Prom UDS/SEM": "{:.1f}", "Min UDS/SEM": "{:.1f}",
                "Max UDS/AÑO": "{:.1f}", "Prom UDS/AÑO": "{:.1f}", "Min UDS/AÑO": "{:.1f}",
                "Max h/SEM": "{:.1f}", "Prom h/SEM": "{:.1f}", "Min h/SEM": "{:.1f}",
                "Max h/AÑO": "{:.1f}", "Prom h/AÑO": "{:.1f}", "Min h/AÑO": "{:.1f}",
            }),
            use_container_width=True,
            hide_index=True
        )

        st.markdown(t("mix_level3"))
        st.caption(t("mix_sim_caption"))

        H_max_plant = float(sum(max_h_week_by_line.values())) if max_h_week_by_line else 0.0
        st.markdown(f"{t('mix_ceiling_label')} {H_max_plant:.2f} h/sem")

        maxH_by_model = {}
        for m in cycle_by_model.keys():
            total_h = 0.0
            for line_id in line_ids_nave:
                h = float(capH_line_model.get((line_id, m), 0.0) or 0.0)
                if h > 0:
                    total_h += h
            if total_h > 0:
                maxH_by_model[m] = total_h

        valid_models = sorted(maxH_by_model.keys())
        share_max_by_model = {m: (maxH_by_model[m] / H_max_plant) if H_max_plant > 0 else 0.0 for m in valid_models}

        if not valid_models or H_max_plant <= 0:
            st.info(t("mix_no_valid_models"))
        else:
            left, right = st.columns([3.2, 1.4], gap="large")

            with right:
                st.markdown(t("mix_plant_agg"))
                total_selected_pct = 0.0

                palette = px.colors.qualitative.Plotly

                segments = []
                for i, m in enumerate(valid_models):
                    max_pct = float(share_max_by_model.get(m, 0.0) * 100.0)
                    if max_pct <= 0:
                        continue
                    key = f"mix_lvl3_pct_{m}"
                    sel_pct = float(st.session_state.get(key, 0.0) or 0.0)
                    sel_pct = max(0.0, min(sel_pct, max_pct))
                    segments.append((m, sel_pct, palette[i % len(palette)]))
                    total_selected_pct += sel_pct

                exceso = max(0.0, total_selected_pct - 100.0)

                range_max = max(120.0, float(total_selected_pct) * 1.10)

                fig = go.Figure(go.Indicator(
                    mode="gauge+number",
                    value=float(total_selected_pct),
                    number={"suffix": "%", "valueformat": ".2f"},
                    gauge={
                        "axis": {"range": [0, range_max]},
                        "bar": {"color": "#A6192E"},
                        "steps": [
                            {"range": [0, 100], "color": "#E8F5E9"},
                            {"range": [100, range_max], "color": "#FFEBEE"},
                        ],
                        "threshold": {"line": {"color": "black", "width": 4}, "thickness": 0.75, "value": 100},
                    },
                ))
                fig.update_layout(height=360, margin=dict(l=10, r=10, t=10, b=10))

                st.plotly_chart(fig, use_container_width=True, key="chart_velocimetro")
                if segments:
                    fig_stack = go.Figure()
                    for m, pct, color in segments:
                        fig_stack.add_bar(y=["Planta"], x=[pct], name=m, orientation="h", marker=dict(color=color))
                    fig_stack.update_layout(
                        barmode="stack",
                        height=120,
                        margin=dict(l=10, r=10, t=10, b=10),
                        xaxis=dict(range=[0, max(100, total_selected_pct)], title="% ocupación estructural"),
                        yaxis=dict(showticklabels=False),
                        legend=dict(orientation="h"),
                    )
                    st.plotly_chart(fig_stack, use_container_width=True, key="chart_stack")

                total_h_week = H_max_plant * (total_selected_pct / 100.0)
                total_h_year = total_h_week * float(weeks_equiv)
                H_max_plant_year = H_max_plant * float(weeks_equiv)

                st.write(f"{t('mix_max_hours_year')} {H_max_plant_year:.2f}")
                if exceso > 0:
                    st.error(f"{t('mix_excess_label')} {exceso:.2f}%")
                st.write(f"{t('mix_occupancy_label')} {total_selected_pct:.2f}%")
                st.write(f"{t('mix_hours_week_label')} {total_h_week:.2f}")
                st.write(f"{t('mix_hours_year_label')} {total_h_year:.2f}")

            with left:
                st.markdown(t("mix_per_model"))
                grid_cols = 2
                rows = [valid_models[i:i+grid_cols] for i in range(0, len(valid_models), grid_cols)]

                for r_models in rows:
                    cols = st.columns(grid_cols, gap="large")
                    for j in range(grid_cols):
                        if j >= len(r_models):
                            cols[j].empty()
                            continue
                        m = r_models[j]
                        maxH = float(maxH_by_model.get(m, 0.0))
                        max_pct = float(share_max_by_model.get(m, 0.0) * 100.0)

                        with cols[j]:
                            st.markdown(f"#### {m}")
                            key = f"mix_lvl3_pct_{m}"
                            sel_pct = st.slider(
                                t("mix_slider_label"),
                                min_value=0.0,
                                max_value=max_pct,
                                value=float(st.session_state.get(key, 0.0) or 0.0),
                                step=0.01,
                                key=key,
                                format="%.2f",
                                help=t("mix_slider_help"),
                            )

                            sel_h_week = H_max_plant * (sel_pct / 100.0)
                            sel_h_year = sel_h_week * float(weeks_equiv)

                            w_m = float(cycle_by_model.get(m, 0.0) or 0.0)
                            sel_u_week = (sel_h_week / w_m) if w_m > 0 else 0.0
                            sel_u_year = sel_u_week * float(weeks_equiv)

                            # Uso del potenciómetro (% del recorrido del slider)
                            pct_of_slider = (sel_pct / max_pct * 100.0) if max_pct > 0 else 0.0
                            st.caption(f"{t('mix_pot_caption')} **{pct_of_slider:.0f}%**")

                            # Equivalencias dinámicas
                            st.markdown(
                                f"<div style='background:#f8f9fa;padding:8px 12px;border-radius:6px;font-size:13px;margin-bottom:8px;'>"
                                f"<b>{t('mix_equiv_label')}</b><br>"
                                f"Uds/sem: <b>{sel_u_week:.2f}</b> · Uds/año: <b>{sel_u_year:.2f}</b><br>"
                                f"h/sem: <b>{sel_h_week:.2f}</b> · h/año: <b>{sel_h_year:.2f}</b>"
                                f"</div>",
                                unsafe_allow_html=True
                            )

                            max_u_week = (maxH / w_m) if w_m > 0 else 0.0
                            max_u_year = max_u_week * float(weeks_equiv)

                            fig_d = go.Figure(go.Pie(
                                labels=["Seleccionado", "Resto hasta máx"],
                                values=[sel_pct, max(0.0, max_pct - sel_pct)],
                                hole=0.65,
                                sort=False,
                                direction="clockwise",
                                textinfo="none",
                            ))
                            fig_d.update_layout(
                                height=240,
                                margin=dict(l=10, r=10, t=10, b=10),
                                showlegend=False,
                                annotations=[dict(
                                    text=f"{sel_pct:.2f}%<br><span style='font-size:12px'>máx {max_pct:.2f}%</span>",
                                    x=0.5, y=0.5, font=dict(size=16), showarrow=False
                                )]
                            )
                            st.plotly_chart(fig_d, use_container_width=True, key=f"chart_donut_{m}")

                            df_m = pd.DataFrame({
                                "": ["Seleccionado", "Máximo modelo"],
                                "Uds/sem": [sel_u_week, max_u_week],
                                "Uds/año": [sel_u_year, max_u_year],
                                "h/sem": [sel_h_week, maxH],
                                "h/año": [sel_h_year, maxH * float(weeks_equiv)],
                            })
                            df_show = df_m.copy()
                            for c in ["Uds/sem", "Uds/año", "h/sem", "h/año"]:
                                df_show[c] = df_show[c].map(lambda x: f"{float(x):.2f}")
                            st.table(df_show)

# =========================================================
# BLOQUE 11 BIS — SIMULACIÓN ANUAL CAPACIDAD VS DEMANDA
# =========================================================
if st.session_state.active_tab == "📅 Simulación anual":
    st.subheader(t("sim_tab_header"))

    # ── Resolver escenario activo ─────────────────────────────────────────────
    # _raw_sc_id: valor puro de session_state — mismo que usa Resultados para _ov_sc_key.
    # _sim_active_sc_id: para display/guard — puede caer a is_active de DB si session es None.
    _raw_sc_id = st.session_state.get(f"_session_sc_id_{plant_id}")
    _sim_active_sc_id = _raw_sc_id
    _sim_sc_name = st.session_state.get(f"_sc_name_map_{plant_id}", {}).get(_raw_sc_id, "—")
    if _sim_sc_name == "—" and _has_db():
        _sim_sc_list = list_scenarios(plant_id)
        _sim_sc_map  = {s["id"]: s["name"] for s in _sim_sc_list}
        if _sim_active_sc_id is None:
            _active_db = next((s for s in _sim_sc_list if s.get("is_active")), None)
            if _active_db:
                _sim_active_sc_id = _active_db["id"]
        if _sim_active_sc_id is not None:
            _sim_sc_name = _sim_sc_map.get(_sim_active_sc_id, "—")


    # Líneas planificadas: mismo criterio que Resultados
    _sim_all_line_ids = sorted(
        stations_df["line_id"].astype(str).str.strip().unique().tolist()
    )
    _sim_planned_line_ids = [
        lid for lid in _sim_all_line_ids
        if st.session_state.get("line_model", {}).get(plant_id, {}).get(lid)
    ]

    # ── Guards ────────────────────────────────────────────────────────────────
    if _sim_active_sc_id is None:
        st.warning(t("sim_no_scenario"))
    elif not _sim_planned_line_ids:
        st.warning(t("sim_no_lines"))
    else:
        # ── Override resolution — mismas claves que Resultados ────────────────
        _sim_ov_sc_key = _raw_sc_id if _raw_sc_id is not None else 0
        _sim_plant_ov = (
            st.session_state.get("line_params_override", {})
            .get(plant_id, {})
            .get(_sim_ov_sc_key, {})
        )
        _sim_proc_ov = (
            st.session_state.get("proc_shift_override", {})
            .get(plant_id, {})
            .get(_sim_ov_sc_key, {})
        )

        def _sim_resolve_he(lid: str) -> float:
            _ov = _sim_plant_ov.get(lid, {})
            if not _ov.get("enabled", False):
                return hours_eff
            return hours_week * int(_ov.get("shifts", shifts)) * float(_ov.get("availability", availability)) * float(_ov.get("efficiency", efficiency))

        def _sim_resolve_proc_he(lid: str):
            _pd = _sim_proc_ov.get(lid)
            if not _pd:
                return None
            _ov_l = _sim_plant_ov.get(lid, {})
            _a = float(_ov_l.get("availability", availability)) if _ov_l.get("enabled") else availability
            _e = float(_ov_l.get("efficiency", efficiency)) if _ov_l.get("enabled") else efficiency
            return {proc: hours_week * float(sh) * _a * _e for proc, sh in _pd.items()}

        # ── Capacidad base: lógica idéntica al loop de Resultados ─────────────
        _sim_tc = times_df.copy()
        _sim_tc["cycle_time"] = pd.to_numeric(_sim_tc["cycle_time"], errors="coerce").fillna(0.0)
        _sim_ctm = _sim_tc.groupby("model")["cycle_time"].sum().to_dict()

        # ── Función local: capacidad de cualquier escenario leyendo desde BD ──
        def _compute_cap_for_scenario(
            sc_id: int,
            times_df_local: pd.DataFrame,
            stations_df_local: pd.DataFrame,
            ctm: dict,
            hours_week_l: float,
            shifts_l: int,
            availability_l: float,
            efficiency_l: float,
            hours_eff_l: float,
        ) -> float:
            _sc_data = load_scenario_by_id(sc_id)
            if not _sc_data:
                return 0.0
            _sc_ov   = load_scenario_line_overrides(sc_id)
            _sc_proc = load_scenario_process_shifts(sc_id)
            _all_lids = sorted(
                stations_df_local["line_id"].astype(str).str.strip().unique().tolist()
            )
            _planned_l = [_lid for _lid in _all_lids if _sc_data["line_model"].get(_lid)]
            _cap_l = 0.0
            for _lid in _planned_l:
                _mdl = _sc_data["line_model"][_lid]
                _ov  = _sc_ov.get(_lid, {})
                if _ov.get("enabled", False):
                    _he = hours_week_l * int(_ov["shifts"]) * float(_ov["availability"]) * float(_ov["efficiency"])
                else:
                    _he = hours_eff_l
                _phe = None
                if _lid in _sc_proc:
                    _a = float(_ov.get("availability", availability_l)) if _ov.get("enabled") else availability_l
                    _e = float(_ov.get("efficiency",   efficiency_l))   if _ov.get("enabled") else efficiency_l
                    _phe = {_p: hours_week_l * float(_sh) * _a * _e for _p, _sh in _sc_proc[_lid].items()}
                if _phe is not None:
                    _, _, _cap_w = compute_line_detail_v2(_lid, _mdl, times_df_local, stations_df_local, _he, _phe)
                else:
                    _, _, _cap_w = compute_line_detail(_lid, _mdl, times_df_local, stations_df_local, _he)
                _cap_l += _cap_w * ctm.get(_mdl, 0.0)
            return _cap_l

        _sim_cap_h_sem = 0.0
        for _slid in _sim_planned_line_ids:
            _smdl = st.session_state.get("line_model", {}).get(plant_id, {}).get(_slid)
            if not _smdl:
                continue
            _s_lhe = _sim_resolve_he(_slid)
            _s_phe = _sim_resolve_proc_he(_slid)
            if _s_phe is not None:
                _, _, _scap_w = compute_line_detail_v2(_slid, _smdl, times_df, stations_df, _s_lhe, _s_phe)
            else:
                _, _, _scap_w = compute_line_detail(_slid, _smdl, times_df, stations_df, _s_lhe)
            _sim_cap_h_sem += _scap_w * _sim_ctm.get(_smdl, 0.0)

        # ── BLOQUE 1 — Cabecera operativa ────────────────────────────────────
        _hdr1, _hdr2, _hdr3, _hdr4 = st.columns([3, 2, 2, 2])
        with _hdr1:
            st.markdown("**Escenario activo**")
            st.markdown(f"{_sim_sc_name}")
        _hdr2.metric(t("sim_n_lines"),        str(len(_sim_planned_line_ids)))
        _hdr3.metric(t("sim_cap_total_base"), f"{_fmt_num(_sim_cap_h_sem)} h/sem")
        _hdr4.metric(t("sim_cap_per_line"),   f"{_fmt_num(_sim_cap_h_sem / len(_sim_planned_line_ids))} h/sem")
        st.caption(t("sim_cap_base_note"))

        def _build_sim_template_xlsx() -> bytes:
            import io
            _buf = io.BytesIO()
            _semanas = [f"Sem {i}" for i in range(1, 53)]

            _df_dem = pd.DataFrame(
                [["SL"] + [0.0] * 52],
                columns=["Modelo"] + _semanas,
            )
            _df_plan = pd.DataFrame(
                [["SL"] + [0.0] * 52],
                columns=["Modelo"] + _semanas,
            )
            _df_esp = pd.DataFrame(
                [
                    {"semana": 1,  "horas_disponibles": 30.0, "motivo": "Semana reducida vacaciones"},
                    {"semana": 25, "horas_disponibles": 0.0,  "motivo": "Parada verano"},
                ],
            )
            _df_res = pd.DataFrame(
                [
                    ["Previsión ventas (h)"]       + [0.0] * 52,
                    ["Plan maestro prod. (h)"]     + [0.0] * 52,
                    ["Disponibilidad real (h)"]    + [0.0] * 52,
                ],
                columns=["Métrica"] + _semanas,
            )

            with pd.ExcelWriter(_buf, engine="openpyxl") as _writer:
                _df_dem.to_excel(_writer,  sheet_name="DEMANDA_HORAS",    index=False)
                _df_plan.to_excel(_writer, sheet_name="PLAN_HORAS",       index=False)
                _df_esp.to_excel(_writer,  sheet_name="SEMANAS_ESPECIALES", index=False)
                _df_res.to_excel(_writer,  sheet_name="RESUMEN_SEMANAL",  index=False)

            return _buf.getvalue()

        # ── BLOQUE 2 — Carga + plantilla ─────────────────────────────────────

        def _parse_sim_excel(file) -> dict:
            import io
            _result = {
                "errors":     [],
                "warnings":   [],
                "demanda":    None,
                "plan":       None,
                "especiales": None,
                "resumen":    None,
                "n_modelos":  0,
                "n_semanas":  0,
            }
            _semanas_esperadas = [f"Sem {i}" for i in range(1, 53)]
            try:
                _xls = pd.ExcelFile(io.BytesIO(file.read()), engine="openpyxl")
            except Exception as _ex:
                _result["errors"].append(f"No se pudo leer el archivo Excel: {_ex}")
                return _result

            _hojas = _xls.sheet_names

            # ── DEMANDA_HORAS — obligatoria ───────────────────────────────────
            if "DEMANDA_HORAS" not in _hojas:
                _result["errors"].append("Falta la hoja obligatoria DEMANDA_HORAS.")
                return _result

            _df_dem = _xls.parse("DEMANDA_HORAS")
            if "Modelo" not in _df_dem.columns:
                _result["errors"].append("DEMANDA_HORAS no tiene la columna 'Modelo'.")
                return _result

            _sems_presentes = [c for c in _df_dem.columns if c in _semanas_esperadas]
            if not _sems_presentes:
                _result["errors"].append(
                    "DEMANDA_HORAS no contiene ninguna columna 'Sem 1'…'Sem 52'."
                )
                return _result

            _df_dem_datos = _df_dem.dropna(subset=["Modelo"])
            if len(_df_dem_datos) == 0:
                _result["errors"].append("DEMANDA_HORAS está vacía (ninguna fila con Modelo).")
                return _result

            # Convertir semanas a numérico — celdas no numéricas → 0.0 + aviso
            for _sc in _sems_presentes:
                _antes = _df_dem_datos[_sc].copy()
                _df_dem_datos = _df_dem_datos.copy()
                _df_dem_datos[_sc] = pd.to_numeric(_df_dem_datos[_sc], errors="coerce").fillna(0.0)
                if _antes.isna().any() or (_antes.astype(str).str.strip() == "").any():
                    pass  # vacías → 0.0, sin aviso extra
                _no_num = pd.to_numeric(_antes, errors="coerce").isna() & _antes.notna() & (_antes.astype(str).str.strip() != "")
                if _no_num.any():
                    _result["warnings"].append(
                        f"DEMANDA_HORAS: valores no numéricos en columna '{_sc}' tratados como 0."
                    )

            _result["demanda"]   = _df_dem_datos
            _result["n_modelos"] = len(_df_dem_datos)
            _result["n_semanas"] = len(_sems_presentes)

            # ── PLAN_HORAS — opcional ─────────────────────────────────────────
            if "PLAN_HORAS" not in _hojas:
                _result["warnings"].append("No se encontró la hoja PLAN_HORAS (opcional).")
            else:
                _df_plan = _xls.parse("PLAN_HORAS")
                for _sc in [c for c in _df_plan.columns if c in _semanas_esperadas]:
                    _df_plan[_sc] = pd.to_numeric(_df_plan[_sc], errors="coerce").fillna(0.0)
                _result["plan"] = _df_plan

            # ── SEMANAS_ESPECIALES — opcional ─────────────────────────────────
            if "SEMANAS_ESPECIALES" not in _hojas:
                _result["warnings"].append("No se encontró la hoja SEMANAS_ESPECIALES (opcional).")
            else:
                _df_esp = _xls.parse("SEMANAS_ESPECIALES")
                _cols_esp = {"semana", "horas_disponibles", "motivo"}
                _faltan_esp = _cols_esp - set(_df_esp.columns)
                if _faltan_esp:
                    _result["warnings"].append(
                        f"SEMANAS_ESPECIALES existe pero le faltan columnas: {sorted(_faltan_esp)}."
                    )
                _result["especiales"] = _df_esp

            # ── RESUMEN_SEMANAL — opcional ────────────────────────────────────
            if "RESUMEN_SEMANAL" not in _hojas:
                _result["warnings"].append("No se encontró la hoja RESUMEN_SEMANAL (opcional).")
            else:
                _result["resumen"] = _xls.parse("RESUMEN_SEMANAL")

            return _result

        _col_up, _col_tmpl = st.columns([3, 1])
        with _col_up:
            _sim_uploaded = st.file_uploader(
                "Selecciona el archivo Excel de simulación",
                type=["xlsx"],
                key=f"sim_upload_{plant_id}",
            )
        with _col_tmpl:
            st.caption("Usa la plantilla como base.")
            st.download_button(
                label="⬇️ Plantilla Excel",
                data=_build_sim_template_xlsx(),
                file_name="plantilla_simulacion_anual.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        if _sim_uploaded is not None:
            import hashlib
            _file_hash = hashlib.sha256(_sim_uploaded.getvalue()).hexdigest()
            _parsed = _parse_sim_excel(_sim_uploaded)

            # Errores bloqueantes
            if _parsed["errors"]:
                for _err in _parsed["errors"]:
                    st.error(f"Error: {_err}")
                st.session_state.pop(f"_sim_parsed_{plant_id}", None)
                st.session_state.pop(f"_sim_result_{plant_id}", None)
            else:
                # Avisos no bloqueantes
                for _warn in _parsed["warnings"]:
                    st.warning(_warn)

                # Resumen de lectura — validación compacta
                st.success(
                    f"✓ {_parsed['n_modelos']} modelos · {_parsed['n_semanas']} semanas"
                    f" · PLAN {'✓' if _parsed['plan'] is not None else '—'}"
                    f" · SEM. ESPECIALES {'✓' if _parsed['especiales'] is not None else '—'}"
                    f" · RESUMEN {'✓' if _parsed['resumen'] is not None else '—'}"
                )
                st.session_state[f"_sim_parsed_{plant_id}"] = _parsed
                if _file_hash != st.session_state.get(f"_sim_file_hash_{plant_id}"):
                    st.session_state.pop(f"_sim_result_{plant_id}", None)
                st.session_state[f"_sim_file_hash_{plant_id}"] = _file_hash

        # ── Cálculo y resultados de simulación ──────────────────────────────
        _sim_parsed = st.session_state.get(f"_sim_parsed_{plant_id}")
        if _sim_parsed is not None:
            _semanas_esperadas = [f"Sem {i}" for i in range(1, 53)]
            _sems_faltantes = [s for s in _semanas_esperadas
                               if s not in _sim_parsed["demanda"].columns]
            _apto_simular = len(_sems_faltantes) == 0

            if not _apto_simular:
                st.warning(
                    "El archivo no es apto para simular. "
                    f"Semanas faltantes en DEMANDA_HORAS: {', '.join(_sems_faltantes)}"
                )
            else:
                if _sim_parsed["plan"] is None:
                    st.info(
                        "PLAN_HORAS no disponible en el archivo — "
                        "se usará DEMANDA_HORAS como plan de referencia."
                    )

                # ── BLOQUE 2B — Configuración de tramos de capacidad ─────────
                st.markdown("---")
                st.markdown("**Tramos de capacidad semanales**")

                # Scenario map for selectboxes
                _tram_sc_map = st.session_state.get(f"_sc_name_map_{plant_id}", {})
                if not _tram_sc_map and _has_db():
                    _tram_raw_list = list_scenarios(plant_id)
                    _tram_sc_map = {s["id"]: s["name"] for s in _tram_raw_list}
                    st.session_state[f"_sc_name_map_{plant_id}"] = _tram_sc_map
                if not _tram_sc_map and _sim_active_sc_id is not None:
                    _tram_sc_map = {_sim_active_sc_id: _sim_sc_name}
                _tram_sc_ids = list(_tram_sc_map.keys())

                # Initialize tramos list (1 tramo por defecto)
                _tram_key = f"_sim_tramos_{plant_id}"
                if _tram_key not in st.session_state or not st.session_state[_tram_key]:
                    _def_sc = _tram_sc_ids[0] if _tram_sc_ids else None
                    st.session_state[_tram_key] = [
                        {"sem_inicio": 1, "sem_fin": 52, "sc_id": _def_sc}
                    ]
                _tramos = st.session_state[_tram_key]

                # Helpers: read current widget values → rebuild list; clear widget keys
                def _tram_read_widgets(trs, pid):
                    _out = []
                    for _ti, _tr in enumerate(trs):
                        _out.append({
                            "sem_inicio": st.session_state.get(f"_tram_{pid}_{_ti}_ini", _tr["sem_inicio"]),
                            "sem_fin":    st.session_state.get(f"_tram_{pid}_{_ti}_fin", _tr["sem_fin"]),
                            "sc_id":      st.session_state.get(f"_tram_{pid}_{_ti}_sc",  _tr["sc_id"]),
                        })
                    return _out

                def _tram_clear_keys(pid, n_max=12):
                    for _ki in range(n_max):
                        for _ks in ["_ini", "_fin", "_sc"]:
                            st.session_state.pop(f"_tram_{pid}_{_ki}{_ks}", None)

                # Validation: check full 1–52 coverage, no gaps, no overlaps
                def _validate_tramos(trs):
                    if not trs:
                        return ["Debe haber al menos 1 tramo."]
                    _errs = []
                    for _vi, _vt in enumerate(trs):
                        if _vt["sem_inicio"] > _vt["sem_fin"]:
                            _errs.append(
                                f"Tramo {_vi + 1}: inicio ({_vt['sem_inicio']}) "
                                f"mayor que fin ({_vt['sem_fin']})."
                            )
                    if _errs:
                        return _errs
                    _sorted = sorted(trs, key=lambda _x: _x["sem_inicio"])
                    if _sorted[0]["sem_inicio"] != 1:
                        _errs.append(
                            f"Hueco: semanas 1–{_sorted[0]['sem_inicio'] - 1} sin cubrir."
                        )
                    _exp = 1
                    for _vt in _sorted:
                        if _vt["sem_inicio"] > _exp:
                            _errs.append(
                                f"Hueco: semanas {_exp}–{_vt['sem_inicio'] - 1} sin cubrir."
                            )
                        elif _vt["sem_inicio"] < _exp:
                            _errs.append(
                                f"Solapamiento: semanas {_vt['sem_inicio']}–{_exp - 1} "
                                "cubiertas más de una vez."
                            )
                        _exp = _vt["sem_fin"] + 1
                    if _exp != 53:
                        _errs.append(
                            f"Hueco: semanas {_exp - 1 if _exp > 1 else _exp}–52 sin cubrir."
                        )
                    return _errs

                # Column headers
                _th0, _th1, _th2, _th3 = st.columns([1, 1, 4, 1])
                _th0.caption("Sem inicio")
                _th1.caption("Sem fin")
                _th2.caption("Escenario")
                _th3.caption("")

                _del_idx = None
                for _ti, _tr in enumerate(_tramos):
                    _tc0, _tc1, _tc2, _tc3 = st.columns([1, 1, 4, 1])
                    _tr["sem_inicio"] = _tc0.number_input(
                        "ini", min_value=1, max_value=52,
                        value=_tr["sem_inicio"],
                        key=f"_tram_{plant_id}_{_ti}_ini",
                        label_visibility="collapsed",
                    )
                    _tr["sem_fin"] = _tc1.number_input(
                        "fin", min_value=1, max_value=52,
                        value=_tr["sem_fin"],
                        key=f"_tram_{plant_id}_{_ti}_fin",
                        label_visibility="collapsed",
                    )
                    if _tram_sc_ids:
                        _cur_sc = _tr["sc_id"] if _tr["sc_id"] in _tram_sc_ids else _tram_sc_ids[0]
                        _tr["sc_id"] = _tc2.selectbox(
                            "sc",
                            options=_tram_sc_ids,
                            format_func=lambda _sid: _tram_sc_map.get(_sid, str(_sid)),
                            index=_tram_sc_ids.index(_cur_sc),
                            key=f"_tram_{plant_id}_{_ti}_sc",
                            label_visibility="collapsed",
                        )
                    else:
                        _tc2.caption("— sin escenarios disponibles —")
                    if len(_tramos) > 1:
                        if _tc3.button("✕", key=f"_tram_{plant_id}_{_ti}_del", use_container_width=True):
                            _del_idx = _ti

                # Apply delete
                if _del_idx is not None:
                    _rebuilt = _tram_read_widgets(_tramos, plant_id)
                    _rebuilt.pop(_del_idx)
                    _tram_clear_keys(plant_id)
                    st.session_state[_tram_key] = _rebuilt
                    st.rerun()

                # Add button — disabled only at max 10 tramos
                if st.button(
                    "＋ Añadir tramo",
                    key=f"_tram_{plant_id}_add",
                    disabled=len(_tramos) >= 10,
                ):
                    _rebuilt = _tram_read_widgets(_tramos, plant_id)
                    _last_fin = _rebuilt[-1]["sem_fin"] if _rebuilt else 0
                    _new_ini  = min(_last_fin + 1, 52)
                    _def_sc   = _tram_sc_ids[0] if _tram_sc_ids else None
                    _rebuilt.append({"sem_inicio": _new_ini, "sem_fin": 52, "sc_id": _def_sc})
                    _tram_clear_keys(plant_id)
                    st.session_state[_tram_key] = _rebuilt
                    st.rerun()

                # Validation feedback
                _tram_errs = _validate_tramos(_tramos)
                if _tram_errs:
                    for _te in _tram_errs:
                        st.error(_te)
                else:
                    st.success("Cobertura completa: Sem 1–52")

                st.markdown("---")

                if st.button("Calcular simulación anual",
                             key=f"sim_calc_btn_{plant_id}",
                             use_container_width=True):
                    # Mapa semanas especiales: {semana_int: horas_disponibles}
                    _esp_map = {}
                    _df_esp_s = _sim_parsed.get("especiales")
                    if _df_esp_s is not None:
                        if "semana" in _df_esp_s.columns and "horas_disponibles" in _df_esp_s.columns:
                            for _, _erow in _df_esp_s.iterrows():
                                try:
                                    _esp_map[int(_erow["semana"])] = float(_erow["horas_disponibles"])
                                except (ValueError, TypeError):
                                    pass

                    # ── Validar tramos y construir mapa semana → capacidad ────
                    _calc_errs = []
                    _week_cap_map = {}  # {w: float} — vacío cuando todos los tramos usan el escenario activo

                    _re_tram_errs = _validate_tramos(_tramos)
                    if _re_tram_errs:
                        _calc_errs.extend(_re_tram_errs)
                    else:
                        # Problema 1: rechazar sc_id None o vacío antes de cualquier otra validación
                        _sc_ids_needed = {t["sc_id"] for t in _tramos}
                        _sc_ids_invalid = {s for s in _sc_ids_needed if s is None or s == ""}
                        if _sc_ids_invalid:
                            _calc_errs.append("Uno o más tramos no tienen escenario asignado.")
                        elif not _has_db():
                            # Sin BD: solo permitir tramos que usen el escenario activo
                            _sc_ids_foreign = _sc_ids_needed - {_sim_active_sc_id}
                            if _sc_ids_foreign:
                                _calc_errs.append(
                                    "El modo multi-tramo con escenarios distintos "
                                    "requiere conexión a base de datos."
                                )
                            # else: _week_cap_map queda vacío → fallback a _sim_cap_h_sem en el bucle
                        else:
                            # Problema 2: validar TODOS los sc_id (incluido el activo) contra lista fresca
                            _fresh_sc_list = list_scenarios(plant_id)
                            _fresh_sc_ids  = {s["id"] for s in _fresh_sc_list}
                            _sc_ids_missing = _sc_ids_needed - _fresh_sc_ids
                            if _sc_ids_missing:
                                _calc_errs.append(
                                    f"Escenario(s) no encontrado(s) en BD: "
                                    f"{', '.join(str(_m) for _m in _sc_ids_missing)}."
                                )
                            else:
                                # Construir cache sc_id → capacidad (una llamada por sc_id distinto)
                                _fresh_sc_name_map = {s["id"]: s["name"] for s in _fresh_sc_list}
                                _sc_cap_cache = {}
                                for _csc in _sc_ids_needed:
                                    _sc_cap_cache[_csc] = (
                                        _sim_cap_h_sem if _csc == _sim_active_sc_id
                                        else _compute_cap_for_scenario(
                                            _csc,
                                            times_df, stations_df, _sim_ctm,
                                            hours_week, shifts, availability, efficiency, hours_eff,
                                        )
                                    )
                                # Problema 4: avisar si un escenario no activo tiene capacidad 0
                                for _csc, _ccap in _sc_cap_cache.items():
                                    if _csc != _sim_active_sc_id and _ccap == 0.0:
                                        _calc_errs.append(
                                            f"El escenario '{_fresh_sc_name_map.get(_csc, _csc)}' "
                                            "tiene capacidad calculada de 0 h/sem. "
                                            "Verifica que tenga líneas planificadas con modelos asignados."
                                        )
                                if not _calc_errs:
                                    for _tr in _tramos:
                                        for _tw in range(_tr["sem_inicio"], _tr["sem_fin"] + 1):
                                            _week_cap_map[_tw] = _sc_cap_cache[_tr["sc_id"]]

                    if _calc_errs:
                        for _cerr in _calc_errs:
                            st.error(_cerr)
                        st.warning(
                            "El gráfico y la tabla mostrados corresponden al último cálculo válido anterior. "
                            "La configuración actual de tramos no se ha aplicado."
                        )
                    else:
                        _rows_sim = []
                        for _w in range(1, 53):
                            _scol = f"Sem {_w}"
                            _dem_w = float(_sim_parsed["demanda"][_scol].sum())
                            if _sim_parsed["plan"] is not None and _scol in _sim_parsed["plan"].columns:
                                _plan_w = float(_sim_parsed["plan"][_scol].sum())
                            else:
                                _plan_w = _dem_w
                            # Prioridad: SEMANAS_ESPECIALES > tramo > escenario activo base
                            _cap_disp_w = _esp_map.get(_w, _week_cap_map.get(_w, _sim_cap_h_sem))
                            _deficit_w  = max(0.0, _dem_w - _cap_disp_w)
                            _sat_w      = round(_dem_w / _cap_disp_w * 100, 1) if _cap_disp_w > 0 else 0.0
                            if _cap_disp_w == 0:
                                _estado_w = "⚫ Parada"
                            elif _deficit_w > 0:
                                _estado_w = "🔴 Déficit"
                            elif _sat_w >= 90.0:
                                _estado_w = "🟡 Atención"
                            else:
                                _estado_w = "🟢 OK"
                            _rows_sim.append({
                                "Semana":              _w,
                                "Demanda (h)":         round(_dem_w, 1),
                                "Plan (h)":            round(_plan_w, 1),
                                "Cap. disponible (h)": round(_cap_disp_w, 1),
                                "Saturación (%)":      _sat_w,
                                "Déficit (h)":         round(_deficit_w, 1),
                                "Estado":              _estado_w,
                            })

                        _tabla_sim       = pd.DataFrame(_rows_sim)
                        _mask_no_parada  = _tabla_sim["Estado"] != "⚫ Parada"
                        _mask_activa     = _tabla_sim["Cap. disponible (h)"] > 0
                        _sat_series      = _tabla_sim.loc[_mask_activa, "Saturación (%)"]
                        _pico_idx        = _sat_series.idxmax() if not _sat_series.empty else None
                        _semana_pico_val = int(_tabla_sim.loc[_pico_idx, "Semana"]) if _pico_idx is not None else "—"
                        _kpis_sim = {
                            "semanas_deficit":   int((_mask_no_parada & (_tabla_sim["Déficit (h)"] > 0)).sum()),
                            "deficit_acumulado": round(float(_tabla_sim["Déficit (h)"].sum()), 1),
                            "semana_pico":       _semana_pico_val,
                            "sat_max":           round(float(_sat_series.max()), 1) if not _sat_series.empty else 0.0,
                            "sat_media":         round(float(_sat_series.mean()), 1) if not _sat_series.empty else 0.0,
                            "usa_tramos":        len(_tramos) > 1,
                        }
                        _tramos_meta = [
                            {
                                "Sem inicio": _tr["sem_inicio"],
                                "Sem fin":    _tr["sem_fin"],
                                "Escenario":  _tram_sc_map.get(_tr["sc_id"], str(_tr["sc_id"])),
                            }
                            for _tr in _tramos
                        ]
                        st.session_state[f"_sim_result_{plant_id}"] = {
                            "tabla":       _tabla_sim,
                            "kpis":        _kpis_sim,
                            "tramos_meta": _tramos_meta,
                            "calc_ts":     datetime.now().strftime("%Y-%m-%d %H:%M"),
                        }

                _sim_result = st.session_state.get(f"_sim_result_{plant_id}")
                if _sim_result is not None:
                    if "semana_pico" not in _sim_result.get("kpis", {}):
                        st.info("Pulsa 'Calcular simulación anual' para actualizar el resultado con el nuevo formato.")
                    else:
                        _rk = _sim_result["kpis"]
                        st.markdown(f"#### Simulación anual — {_sim_sc_name}")
                        _kr1, _kr2, _kr3, _kr4, _kr5 = st.columns(5)
                        _kr1.metric("Semanas con déficit",
                                    str(_rk["semanas_deficit"]))
                        _kr2.metric("Déficit acumulado (h)",
                                    _fmt_num(_rk["deficit_acumulado"]))
                        _kr3.metric("Semana pico",
                                    f"Sem {_rk['semana_pico']}" if _rk["semana_pico"] != "—" else "—")
                        _kr4.metric("Saturación máxima (%)",
                                    f"{_rk['sat_max']} %")
                        _kr5.metric("Saturación media (%)",
                                    f"{_rk['sat_media']} %")
                        _df_esp_render = _sim_parsed.get("especiales") if _sim_parsed else None
                        _n_esp_red = len(_df_esp_render) if _df_esp_render is not None else 0
                        _cap_note = f" · {_n_esp_red} sem. con capacidad reducida" if _n_esp_red > 0 else ""
                        _usa_tramos_r = _rk.get("usa_tramos", False)
                        if _usa_tramos_r:
                            _cap_lbl = f"Capacidad calculada por tramos{_cap_note}"
                        else:
                            _cap_lbl = f"Capacidad base de planta: {_fmt_num(_sim_cap_h_sem)} h/sem{_cap_note}"
                        st.caption(_cap_lbl)
                        # ── Gráfico principal ─────────────────────────────────────────
                        import plotly.graph_objects as go
                        _df_g = _sim_result["tabla"]
                        _color_map_g = {
                            "🟢 OK":       "#7a9fc4",
                            "🟡 Atención": "#f0a500",
                            "🔴 Déficit":  "#d94f4f",
                            "⚫ Parada":   "#5a5a5a",
                        }
                        _bar_colors_g = [_color_map_g.get(e, "#7a9fc4") for e in _df_g["Estado"]]
                        _custom_g = _df_g[
                            ["Plan (h)", "Cap. disponible (h)", "Saturación (%)", "Déficit (h)", "Estado"]
                        ].values

                        # ── Banda estructural según mix ──────────────────────────────
                        _mix_struct = compute_plant_structural_capacity(
                            plant_id, selected_plant_name,
                            settings_df, models_df, times_df, stations_df, compat_df,
                            hours_week_override=hours_week,
                            shifts_override=shifts,
                            availability_override=availability,
                            efficiency_override=efficiency,
                        )
                        _mix_min_h  = _mix_struct.get("min_h_sem", 0.0)
                        _mix_prom_h = _mix_struct.get("prom_h_sem", 0.0)
                        _mix_max_h  = _mix_struct.get("max_h_sem", 0.0)
                        _mix_weeks  = list(range(1, 53))

                        _fig_sim = go.Figure()

                        if _mix_max_h > 0:
                            _fig_sim.add_trace(go.Scatter(
                                x=_mix_weeks, y=[_mix_min_h] * 52,
                                mode="lines", line=dict(color="rgba(60,120,60,0.6)", width=1),
                                showlegend=False, hoverinfo="skip",
                            ))
                            _fig_sim.add_trace(go.Scatter(
                                x=_mix_weeks, y=[_mix_max_h] * 52,
                                mode="lines", line=dict(color="rgba(60,120,60,0.6)", width=1),
                                fill="tonexty",
                                fillcolor="rgba(100,160,100,0.35)",
                                name="Rango estructural según mix — sin overrides",
                                hoverinfo="skip",
                            ))
                            _fig_sim.add_trace(go.Scatter(
                                x=_mix_weeks, y=[_mix_prom_h] * 52,
                                mode="lines",
                                line=dict(color="rgba(60,120,60,0.8)", width=1.5, dash="dashdot"),
                                name="Promedio estructural según mix",
                                hoverinfo="skip",
                            ))

                        _fig_sim.add_trace(go.Bar(
                            x=_df_g["Semana"],
                            y=_df_g["Demanda (h)"],
                            name="Demanda (h)",
                            marker_color=_bar_colors_g,
                            opacity=0.65,
                            width=0.55,
                            customdata=_custom_g,
                            hovertemplate=(
                                "<b>Sem %{x}</b><br>"
                                "Demanda: %{y} h<br>"
                                "Plan: %{customdata[0]} h<br>"
                                "Cap. disponible: %{customdata[1]} h<br>"
                                "Saturación: %{customdata[2]} %<br>"
                                "Déficit: %{customdata[3]} h<br>"
                                "Estado: %{customdata[4]}"
                                "<extra></extra>"
                            ),
                        ))

                        _fig_sim.add_trace(go.Scatter(
                            x=_df_g["Semana"],
                            y=_df_g["Cap. disponible (h)"],
                            name="Cap. disponible (tramos)" if _rk.get("usa_tramos", False) else "Cap. disponible (escenario activo)",
                            mode="lines",
                            line=dict(color="#3d3d3d", width=2.3, shape="hv"),
                            hoverinfo="skip",
                        ))

                        _df_parada_g = _df_g[_df_g["Estado"] == "⚫ Parada"]
                        if not _df_parada_g.empty:
                            _fig_sim.add_trace(go.Scatter(
                                x=_df_parada_g["Semana"],
                                y=[0] * len(_df_parada_g),
                                name="Parada",
                                mode="markers",
                                marker=dict(symbol="x", color="#000000", size=10),
                                hoverinfo="skip",
                            ))

                        _fig_sim.update_layout(
                            height=420,
                            margin=dict(t=20, b=40, l=0, r=0),
                            legend=dict(orientation="h", yanchor="bottom", y=1.01,
                                        xanchor="right", x=1),
                            xaxis=dict(title="Semana", tickmode="linear", tick0=1, dtick=2, tickangle=-45),
                            yaxis=dict(title="Horas"),
                            plot_bgcolor="rgba(0,0,0,0)",
                            paper_bgcolor="rgba(0,0,0,0)",
                            bargap=0.15,
                        )
                        st.plotly_chart(_fig_sim, use_container_width=True)
                        if _mix_max_h > 0:
                            st.caption(
                                f"Rango estructural según mix con parámetros globales actuales, "
                                f"sin overrides de escenario — "
                                f"Mín: {_fmt_num(_mix_min_h)} h/sem · "
                                f"Prom: {_fmt_num(_mix_prom_h)} h/sem · "
                                f"Máx: {_fmt_num(_mix_max_h)} h/sem."
                            )

                        # ── Resumen de tramos aplicados ──────────────────────────────
                        _tramos_meta_r = _sim_result.get("tramos_meta")
                        if _tramos_meta_r:
                            _calc_ts_r = _sim_result.get("calc_ts", "")
                            _ts_lbl = f" · calculado {_calc_ts_r}" if _calc_ts_r else ""
                            st.markdown(f"**Tramos aplicados**{_ts_lbl}")
                            st.dataframe(
                                pd.DataFrame(_tramos_meta_r),
                                use_container_width=True,
                                hide_index=True,
                            )

                        # ── BLOQUE 7 — Tabla operativa ───────────────────────────────
                        _tabla_critica = _sim_result["tabla"][
                            _sim_result["tabla"]["Estado"].isin(
                                ["🔴 Déficit", "🟡 Atención", "⚫ Parada"]
                            )
                        ].copy()
                        if not _tabla_critica.empty:
                            st.markdown("##### Semanas críticas")
                            st.dataframe(_tabla_critica, use_container_width=True, hide_index=True)
                        else:
                            st.success("Sin semanas críticas — todas las semanas dentro de capacidad.")
                        with st.expander("Ver tabla completa (52 semanas)", expanded=False):
                            st.dataframe(_sim_result["tabla"], use_container_width=True, hide_index=True)

                        # ── Export Excel ──────────────────────────────────────────────
                        def _build_sim_export_xlsx(_tabla, _kpis, _cap_base, _n_esp, _pname, _scname, _tramos_meta_exp=None):
                            import io as _io_exp
                            _buf_exp = _io_exp.BytesIO()
                            _estado_map_exp = {
                                "🟢 OK":       "OK",
                                "🟡 Atención": "Atención",
                                "🔴 Déficit":  "Déficit",
                                "⚫ Parada":   "Parada",
                            }
                            _tabla_exp = _tabla.copy()
                            _tabla_exp["Estado"] = _tabla_exp["Estado"].map(
                                lambda v: _estado_map_exp.get(v, v)
                            )
                            _pico_str = f"Sem {_kpis['semana_pico']}" if _kpis["semana_pico"] != "—" else "—"
                            _resumen_exp = pd.DataFrame([
                                {"Parámetro": "Planta",                    "Valor": _pname},
                                {"Parámetro": "Escenario",                 "Valor": _scname},
                                {"Parámetro": "Fecha de cálculo",          "Valor": datetime.now().strftime("%Y-%m-%d")},
                                {"Parámetro": "Capacidad base (h/sem)",    "Valor": _cap_base},
                                {"Parámetro": "Semanas con cap. especial", "Valor": _n_esp},
                                {"Parámetro": "Semanas con déficit",       "Valor": _kpis["semanas_deficit"]},
                                {"Parámetro": "Déficit acumulado (h)",     "Valor": _kpis["deficit_acumulado"]},
                                {"Parámetro": "Semana pico",               "Valor": _pico_str},
                                {"Parámetro": "Saturación máxima (%)",     "Valor": _kpis["sat_max"]},
                                {"Parámetro": "Saturación media (%)",      "Valor": _kpis["sat_media"]},
                            ])
                            with pd.ExcelWriter(_buf_exp, engine="openpyxl") as _writer_exp:
                                _resumen_exp.to_excel(_writer_exp, sheet_name="RESUMEN", index=False)
                                _tabla_exp.to_excel(_writer_exp, sheet_name="SIMULACIÓN_SEMANAL", index=False)
                                if _tramos_meta_exp:
                                    pd.DataFrame(_tramos_meta_exp).to_excel(
                                        _writer_exp, sheet_name="TRAMOS", index=False
                                    )
                            return _buf_exp.getvalue()

                        _export_fname = datetime.now().strftime("simulacion_anual_%Y-%m-%d_%H-%M.xlsx")
                        st.download_button(
                            label="⬇️ Exportar simulación a Excel",
                            data=_build_sim_export_xlsx(
                                _sim_result["tabla"],
                                _sim_result["kpis"],
                                _sim_cap_h_sem,
                                _n_esp_red,
                                selected_plant_name,
                                _sim_sc_name,
                                _sim_result.get("tramos_meta"),
                            ),
                            file_name=_export_fname,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
